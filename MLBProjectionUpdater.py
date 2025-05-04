"""
MLB Projections Automatic Updater

This script automatically updates MLB projections data from various sources:
- FanGraphs (hitter and pitcher stats)
- DraftKings (salaries) from Fantasy Pros
- Fangraphs Probable pitchers
- Baseball Savant batted ball data

Prerequisites:
- Python 3.8+
- Required packages: pandas, openpyxl, pybaseball, requests, selenium, webdriver_manager
"""

import pandas as pd
import time
import os
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import schedule
import logging
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from pybaseball import batting_stats, pitching_stats, batting_stats_range, pitching_stats_range, team_batting, playerid_lookup, statcast
from io import StringIO
import re
import shutil
import xlwings as xw
import unicodedata
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import Ridge
import math

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("mlb_updater.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("MLB Updater")

# File path configuration
EXCEL_FILE_PATH = "MLBProjections.xlsm"  # Update with absolute path if needed
BACKUP_FOLDER = "backups"

# Safely log potentially problematic characters
def safe_log_info(message):
    """Log information safely, handling any encoding issues"""
    try:
        logger.info(message)
    except UnicodeEncodeError:
        # Replace problematic unicode characters with ASCII
        safe_message = message.encode('ascii', 'replace').decode('ascii')
        logger.info(safe_message)
        logger.info("(Note: Some special characters were replaced due to encoding limitations)")

def find_file(filename, possible_locations=None):
    """
    Search for a file in multiple possible locations
    
    Parameters:
    filename (str): Name of the file to find
    possible_locations (list): List of directories to search in. If None, uses default locations.
    
    Returns:
    str: Full path to the file if found, None otherwise
    """
    if possible_locations is None:
        # Default locations to check, change based on your relative paths
        possible_locations = [
            os.path.dirname(os.path.abspath(__file__)),  # Script directory
            os.path.join(os.path.expanduser("~"), "Desktop"),  # Regular Desktop
            os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop"),  # OneDrive Desktop
            os.path.join(os.path.expanduser("~"), "OneDrive - Personal", "Desktop"),  # OneDrive Personal
            os.path.join(os.path.expanduser("~"), "OneDrive - Business", "Desktop"),  # OneDrive Business
            os.getcwd()  # Current working directory
        ]
    
    # Search through locations
    for location in possible_locations:
        if os.path.exists(location):  # Ensure the directory exists
            file_path = os.path.join(location, filename)
            if os.path.exists(file_path):
                logger.info(f"Found file: {file_path}")
                return file_path
    
    logger.warning(f"File not found: {filename}")
    return None

def initialize_file_paths():
    """
    Initialize global file paths based on where files are actually located
    """
    global EXCEL_FILE_PATH
    global BACKUP_FOLDER
    
    # Try to find the Excel file, insert your excel
    excel_file = find_file("MLBProjections.xlsm")
    if excel_file:
        EXCEL_FILE_PATH = excel_file
        logger.info(f"Using Excel file: {EXCEL_FILE_PATH}")
        
        # Create backup folder relative to the Excel file location
        excel_dir = os.path.dirname(excel_file)
        BACKUP_FOLDER = os.path.join(excel_dir, "backups")
    else:
        logger.error("MLBProjections.xlsm not found in any expected location")
        
        # Use current directory for backup folder if Excel file not found
        BACKUP_FOLDER = os.path.join(os.getcwd(), "backups")
    
    logger.info(f"Using backup folder: {BACKUP_FOLDER}")

# Initialize file paths at the start
initialize_file_paths()

def create_backup():
    """Create a backup of the original file"""
    global BACKUP_FOLDER
    
    if not os.path.exists(BACKUP_FOLDER):
        os.makedirs(BACKUP_FOLDER)
        logger.info(f"Created backup folder: {BACKUP_FOLDER}")
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_FOLDER, f"MLBProjections_backup_{timestamp}.xlsm")
    
    try:
        if os.path.exists(EXCEL_FILE_PATH):
            shutil.copy2(EXCEL_FILE_PATH, backup_path)
            logger.info(f"Created backup at {backup_path}")
            return True
    except Exception as e:
        logger.error(f"Backup failed: {str(e)}")
        return False
    
    return False

def get_latest_backup():
    """
    Returns the path to the latest backup file
    
    Returns:
    str: Path to the latest backup file, or None if no backups exist
    """
    if not os.path.exists(BACKUP_FOLDER):
        return None
    
    backup_files = [os.path.join(BACKUP_FOLDER, f) for f in os.listdir(BACKUP_FOLDER) 
                   if f.startswith("MLBProjections_backup_") and f.endswith(".xlsm")]
    
    if not backup_files:
        return None
    
    # Sort by modification time (newest first)
    backup_files.sort(key=os.path.getmtime, reverse=True)
    return backup_files[0]
    
def get_fangraphs_hitters(period="season"):
    """
    Gets fangraphs data for hitters from pyBaseball's batting_stats

    Parameters:
    period (int): season or multiple seasons that you want stats for
    """
    logger.info(f"Getting FanGraphs hitters data ({period})...")
    
    current_year = datetime.now().year

    try:
        if period == "season":
            df = batting_stats(current_year, current_year, qual=1)  # Set qual=1 for minimum 1 PA
        elif period == "last3years":
            start_year = current_year - 3
            df = batting_stats(start_year, current_year, qual=1)  # Set qual=1
            logger.info(f"Retrieved hitter data from {start_year} to {current_year}")

            # Add a year column if it doesn't exist
            if 'Season' in df.columns:
                df['Year'] = df['Season']
            elif 'year' in df.columns:
                df['Year'] = df['year']
            
            # Count how many seasons each player has played
            seasons_played = df.groupby('Name')['Year'].nunique().reset_index()
            seasons_played.rename(columns={'Year': 'SeasonsPlayed'}, inplace=True)
            
            # Group raw batted ball events for proper percentage calculation
            batted_ball_cols = ['GB', 'FB', 'LD', 'IFFB']
            available_bb_cols = [col for col in batted_ball_cols if col in df.columns]
            
            # Sum batted ball events by player to get all BIP data
            bb_sums = df.groupby('Name')[available_bb_cols].sum().reset_index()
            
            # Sum all non-percentage stats
            count_stats = ['G', 'AB', 'PA', 'H', '1B', '2B', '3B', 'HR', 'R', 'RBI', 
                          'BB', 'IBB', 'SO', 'HBP', 'SF', 'SH', 'GDP', 'SB', 'CS']
            available_count_stats = [col for col in count_stats if col in df.columns]
            
            df_summed = df.groupby('Name')[available_count_stats].sum().reset_index()
            
            # Now calculate percentage stats from raw numbers
            if all(col in bb_sums.columns for col in ['GB', 'FB', 'LD']):
                total_balls = bb_sums['GB'] + bb_sums['FB'] + bb_sums['LD']
                # Add IFFB if available
                if 'IFFB' in bb_sums.columns:
                    # Don't add IFFB to total as they're already included in FB
                    pass
                
                # Avoid division by zero
                total_balls = total_balls.apply(lambda x: max(x, 1))
                
                # Calculate percentages
                bb_sums['GB%'] = bb_sums['GB'] / total_balls
                bb_sums['FB%'] = bb_sums['FB'] / total_balls
                bb_sums['LD%'] = bb_sums['LD'] / total_balls
                
                if 'IFFB' in bb_sums.columns and 'FB' in bb_sums.columns:
                    # Avoid division by zero
                    bb_sums['IFFB%'] = bb_sums.apply(
                        lambda row: row['IFFB'] / max(row['FB'], 1), axis=1
                    )
                
                # Add these calculated percentages to df_summed
                df_summed = pd.merge(df_summed, bb_sums, on='Name', how='left')
                
                # Calculate HR/FB from raw numbers, found some inconsistencies so calculating from raw data
                if 'HR' in df_summed.columns and 'FB' in df_summed.columns:
                    df_summed['HR/FB'] = df_summed.apply(
                        lambda row: row['HR'] / max(row['FB'], 1), axis=1
                    )
                
                logger.info("Calculated batted ball percentages from raw event counts")
            
            # Calculate other rate stats that need weighted averages
            rate_stats = ['AVG', 'OBP', 'SLG', 'OPS', 'ISO', 'BABIP', 'wOBA', 'wRC+', 
                         'BB%', 'K%', 'SwStr%', 'Pull%', 'Cent%', 'Oppo%',
                         'Soft%', 'Med%', 'Hard%', 'GB/FB', 'WPA', 'WPA/LI']
            
            # Remove any stats we calculated directly from raw counts
            for stat in ['GB%', 'FB%', 'LD%', 'IFFB%', 'HR/FB']:
                if stat in rate_stats:
                    rate_stats.remove(stat)
            
            available_rate_stats = [stat for stat in rate_stats if stat in df.columns]
            
            # Weigh by PA
            weight_col = 'PA'
            
            # Create a dataframe for weighted rate stats
            weighted_rates = pd.DataFrame({'Name': df['Name'].unique()})
            
            # Calculate weighted averages for rate stats
            for stat in available_rate_stats:
                # Create temp df
                temp_df = df[['Name', stat, weight_col]].copy()
                # Remove na
                temp_df = temp_df.dropna(subset=[stat, weight_col])
                
                if not temp_df.empty:
                    # Weigh the stat by PA
                    temp_df[f'{stat}_weighted'] = temp_df[stat] * temp_df[weight_col]
                    
                    # Group and calculate weighted average
                    weighted_sum = temp_df.groupby('Name')[f'{stat}_weighted'].sum()
                    weight_sum = temp_df.groupby('Name')[weight_col].sum()
                    
                    # Calculate the weighted average
                    weighted_avg = (weighted_sum / weight_sum).reset_index()
                    weighted_avg.columns = ['Name', stat]
                    
                    # Merge into our weighted rates df
                    weighted_rates = pd.merge(weighted_rates, weighted_avg, on='Name', how='left')
            
            # Merge
            if not weighted_rates.empty:
                # Get all stats we want to merge
                stats_to_merge = [stat for stat in available_rate_stats if stat in weighted_rates.columns]
                
                if stats_to_merge:
                    logger.info(f"Merging rate stats for hitters: {stats_to_merge}")
                    
                    # Select just the columns we want (Name + stats)
                    merge_df = weighted_rates[['Name'] + stats_to_merge].copy()
                    
                    # Do a single merge for all stats at once
                    df_summed = pd.merge(df_summed, merge_df, on='Name', how='left')
                    
                    # Log the merge
                    logger.info(f"After merging rate stats, df_summed has {len(df_summed)} rows")
                
            # Add the seasons played column
            df_summed = pd.merge(df_summed, seasons_played, on='Name', how='left')
            df_summed['SeasonsPlayed'] = df_summed['SeasonsPlayed'].fillna(1)
            
            # Return the processed dataframe
            df = df_summed

        else:
            df = batting_stats(current_year, current_year, qual=1)  # Set qual=1
            logger.info(f"Note: Using full season data instead of {period} data")

        logger.info(f"Successfully retrieved {len(df)} hitter records")

        if 'Name' in df.columns:
            logger.info("Final name standardization within get_fangraphs_hitters")
            # Create a standardization
            df['Name'] = df['Name'].apply(lambda x: unicodedata.normalize('NFC', x) if isinstance(x, str) else x)

        return df

    except Exception as e:
        logger.error(f"Error getting FanGraphs hitters: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return pd.DataFrame()
    
def ip_to_outs(ip):
    """
    Convert IP (innings pitched) to number of outs.
    ip (int): innings pitched

    Returns:
    outs (int): converted innings pitched to total number of outs
    """
    whole_innings = int(ip)
    partial = ip - whole_innings
    if abs(partial - 0.1) < 1e-6:
        outs = whole_innings * 3 + 1
    elif abs(partial - 0.2) < 1e-6:
        outs = whole_innings * 3 + 2
    elif abs(partial) < 1e-6:
        outs = whole_innings * 3
    else:
        raise ValueError(f"Unexpected IP value: {ip}")
    return outs

def outs_to_ip(outs):
    """
    Convert number of outs back to IP format.
    outs (int): total outs for a pitcher

    Returns:
    ip (int): converted outs to innings pitched
    """
    whole_innings = outs // 3
    remaining_outs = outs % 3
    return whole_innings + remaining_outs * 0.1


def get_fangraphs_pitchers(period="season"):
    """
    Gets fangraphs data for pitchers from pyBaseball's pitching_stats

    Parameters:
    period (int): season or multiple seasons that you want stats for
    """
    logger.info(f"Getting FanGraphs pitchers data ({period})...")
    
    current_year = datetime.now().year

    try:
        if period == "season":
            df = pitching_stats(current_year, current_year, qual=1)  # Set qual=1 for minimum 1 IP

        elif period == "last3years":
            start_year = current_year - 3
            df = pitching_stats(start_year, current_year, qual=1)  # Set qual=1
            logger.info(f"Retrieved pitcher data from {start_year} to {current_year}")

            # Add a year column if it doesn't exist
            if 'Season' in df.columns:
                df['Year'] = df['Season']
            elif 'year' in df.columns:
                df['Year'] = df['year']
            
            # Count how many seasons each player has played
            seasons_played = df.groupby('Name')['Year'].nunique().reset_index()
            seasons_played.rename(columns={'Year': 'SeasonsPlayed'}, inplace=True)
            
            # Make sure quality of contact metrics are available
            has_hard_pct = 'Hard%' in df.columns
            has_soft_pct = 'Soft%' in df.columns
            has_med_pct = 'Med%' in df.columns
            
            if has_hard_pct:
                logger.info("Found Hard% in data, will include")
            else:
                logger.warning("Hard% not found in data, will add placeholder")
                
            if has_soft_pct:
                logger.info("Found Soft% in data, will include")
            else:
                logger.warning("Soft% not found in data, will add placeholder")
            
            # First, group batted ball events for proper calculation
            batted_ball_cols = ['GB', 'FB', 'LD', 'IFFB']
            available_bb_cols = [col for col in batted_ball_cols if col in df.columns]
            
            # Sum batted ball events by player
            bb_sums = df.groupby('Name')[available_bb_cols].sum().reset_index()
            
            # Sum all non-percentage stats
            count_stats = ['G', 'GS', 'W', 'L', 'SV', 'IP', 'TBF', 'H', 'R', 'ER', 
                          'HR', 'BB', 'IBB', 'HBP', 'SO', 'BK', 'WP']
            available_count_stats = [col for col in count_stats if col in df.columns]
            
            df_summed = df.groupby('Name')[available_count_stats].sum().reset_index()
            
            # Now calculate percentage stats from batted ball numbers
            if all(col in bb_sums.columns for col in ['GB', 'FB', 'LD']):
                total_balls = bb_sums['GB'] + bb_sums['FB'] + bb_sums['LD']
                # Add IFFB if available
                if 'IFFB' in bb_sums.columns:
                    total_balls += bb_sums['IFFB']
                
                # Avoid division by zero
                total_balls = total_balls.apply(lambda x: max(x, 1))
                
                # Calculate percentages
                bb_sums['GB%'] = bb_sums['GB'] / total_balls
                bb_sums['FB%'] = bb_sums['FB'] / total_balls
                bb_sums['LD%'] = bb_sums['LD'] / total_balls
                
                if 'IFFB' in bb_sums.columns and 'FB' in bb_sums.columns:
                    # Avoid division by zero
                    bb_sums['IFFB%'] = bb_sums.apply(
                        lambda row: row['IFFB'] / max(row['FB'], 1), axis=1
                    )
                
                # Add these calculated percentages to df_summed
                df_summed = pd.merge(df_summed, bb_sums, on='Name', how='left')
                
                # Calculate HR/FB from raw numbers
                if 'HR' in df_summed.columns and 'FB' in df_summed.columns:
                    df_summed['HR/FB'] = df_summed.apply(
                        lambda row: row['HR'] / max(row['FB'], 1), axis=1
                    )
                
                logger.info("Calculated batted ball percentages from raw event counts")
            
            # Calculate other rate stats that require weighted averages
            rate_stats = ['ERA', 'FIP', 'xFIP', 'BABIP', 'LOB%', 'AVG', 'WHIP', 'OBP', 'SLG', 'wOBA']
            
            # If quality of contact metrics exist, add
            if has_hard_pct:
                rate_stats.append('Hard%')
            if has_soft_pct:
                rate_stats.append('Soft%')
            if has_med_pct:
                rate_stats.append('Med%')
                
            available_rate_stats = [col for col in rate_stats if col in df.columns]
            
            # Weight by IP or TBF as fallback
            weight_col = 'IP' if 'IP' in df.columns else 'TBF'
            
            # Create a dataframe for weighted rate stats
            weighted_rates = pd.DataFrame({'Name': df['Name'].unique()})
            
            # Calculate weighted averages for rate stats
            for stat in available_rate_stats:
                # Create temp df with needed columns
                temp_df = df[['Name', stat, weight_col]].copy()
                # Remove any na
                temp_df = temp_df.dropna(subset=[stat, weight_col])
                
                if not temp_df.empty:
                    # Weigh the stat by IP or TBF
                    temp_df[f'{stat}_weighted'] = temp_df[stat] * temp_df[weight_col]
                    
                    # Group and calculate weighted average
                    weighted_sum = temp_df.groupby('Name')[f'{stat}_weighted'].sum()
                    weight_sum = temp_df.groupby('Name')[weight_col].sum()
                    
                    # Calculate the weighted average
                    weighted_avg = (weighted_sum / weight_sum).reset_index()
                    weighted_avg.columns = ['Name', stat]
                    
                    # Merge into our weighted rates df
                    weighted_rates = pd.merge(weighted_rates, weighted_avg, on='Name', how='left')
            
            # Add other stats like K/9, BB/9, etc.
            if all(col in df_summed.columns for col in ['SO', 'IP']):
                df_summed['K/9'] = (df_summed['SO'] * 9) / df_summed['IP']
            
            if all(col in df_summed.columns for col in ['BB', 'IP']):
                df_summed['BB/9'] = (df_summed['BB'] * 9) / df_summed['IP']
            
            if all(col in df_summed.columns for col in ['HR', 'IP']):
                df_summed['HR/9'] = (df_summed['HR'] * 9) / df_summed['IP']
            
            if all(col in df_summed.columns for col in ['SO', 'TBF']):
                df_summed['K%'] = df_summed['SO'] / df_summed['TBF']
            
            if all(col in df_summed.columns for col in ['BB', 'TBF']):
                df_summed['BB%'] = df_summed['BB'] / df_summed['TBF']
            
            # FMerge
            if not weighted_rates.empty:
                # Get all stats we want to merge
                stats_to_merge = [stat for stat in available_rate_stats if stat in weighted_rates.columns]
                
                if stats_to_merge:
                    logger.info(f"Merging rate stats: {stats_to_merge}")
                    
                    # Select just the columns we want (Name + stats)
                    merge_df = weighted_rates[['Name'] + stats_to_merge].copy()
                    
                    # Do a single merge for all stats at once
                    df_summed = pd.merge(df_summed, merge_df, on='Name', how='left')
                    
                    # Log merge results
                    logger.info(f"After merging rate stats, df_summed has {len(df_summed)} rows")
            
            # Add the seasons played column
            df_summed = pd.merge(df_summed, seasons_played, on='Name', how='left')
            df_summed['SeasonsPlayed'] = df_summed['SeasonsPlayed'].fillna(1)
            
            # Add placeholder columns for missing metrics
            if not has_hard_pct:
                df_summed['Hard%'] = 0.35  # Default average value
                logger.info("Added placeholder Hard% column with default value")
                
            if not has_soft_pct:
                df_summed['Soft%'] = 0.15  # Default average value
                logger.info("Added placeholder Soft% column with default value")
                
            if not has_med_pct and 'Med%' not in df_summed.columns:
                # If we have Hard% and Soft%, calculate Med% = 1 - Hard% - Soft%
                if 'Hard%' in df_summed.columns and 'Soft%' in df_summed.columns:
                    df_summed['Med%'] = 1 - df_summed['Hard%'] - df_summed['Soft%']
                    logger.info("Calculated Med% from Hard% and Soft%")
                else:
                    df_summed['Med%'] = 0.5  # Default average value
                    logger.info("Added placeholder Med% column with default value")
            
            # Ensure BABIP is present for all pitchers
            if 'BABIP' not in df_summed.columns:
                df_summed['BABIP'] = 0.300  # League average
                logger.info("Added BABIP column with default values for all pitchers")
            else:
                # Fill any na values
                null_count = df_summed['BABIP'].isnull().sum()
                if null_count > 0:
                    logger.info(f"Filling {null_count} null BABIP values with default")
                    df_summed['BABIP'] = df_summed['BABIP'].fillna(0.300)
            
            # Return the processed dataframe
            df = df_summed

        else:
            df = pitching_stats(current_year, current_year, qual=1)  # Set qual=1
            logger.info(f"Note: Using full season data instead of {period} data")
            
            # Add placeholders if metrics are missing
            if 'Hard%' not in df.columns:
                df['Hard%'] = 0.35
                logger.info("Added placeholder Hard% column with default value")
                
            if 'Soft%' not in df.columns:
                df['Soft%'] = 0.15
                logger.info("Added placeholder Soft% column with default value")
                
            if 'Med%' not in df.columns:
                # If we have Hard% and Soft%, calculate Med% = 1 - Hard% - Soft%
                if 'Hard%' in df.columns and 'Soft%' in df.columns:
                    df['Med%'] = 1 - df['Hard%'] - df['Soft%']
                    logger.info("Calculated Med% from Hard% and Soft%")
                else:
                    df['Med%'] = 0.5
                    logger.info("Added placeholder Med% column with default value")

        # Log the number of records
        logger.info(f"Successfully retrieved {len(df)} pitcher records")
        
        # Ensure percentage columns are in decimal format
        pct_cols = ['GB%', 'FB%', 'LD%', 'IFFB%', 'HR/FB', 'K%', 'BB%', 'LOB%', 'Hard%', 'Soft%', 'Med%']
        for col in pct_cols:
            if col in df.columns:
                # If the column contains values > 1.0, assume it's already in percentage format (0-100)
                if df[col].max() <= 1.0:
                    # It's in decimal format (0-1), which is what we want
                    pass
                else:
                    # Convert from percentage (0-100) to decimal (0-1)
                    df[col] = df[col] / 100
                    logger.info(f"Converted {col} from percentage to decimal format")
            
        return df

    except Exception as e:
        logger.error(f"Error getting FanGraphs pitchers: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return pd.DataFrame()


def get_recent_stats(days=7, player_type="hitters"):
    """
    Get recent stats using pybaseball's date range functionality
    
    Parameters:
    days (int): Number of days to look back
    player_type (str): "hitters" or "pitchers"
    
    Returns:
    pandas DataFrame: Recent stats data
    """
    logger.info(f"Getting last {days} days {player_type} data...")
    
    end_date = datetime.now().strftime("%Y-%m-%d")
    start_date = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    
    try:
        if player_type == "hitters":
            # Remove qual parameters
            df = batting_stats_range(start_date, end_date)
        else:
            # Remove qual parameters
            df = pitching_stats_range(start_date, end_date)
        
        logger.info(f"Successfully retrieved {len(df)} {player_type} records for last {days} days")
        return df
    
    except Exception as e:
        logger.error(f"Error getting last {days} days {player_type} data: {str(e)}")
        return pd.DataFrame()

def calculate_dk_points_hitter(player_row):
    """
    Calculate DraftKings fantasy points for a hitter based on their stats
    
    Parameters:
    player_row (pandas Series): A row of player stats
    
    Returns:
    float: Calculated fantasy points per game
    """
    points = 0
    
    # Always calculate singles from H - 2B - 3B - HR
    h = player_row.get('H', 0) 
    doubles = player_row.get('2B', 0)
    triples = player_row.get('3B', 0)
    hr = player_row.get('HR', 0)
    singles = h - doubles - triples - hr
    
    # Add points for singles
    points += singles * 3
    
    # Doubles
    points += doubles * 5
    
    # Triples
    points += triples * 8
    
    # Home Runs
    points += hr * 10
    
    # RBIs
    rbi = player_row.get('RBI', 0)
    points += rbi * 2
    
    # Runs
    runs = player_row.get('R', 0)
    points += runs * 2
    
    # Walks
    bb = player_row.get('BB', 0)
    points += bb * 2
    
    # Hit By Pitch
    hbp = player_row.get('HBP', 0)
    points += hbp * 2
    
    # Stolen Bases
    sb = player_row.get('SB', 0)
    points += sb * 5
    
    # Get games played
    g = player_row.get('G', 0)
    
    # Return average points per game
    if g > 0:
        return points / g
    else:
        return 0

def calculate_dk_points_pitcher(player_row):
    """
    Calculate DraftKings fantasy points for a pitcher based on their stats
    
    Parameters:
    player_row (pandas Series): A row of player stats
    
    Returns:
    float: Calculated fantasy points per game
    """
    
    points = 0
    
    # Innings Pitched - handle all possible formats
    ip = player_row.get('IP', 0)
    if pd.isna(ip):
        ip = 0
    
    # Convert string IP to float if needed
    if isinstance(ip, str):
        # Handle formats like "179.1" or "179.2" (meaning 179 and 1/3 or 179 and 2/3 innings)
        try:
            ip = float(ip)
        except ValueError:
            # Handle other potential string formats
            ip = 0
    
    # Calculate outs from IP, handling correct baseball notation
    # In baseball, IP like 179.1 means 179 1/3 innings
    whole_innings = int(ip)
    partial = ip - whole_innings
    
    # Each full inning = 3 outs
    outs = whole_innings * 3
    
    # For the fractional part, 0.1 = 1/3 inning (1 out), 0.2 = 2/3 inning (2 outs)
    if abs(partial - 0.1) < 0.05:  # Allow some tolerance for floating point
        outs += 1
    elif abs(partial - 0.2) < 0.05:
        outs += 2
    elif abs(partial - 0.3) < 0.05:  # Handle potential 0.3 which should be interpreted as 0.1
        outs += 1
    elif partial > 0 and partial < 0.05:  # Very small remainder, just round
        pass
    elif partial > 0:  # Any other value is likely an error in notation
        decimals = round(partial * 3) / 3
        outs += round(decimals * 3)
    
    points += outs * 0.75

    # Helper to get stats safely
    def safe_stats(stat):
        val = player_row.get(stat, 0)
        if pd.isna(val) or val == '':
            return 0
        return val

    # Strikeouts
    so = safe_stats('SO')
    points += so * 2

    # Wins
    w = safe_stats('W')
    points += w * 4

    # Earned Runs
    er = safe_stats('ER')
    points -= er * 2

    # Hits Against
    h = safe_stats('H')
    points -= h * 0.6

    # Walks
    bb = safe_stats('BB')
    points -= bb * 0.6

    # Hit Batsmen
    hbp = safe_stats('HBP')
    points -= hbp * 0.6

    # Games Played
    g = safe_stats('G')
    
    if g > 0:
        return points / g
    else:
        return 0


def add_dk_points_to_dataframe(df, player_type="hitters"):
    """
    Adds a DK_Points column to a DF containing player stats
    
    Parameters:
    df (pandas DataFrame): DataFrame with player stats
    player_type (str): "hitters" or "pitchers"
    
    Returns:
    pandas DataFrame: Original DataFrame with DK_Points column added
    """
    if df.empty:
        return df
    
    # Make a copy to avoid modifying the original
    result_df = df.copy()
    
    # Apply the appropriate calculation function
    if player_type == "hitters":
        result_df['DK_Points'] = result_df.apply(calculate_dk_points_hitter, axis=1)
    else:
        result_df['DK_Points'] = result_df.apply(calculate_dk_points_pitcher, axis=1)
    
    # Round to 2 decimal places for readability
    result_df['DK_Points'] = result_df['DK_Points'].round(2)
    
    return result_df

def convert_salary(salary_str):
    """
    Convert a salary value like '$7,400' to 7400

    Parameters:
    salary_str (string): string representation of DK salary

    Returns:
    Integer representation of salary.
    """
    if not salary_str or pd.isna(salary_str):
        return 0
    
    try:
        # Convert to string first to handle different input types
        salary_str = str(salary_str)
        # Remove dollar sign, commas and any whitespace
        cleaned = salary_str.replace('$', '').replace(',', '').strip()
        # Convert to integer
        return int(float(cleaned))
    except (ValueError, TypeError):
        # Return 0 if conversion fails
        return 0

def get_draftkings_salaries_csv():
    """
    Get DraftKings MLB salaries from FantasyPros. Player name format FirstName LastName (TeamCode - Position)

    Returns:
    Salary data as a DF
    """
    logger.info("Getting DraftKings salaries from FantasyPros...")
    
    try:
        # FantasyPros URL for DraftKings MLB salaries
        url = "https://www.fantasypros.com/daily-fantasy/mlb/draftkings-salary-changes.php"
        
        # Setup headless Chrome browser
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--ignore-certificate-errors")
        chrome_options.add_argument("--ignore-ssl-errors=yes")
        chrome_options.add_argument("--allow-running-insecure-content")
        
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.get(url)
        
        # Wait for page to load
        time.sleep(5)
        
        try:
            # Find and parse the salary table
            tables = pd.read_html(StringIO(driver.page_source))
            
            # Find the table with salary information
            df = None
            for table in tables:
                if 'Today' in table.columns:
                    df = table
                    break
            
            if df is None and len(tables) > 0:
                df = tables[0]
                logger.warning("Could not identify the salary table by column name. Using first table found.")
            
            if df is not None:
                # Convert salaries to integers
                df['Salary'] = df['Today'].apply(lambda x: int(str(x).replace('$', '').replace(',', '')))
                
                # Keep the original name with team/position for reference
                df['FullName'] = df['Player']
                
                # Extract just the player name without team/position for VLOOKUPs
                def extract_name_only(full_name):
                    if '(' in full_name:
                        return full_name.split('(')[0].strip()
                    return full_name
                
                # Create a new column with just the player name
                df['Name'] = df['Player'].apply(extract_name_only)
                
                # Extract team code from the parentheses
                def extract_team(full_name):
                    if '(' in full_name and ')' in full_name and ' - ' in full_name:
                        team_part = full_name.split('(')[1].split(' - ')[0].strip()
                        return team_part
                    return ""
                
                # Extract position from the parentheses
                def extract_position(full_name):
                    if '(' in full_name and ')' in full_name and ' - ' in full_name:
                        pos_part = full_name.split(' - ')[1].split(')')[0].strip()
                        return pos_part
                    return ""
                
                df['TeamCode'] = df['Player'].apply(extract_team)
                df['Position'] = df['Player'].apply(extract_position)
                
                # Select and order columns for the output
                columns_to_keep = ['Name', 'TeamCode', 'Position', 'Salary', 'FullName']
                output_df = df[columns_to_keep]
                
                logger.info(f"Successfully processed {len(df)} salary records")
                driver.quit()
                return output_df
                
        except Exception as inner_e:
            logger.warning(f"Error parsing FantasyPros tables: {str(inner_e)}")
            
        # Fallback to dummy data
        logger.warning("Using dummy salary data")
        df = pd.DataFrame({
            "Name": ["Brandon Young", "John Smith", "Mike Johnson"],
            "TeamCode": ["BAL", "NYY", "LAD"],
            "Position": ["P", "OF", "1B"],
            "Salary": [7400, 8500, 9200],
            "FullName": ["Brandon Young (BAL - P)", "John Smith (NYY - OF)", "Mike Johnson (LAD - 1B)"]
        })
        
        driver.quit()
        return df
    
    except Exception as e:
        logger.error(f"Error getting DraftKings salaries: {str(e)}")
        return pd.DataFrame()
    
def standardize_all_player_names(df, reference_df=None, name_column='Name', id_column=None):
    """
    Comprehensive function to standardize player names across all data sources
    Handles all encoding issues, including escaped sequences like \xc3\xad
    
    Parameters:
    df (pandas DataFrame): DataFrame containing names to be standardized
    reference_df (pandas DataFrame): Reference DataFrame with "correct" names (e.g., Salaries)
    name_column (str): Column name containing player names
    id_column (str, optional): Column name containing player IDs for more accurate matching
    
    Returns:
    pandas DataFrame: DataFrame with standardized names
    """
    logger.info(f"Standardizing {len(df)} player names with player ID support")
    
    if df.empty or name_column not in df.columns:
        return df
    
    result_df = df.copy()
    
    # Function to clean a name and remove all accents
    def clean_name(name):
        if pd.isna(name) or not isinstance(name, str):
            return name, ""
            
        # Handle escaped hex sequences like \xc3\xad (í)
        cleaned = name
        
        # Try to handle escaped sequences
        if '\\x' in cleaned:
            try:
                # Convert the string with escape sequences to bytes and decode
                cleaned = bytes(cleaned, 'latin1').decode('unicode_escape').encode('latin1').decode('utf-8')
            except Exception:
                pass
                
        # Handle double-encoded UTF-8 (common in L7/L30 data)
        try:
            if any(c in cleaned for c in ['Ã', 'Â', 'Ä', 'Å']):
                # This handles cases where UTF-8 was decoded as Latin-1
                decoded = cleaned.encode('latin1').decode('utf-8')
                if decoded != cleaned:
                    cleaned = decoded
        except Exception:
            pass
        
        # Extract suffix if it exists
        suffix = ""
        suffixes = ['Jr.', 'Sr.', 'III', 'II', 'IV']
        
        # Check for suffix in the name
        name_parts = cleaned.split()
        if len(name_parts) >= 2 and name_parts[-1] in suffixes:
            suffix = name_parts.pop()
            cleaned = ' '.join(name_parts)
            
        # Remove all accents
        import unicodedata
        normalized = unicodedata.normalize('NFKD', cleaned)
        cleaned = ''.join([c for c in normalized if not unicodedata.combining(c)])
            
        return cleaned, suffix
    
    # Function to standardize a name with ID support
    def standardize_name(row):
        if isinstance(row, dict):
            name = row.get(name_column)
            player_id = row.get(id_column) if id_column else None
        else:
            name = row
            player_id = None
        
        # If valid ID and it exists in ID mappings, use that
        if player_id is not None and pd.notna(player_id) and str(player_id) in id_mappings:
            return id_mappings[str(player_id)]
            
        # Otherwise fall back to name-based standardization
        if pd.isna(name) or not isinstance(name, str):
            return name
            
        # First clean the name and extract suffix
        cleaned, original_suffix = clean_name(name)
        
        # Convert to lowercase for matching purposes
        cleaned_lower = cleaned.lower()
        
        # Check if we have a match in reference names
        if cleaned_lower in name_mappings:
            # Get the reference name - this is the name format from Salaries
            ref_name = name_mappings[cleaned_lower]
            
            # We prioritize the Salaries format
            return ref_name  # Return the reference name as is
        
        # If no match in reference data, return the original name (with suffix if it had one)
        if original_suffix:
            return f"{cleaned} {original_suffix}"
        else:
            return cleaned
    
    # Create mappings from reference data if available
    name_mappings = {}
    id_mappings = {}
    
    if reference_df is not None and not reference_df.empty and name_column in reference_df.columns:
        logger.info(f"Using reference data with {len(reference_df)} rows to guide standardization")
        
        # Check for player ID in reference data
        ref_id_column = None
        for col in ['PlayerID', 'player_id', 'mlbam_id', 'id']:
            if col in reference_df.columns:
                ref_id_column = col
                break
        
        # Create ID-based mapping
        if ref_id_column and id_column and id_column in df.columns:
            logger.info(f"Using player IDs for name standardization")
            for _, row in reference_df.iterrows():
                if pd.notna(row.get(ref_id_column)) and pd.notna(row.get(name_column)):
                    id_mappings[str(row[ref_id_column])] = row[name_column]
        
        # Create name-based mapping as fallback
        for _, row in reference_df.iterrows():
            ref_name = row[name_column]
            if pd.isna(ref_name) or not isinstance(ref_name, str):
                continue
                
            # Clean the reference name (without suffix initially)
            clean_ref, _ = clean_name(ref_name)
            
            # Use the reference names as our standard format
            name_mappings[clean_ref.lower()] = ref_name
    
    # Apply standardization with ID if available
    if id_column and id_column in result_df.columns:
        # Row by row, check both name and ID
        for idx, row in result_df.iterrows():
            row_dict = row.to_dict()
            result_df.at[idx, name_column] = standardize_name(row_dict)
    else:
        # Apply the original name-only standardization
        result_df[name_column] = result_df[name_column].apply(standardize_name)
    
    # Count and log changes
    if name_column in df.columns:
        changes = sum(df[name_column] != result_df[name_column])
        logger.info(f"Standardized {changes} player names")
        
        # Show changes in log
        if changes > 0:
            examples = []
            for idx, (old, new) in enumerate(zip(df[name_column], result_df[name_column])):
                if old != new and len(examples) < 10:
                    examples.append((old, new))
            
            if examples:
                logger.info("Name standardization examples:")
                for old, new in examples:
                    logger.info(f"  '{old}' -> '{new}'")
    
    return result_df

def process_splits_data(splits_dict, reference_df=None):
    """
    Process and standardize handedness splits data
    
    Parameters:
    splits_dict (dict): Dictionary with splits DataFrames
    reference_df (pandas DataFrame): Reference DataFrame for name standardization
    
    Returns:
    dict: Dictionary with processed DataFrames
    """
    logger.info("Processing handedness splits data")
    
    processed = {}
    
    for split_type, df in splits_dict.items():
        if df.empty:
            processed[split_type] = df
            continue
            
        logger.info(f"Processing {split_type} data ({len(df)} rows)")
        
        # Make a copy to avoid overwriting original
        processed_df = df.copy()
        
        # Check for player ID column
        player_id_col = None
        for col in processed_df.columns:
            if col.lower() in ['playerid', 'mlbamid', 'id']:
                player_id_col = col
                break
        
        # Extract team information from the Team column
        if 'Team' in processed_df.columns:
            # FanGraphs sometimes has the team in a format like "Team Totals (LAD)"
            def extract_team_code(team_text):
                if pd.isna(team_text) or not isinstance(team_text, str):
                    return ""
                
                # Try to extract team code from parentheses
                if '(' in team_text and ')' in team_text:
                    team_code = team_text.split('(')[-1].split(')')[0].strip()
                    if len(team_code) <= 3 and team_code.isalpha():
                        return team_code
                
                # If no team code in parentheses, just return the original
                return team_text
            
            processed_df['TeamCode'] = processed_df['Team'].apply(extract_team_code)
        
        # Standardize player names if reference data is available
        if reference_df is not None and 'Name' in processed_df.columns:
            processed_df = standardize_all_player_names(
                processed_df, 
                reference_df, 
                'Name', 
                player_id_col
            )
            logger.info(f"Standardized player names in {split_type} data")
        
        # Convert percentage columns to decimal format for consistency
        for col in processed_df.columns:
            # Look for percentage columns that might be in percent format (0-100)
            if any(suffix in col for suffix in ['Pct', 'PCT', '%']) or any(col.endswith(rate) for rate in ['rate', 'Rate']):
                # Check if values are already in decimal format (0-1)
                if processed_df[col].max() <= 1.0:
                    # Already in decimal format
                    pass
                else:
                    # Convert from percentage (0-100) to decimal (0-1)
                    processed_df[col] = processed_df[col] / 100
                    logger.info(f"Converted {col} from percentage to decimal format")
        
        # Calculate fantasy points
        if split_type.startswith("Batter"):
            # First check if we have the necessary columns for DK points
            required_cols = ['H', '2B', '3B', 'HR', 'RBI', 'R', 'BB', 'HBP', 'SB', 'G']
            
            # Check which columns are missing and available
            missing_cols = [col for col in required_cols if col not in processed_df.columns 
                           and f"vsL_{col}" not in processed_df.columns 
                           and f"vsR_{col}" not in processed_df.columns]
            
            if missing_cols:
                logger.warning(f"Missing columns for DK points calculation: {missing_cols}")
                # Cbeck if column name is not standard but means the same thing
                col_mapping = {
                    'H': ['H', 'Hits', 'hits'],
                    '2B': ['2B', 'Doubles', 'doubles'],
                    '3B': ['3B', 'Triples', 'triples'],
                    'HR': ['HR', 'HomeRuns', 'Home Runs', 'homers'],
                    'RBI': ['RBI', 'Runs Batted In', 'runsBattedIn'],
                    'R': ['R', 'Runs', 'runs'],
                    'BB': ['BB', 'Walks', 'Base on Balls', 'walks'],
                    'HBP': ['HBP', 'Hit By Pitch', 'hitByPitch'],
                    'SB': ['SB', 'Stolen Bases', 'stolenBases'],
                    'G': ['G', 'Games', 'GP', 'gamesPlayed']
                }
                
                # Look for columns in both vsL_ and vsR_ prefixes
                for prefix in ['vsL_', 'vsR_', '']:
                    # Try to map columns
                    for missing_col in missing_cols:
                        for alt_col in col_mapping.get(missing_col, []):
                            prefixed_alt = f"{prefix}{alt_col}"
                            if prefixed_alt in processed_df.columns:
                                if prefix:
                                    # Found a prefixed alternative
                                    processed_df[f"{prefix}{missing_col}"] = processed_df[prefixed_alt]
                                    logger.info(f"Mapped {prefixed_alt} to {prefix}{missing_col}")
                                else:
                                    # Found an unprefixed alternative
                                    processed_df[missing_col] = processed_df[prefixed_alt]
                                    logger.info(f"Mapped {prefixed_alt} to {missing_col}")
                                break
            
            # Calculate DK points for both vsL and vsR splits
            for prefix in ['vsL_', 'vsR_']:
                prefix_cols = [col for col in processed_df.columns if col.startswith(prefix)]
                
                if prefix_cols:
                    # Create a temporary DataFrame with renamed columns for the DK points calculation
                    temp_df = processed_df.copy()
                    
                    # Rename columns to remove prefix for calculation
                    for col in prefix_cols:
                        unprefixed = col[len(prefix):]
                        temp_df[unprefixed] = temp_df[col]
                    
                    # Now calculate DK points using the standard function
                    try:
                        temp_df = add_dk_points_to_dataframe(temp_df, "hitters")
                        
                        # Add the result back to the original DataFrame with the prefix
                        processed_df[f"{prefix}DK_Points"] = temp_df['DK_Points']
                        logger.info(f"Calculated DK points for {prefix} splits")
                    except Exception as dk_e:
                        logger.warning(f"Error calculating {prefix}DK_Points: {str(dk_e)}")
        
        elif split_type.startswith("Pitcher"):
            # Check if we have the necessary columns for DK points
            required_cols = ['IP', 'ER', 'H', 'BB', 'SO', 'W', 'G']
            
            # Check which required columns are missing and available
            missing_cols = [col for col in required_cols if col not in processed_df.columns 
                           and f"vsL_{col}" not in processed_df.columns 
                           and f"vsR_{col}" not in processed_df.columns]
            
            if missing_cols:
                logger.warning(f"Missing columns for pitcher DK points calculation: {missing_cols}")
                # Try to find alternative column names
                col_mapping = {
                    'IP': ['IP', 'InningsPitched', 'Innings'],
                    'ER': ['ER', 'EarnedRuns', 'earnedRuns'],
                    'H': ['H', 'Hits', 'hits', 'HitsAllowed'],
                    'BB': ['BB', 'Walks', 'BasesOnBalls', 'walks'],
                    'SO': ['SO', 'K', 'Strikeouts', 'strikeouts'],
                    'W': ['W', 'Wins', 'wins'],
                    'G': ['G', 'Games', 'appearances']
                }
                
                # Look for columns in both vsL_ and vsR_ prefixes
                for prefix in ['vsL_', 'vsR_', '']:
                    # Try to map columns
                    for missing_col in missing_cols:
                        for alt_col in col_mapping.get(missing_col, []):
                            prefixed_alt = f"{prefix}{alt_col}"
                            if prefixed_alt in processed_df.columns:
                                if prefix:
                                    # Found a prefixed alternative
                                    processed_df[f"{prefix}{missing_col}"] = processed_df[prefixed_alt]
                                    logger.info(f"Mapped {prefixed_alt} to {prefix}{missing_col}")
                                else:
                                    # Found an unprefixed alternative
                                    processed_df[missing_col] = processed_df[prefixed_alt]
                                    logger.info(f"Mapped {prefixed_alt} to {missing_col}")
                                break
            
            # Calculate DK points for both vsL and vsR splits
            for prefix in ['vsL_', 'vsR_']:
                prefix_cols = [col for col in processed_df.columns if col.startswith(prefix)]
                
                if prefix_cols:
                    # Create a temporary DataFrame with renamed columns for the DK points calculation
                    temp_df = processed_df.copy()
                    
                    # Rename columns to remove prefix for calculation
                    for col in prefix_cols:
                        unprefixed = col[len(prefix):]
                        temp_df[unprefixed] = temp_df[col]
                    
                    # Now calculate DK points using the standard function
                    try:
                        temp_df = add_dk_points_to_dataframe(temp_df, "pitchers")
                        
                        # Add the result back to the original DataFrame with the prefix
                        processed_df[f"{prefix}DK_Points"] = temp_df['DK_Points']
                        logger.info(f"Calculated DK points for {prefix} splits")
                    except Exception as dk_e:
                        logger.warning(f"Error calculating {prefix}DK_Points: {str(dk_e)}")
        
        # Add timestamp for when this data was updated
        processed_df['LastUpdated'] = datetime.now().strftime("%Y-%m-%d")
        
        # Add some metadata about the data source and date range
        processed_df['DataSource'] = 'FanGraphs'
        processed_df['YearRange'] = '2022-2025'
        processed_df['SplitType'] = split_type
        
        # Calculate sample size metrics for reference
        if split_type.startswith("Batter"):
            # Use PA or AB for batter sample size
            for prefix in ['vsL_', 'vsR_']:
                pa_col = f"{prefix}PA"
                ab_col = f"{prefix}AB"
                
                if pa_col in processed_df.columns:
                    processed_df[f"{prefix}SampleSize"] = processed_df[pa_col]
                elif ab_col in processed_df.columns:
                    processed_df[f"{prefix}SampleSize"] = processed_df[ab_col]
        else:
            # Use TBF or IP for pitcher sample size
            for prefix in ['vsL_', 'vsR_']:
                tbf_col = f"{prefix}TBF"
                ip_col = f"{prefix}IP"
                
                if tbf_col in processed_df.columns:
                    processed_df[f"{prefix}SampleSize"] = processed_df[tbf_col]
                elif ip_col in processed_df.columns:
                    # IP * 3 is a rough estimate of batters faced
                    processed_df[f"{prefix}SampleSize"] = processed_df[ip_col] * 3
        
        # Log some stats about the processed data
        logger.info(f"Processed {len(processed_df)} rows for {split_type}")
        logger.info(f"Final columns: {processed_df.columns.tolist()}")
        
        processed[split_type] = processed_df
    
    return processed

def update_excel_data_only(data_df, sheet_name, old_sheet_name=None):
    """
    Safely updates only the data in the specified sheet without deleting macros.
    If the sheet doesn't exist but old_sheet_name is provided, renames that sheet.
    If neither exists, creates a new sheet.
    
    Parameters:
    data_df (pandas DataFrame): New data to insert
    sheet_name (str): Name of the sheet to update
    old_sheet_name (str, optional): Old name of the sheet to rename if sheet_name doesn't exist
    """
    logger.info(f"Safely updating data in sheet: {sheet_name}")

    try:
        # Check if file exists
        if not os.path.exists(EXCEL_FILE_PATH):
            logger.error(f"Excel file not found: {EXCEL_FILE_PATH}")
            return False

        # Load workbook with VBA support
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH, keep_vba=True)

        # Check if the sheet exists
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            
            # If old_sheet_name is provided and exists, rename it
            if old_sheet_name and old_sheet_name in wb.sheetnames:
                logger.info(f"Renaming sheet '{old_sheet_name}' to '{sheet_name}'")
                ws = wb[old_sheet_name]
                ws.title = sheet_name
            else:
                # Create a new sheet with the desired name
                logger.info(f"Creating new sheet '{sheet_name}'")
                ws = wb.create_sheet(sheet_name)
        else:
            # Get the existing worksheet
            ws = wb[sheet_name]

        # Clear existing sheet content (except headers if needed)
        logger.info(f"Clearing existing data in '{sheet_name}'")
        if ws.max_row > 1:  # Only if there's data to clear
            ws.delete_rows(2, ws.max_row)  # Delete all rows *after* the header

        # Write DataFrame rows into the sheet starting from row 2
        logger.info(f"Writing {len(data_df)} new rows into '{sheet_name}'")

        # Write header manually
        for col_idx, col_name in enumerate(data_df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write each data row
        for row_idx, row in enumerate(data_df.itertuples(index=False, name=None), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Save workbook safely
        wb.save(EXCEL_FILE_PATH)
        logger.info(f"Successfully updated '{sheet_name}' and saved {EXCEL_FILE_PATH}")

        return True

    except Exception as e:
        logger.error(f"Error updating Excel sheet '{sheet_name}': {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False

    
def update_player_data():
    """
    Update player data in Hitter and Pitcher sheets based on Salaries data.
    Only updates player name, team, and salary columns to preserve VLOOKUPs.
    Enhanced to maintain all rows in the sheet, including dragged-down formulas.
    """
    logger.info("Updating player data from Salaries sheet while preserving VLOOKUPs...")
    
    try:
        # Try to read from the file directly
        salaries_df = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Salaries')
        
        if salaries_df.empty:
            logger.error("Salaries sheet is empty")
            return False
        
        # Make sure required columns exist
        required_cols = ['Name', 'Position', 'TeamCode', 'Salary']
        missing_cols = [col for col in required_cols if col not in salaries_df.columns]
        
        if missing_cols:
            logger.error(f"Missing columns in Salaries sheet: {missing_cols}")
            return False
        
        # Separate pitchers and hitters based on position
        pitchers_df = salaries_df[salaries_df['Position'] == 'P']
        hitters_df = salaries_df[salaries_df['Position'] != 'P']
        
        logger.info(f"Found {len(pitchers_df)} pitchers and {len(hitters_df)} hitters in Salaries sheet")
        
        # Now load the workbook with openpyxl for updating, pandas struggles with preserving formulas
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH, keep_vba=True)
        
        # Update Hitter sheet
        if 'Hitter' in wb.sheetnames and not hitters_df.empty:
            hitter_sheet = wb['Hitter']
            
            # Store all existing data and formulas first
            original_data = {}
            for row in range(2, hitter_sheet.max_row + 1):
                # Get the existing row data
                row_data = {}
                for col in range(1, hitter_sheet.max_column + 1):
                    cell = hitter_sheet.cell(row=row, column=col)
                    # Properly check if cell contains a formula
                    is_formula = False
                    formula_value = None
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        is_formula = True
                        formula_value = cell.value
                        
                    row_data[col] = {
                        'value': cell.value,
                        'is_formula': is_formula,
                        'formula_value': formula_value
                    }
                
                # Use the row number as a key to preserve all rows
                original_data[row] = row_data
            
            # Clear only specific columns in these rows (preserve others for VLOOKUPs)
            columns_to_update = [1, 2, 7, 11]  # Position, Name, Salary, Team
            
            for row in range(2, hitter_sheet.max_row + 1):
                for col in columns_to_update:
                    hitter_sheet.cell(row=row, column=col).value = None
            
            # Add new data from hitters_df
            logger.info(f"Adding {len(hitters_df)} rows to Hitter sheet")
            
            for idx, (_, row) in enumerate(hitters_df.iterrows(), start=2):
                # Position in column A (1)
                hitter_sheet.cell(row=idx, column=1).value = row['Position']
                # Name in column B (2)
                hitter_sheet.cell(row=idx, column=2).value = row['Name']
                # Salary in column G (7) 
                hitter_sheet.cell(row=idx, column=7).value = row['Salary']
                # Team in column K (11)
                hitter_sheet.cell(row=idx, column=11).value = row['TeamCode']
            
            # Restore formulas for rows that might have been disrupted
            for row in range(2, min(len(hitters_df) + 2, hitter_sheet.max_row + 1)):
                if row in original_data:
                    # For columns we're NOT updating, preserve any formulas
                    for col in range(1, hitter_sheet.max_column + 1):
                        if col not in columns_to_update and original_data[row][col]['is_formula']:
                            formula = original_data[row][col]['formula_value']
                            hitter_sheet.cell(row=row, column=col).value = formula
            
            # Restore formulas for all rows beyond our data
            for row in range(len(hitters_df) + 2, hitter_sheet.max_row + 1):
                if row in original_data:
                    for col in range(1, hitter_sheet.max_column + 1):
                        if col not in columns_to_update and original_data[row][col]['is_formula']:
                            hitter_sheet.cell(row=row, column=col).value = original_data[row][col]['formula_value']
            
            logger.info(f"Preserved formulas in {hitter_sheet.max_row - (len(hitters_df) + 1)} additional rows")
            
        # Update Pitcher sheet - Similar approach as for hitters
        if 'Pitcher' in wb.sheetnames and not pitchers_df.empty:
            pitcher_sheet = wb['Pitcher']
            
            # Store all existing data and formulas first
            original_data = {}
            for row in range(2, pitcher_sheet.max_row + 1):
                # Get the existing row data
                row_data = {}
                for col in range(1, pitcher_sheet.max_column + 1):
                    cell = pitcher_sheet.cell(row=row, column=col)
                    # Properly check if cell contains a formula
                    is_formula = False
                    formula_value = None
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        is_formula = True
                        formula_value = cell.value
                        
                    row_data[col] = {
                        'value': cell.value,
                        'is_formula': is_formula,
                        'formula_value': formula_value
                    }
                
                # Use the row number as a key to preserve all rows
                original_data[row] = row_data
            
            # Clear only specific columns in all rows (preserve others for VLOOKUPs)
            columns_to_update = [1, 2, 6, 10]  # Position, Name, Salary, Team
            for row in range(2, pitcher_sheet.max_row + 1):
                for col in columns_to_update:
                    pitcher_sheet.cell(row=row, column=col).value = None
            
            # Add new data from pitchers_df
            logger.info(f"Adding {len(pitchers_df)} rows to Pitcher sheet")
            
            for idx, (_, row) in enumerate(pitchers_df.iterrows(), start=2):
                # Position in column A (1)
                pitcher_sheet.cell(row=idx, column=1).value = "P"
                # Name in column B (2)
                pitcher_sheet.cell(row=idx, column=2).value = row['Name']
                # Salary in column F (6)
                pitcher_sheet.cell(row=idx, column=6).value = row['Salary']
                # Team in column J (10)
                pitcher_sheet.cell(row=idx, column=10).value = row['TeamCode']
            
            # Restore formulas for rows we just updated
            for row in range(2, min(len(pitchers_df) + 2, pitcher_sheet.max_row + 1)):
                if row in original_data:
                    # Preserve any formulas
                    for col in range(1, pitcher_sheet.max_column + 1):
                        if col not in columns_to_update and original_data[row][col]['is_formula']:
                            formula = original_data[row][col]['formula_value']
                            pitcher_sheet.cell(row=row, column=col).value = formula
            
            # Restore formulas for all rows
            for row in range(len(pitchers_df) + 2, pitcher_sheet.max_row + 1):
                if row in original_data:
                    for col in range(1, pitcher_sheet.max_column + 1):
                        if col not in columns_to_update and original_data[row][col]['is_formula']:
                            pitcher_sheet.cell(row=row, column=col).value = original_data[row][col]['formula_value']
            
            logger.info(f"Preserved formulas in {pitcher_sheet.max_row - (len(pitchers_df) + 1)} additional rows")
        
        # Save the workbook
        try:
            wb.save(EXCEL_FILE_PATH)
            logger.info(f"Successfully saved updated workbook to {EXCEL_FILE_PATH}")
            return True
        except Exception as save_e:
            logger.error(f"Error saving workbook: {str(save_e)}")
            return False
    
    except Exception as e:
        logger.error(f"Error updating player data: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False
    
def update_probables_with_formulas(excel_path, probables_df, sheet_name="probables"):
    """
    Update probables sheet while preserving any existing formulas
    
    Parameters:
    excel_path (str): Path to the Excel file
    probables_df (pandas DataFrame): New probables data
    sheet_name (str): Name of the sheet to update
    
    Returns:
    bool: True if successful, False otherwise
    """
    logger.info(f"Updating {sheet_name} with formula preservation...")
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        
        # Check if the sheet exists
        if sheet_name not in wb.sheetnames:
            logger.info(f"Creating new sheet '{sheet_name}'")
            sheet = wb.create_sheet(sheet_name)
            
            # Write headers
            for col_idx, col_name in enumerate(probables_df.columns, 1):
                sheet.cell(row=1, column=col_idx, value=col_name)
            
            # Write data
            for row_idx, row in enumerate(probables_df.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
        else:
            # Get the existing sheet
            sheet = wb[sheet_name]
            
            # Store all existing data and formulas first
            original_data = {}
            for row in range(1, sheet.max_row + 1):
                row_data = {}
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    is_formula = False
                    formula_value = None
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        is_formula = True
                        formula_value = cell.value
                    
                    row_data[col] = {
                        'value': cell.value,
                        'is_formula': is_formula,
                        'formula_value': formula_value
                    }
                
                original_data[row] = row_data
            
            # Identify which columns contain formulas
            formula_columns = set()
            for row_data in original_data.values():
                for col, cell_data in row_data.items():
                    if cell_data['is_formula']:
                        formula_columns.add(col)
            
            logger.info(f"Found formulas in columns: {sorted(formula_columns)}")
            
            # Get current headers
            header_map = {}
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header:
                    header_map[header] = col
            
            # Identify data columns from the DataFrame
            data_columns = []
            for col_name in probables_df.columns:
                if col_name in header_map:
                    col_idx = header_map[col_name]
                    if col_idx not in formula_columns:
                        data_columns.append((col_name, col_idx))
            
            logger.info(f"Will update data columns: {[col[0] for col in data_columns]}")
            
            # Clear only the data columns we'll update
            for row in range(2, sheet.max_row + 1):
                for col_name, col_idx in data_columns:
                    sheet.cell(row=row, column=col_idx).value = None
            
            # Write new data
            for row_idx, (_, row) in enumerate(probables_df.iterrows(), start=2):
                for col_name, col_idx in data_columns:
                    if col_name in row.index:
                        sheet.cell(row=row_idx, column=col_idx).value = row[col_name]
            
            # Restore or copy formulas
            max_data_row = len(probables_df) + 1
            
            # Restore formulas for existing rows
            for row in range(2, min(max_data_row + 1, sheet.max_row + 1)):
                if row in original_data:
                    for col in formula_columns:
                        if col in original_data[row] and original_data[row][col]['is_formula']:
                            formula = original_data[row][col]['formula_value']
                            sheet.cell(row=row, column=col).value = formula
            
            # For new rows beyond the original data, copy formulas from the last template row
            if max_data_row > len(original_data):
                # Find the last row with formulas to use as template
                template_row = None
                for row in sorted(original_data.keys(), reverse=True):
                    has_formulas = False
                    for col in formula_columns:
                        if col in original_data[row] and original_data[row][col]['is_formula']:
                            has_formulas = True
                            break
                    if has_formulas:
                        template_row = row
                        break
                
                if template_row:
                    # Copy formulas to new rows
                    for new_row in range(len(original_data) + 1, max_data_row + 1):
                        for col in formula_columns:
                            if col in original_data[template_row] and original_data[template_row][col]['is_formula']:
                                formula = original_data[template_row][col]['formula_value']
                                # Adjust formula references
                                adjusted_formula = adjust_formula_row_references(formula, template_row, new_row)
                                sheet.cell(row=new_row, column=col).value = adjusted_formula
            
            # Preserve formulas for rows beyond the data
            for row in range(max_data_row + 1, sheet.max_row + 1):
                if row in original_data:
                    for col in formula_columns:
                        if col in original_data[row] and original_data[row][col]['is_formula']:
                            formula = original_data[row][col]['formula_value']
                            sheet.cell(row=row, column=col).value = formula
        
        # Save the workbook
        wb.save(excel_path)
        logger.info(f"Successfully updated {sheet_name} with formula preservation")
        return True
        
    except Exception as e:
        logger.error(f"Error updating {sheet_name}: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False
    
def update_splits_data_with_formulas(sheet_name, data_df):
    """
    Update splits data while preserving formulas, similar to update_player_data.
    Carefully updates only the data columns while preserving formula columns.
    
    Parameters:
    sheet_name (str): Name of the sheet to update
    data_df (pandas DataFrame): New data to insert
    
    Returns:
    bool: True if successful, False otherwise
    """
    logger.info(f"Updating {sheet_name} with formula preservation...")
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH, keep_vba=True)
        
        # Check if the sheet exists
        if sheet_name not in wb.sheetnames:
            logger.info(f"Creating new sheet '{sheet_name}'")
            sheet = wb.create_sheet(sheet_name)
            
            # Write headers
            for col_idx, col_name in enumerate(data_df.columns, 1):
                sheet.cell(row=1, column=col_idx, value=col_name)
            
            # Write data
            for row_idx, row in enumerate(data_df.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
        else:
            # Get the existing sheet
            sheet = wb[sheet_name]
            
            # Store all existing data and formulas first
            original_data = {}
            for row in range(2, sheet.max_row + 1):
                row_data = {}
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    is_formula = False
                    formula_value = None
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        is_formula = True
                        formula_value = cell.value
                    
                    row_data[col] = {
                        'value': cell.value,
                        'is_formula': is_formula,
                        'formula_value': formula_value
                    }
                
                original_data[row] = row_data
            
            # Identify which columns contain formulas (like wOBA)
            formula_columns = set()
            for row_data in original_data.values():
                for col, cell_data in row_data.items():
                    if cell_data['is_formula']:
                        formula_columns.add(col)
            
            logger.info(f"Found formulas in columns: {sorted(formula_columns)}")
            
            # Identify columns to update (non-formula columns from the data)
            header_map = {}
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header:
                    header_map[header] = col
            
            columns_to_update = []
            for col_name in data_df.columns:
                if col_name in header_map:
                    col_idx = header_map[col_name]
                    if col_idx not in formula_columns:
                        columns_to_update.append((col_name, col_idx))
            
            logger.info(f"Will update columns: {[col[0] for col in columns_to_update]}")
            
            # Clear only the data columns we'll update
            for row in range(2, sheet.max_row + 1):
                for col_name, col_idx in columns_to_update:
                    sheet.cell(row=row, column=col_idx).value = None
            
            # Write new data
            for row_idx, (_, row) in enumerate(data_df.iterrows(), start=2):
                for col_name, col_idx in columns_to_update:
                    if col_name in row.index:
                        sheet.cell(row=row_idx, column=col_idx).value = row[col_name]
            
            # Restore formulas for existing rows
            max_data_row = len(data_df) + 1
            for row in range(2, min(max_data_row + 1, sheet.max_row + 1)):
                if row in original_data:
                    for col in formula_columns:
                        if col in original_data[row] and original_data[row][col]['is_formula']:
                            formula = original_data[row][col]['formula_value']
                            sheet.cell(row=row, column=col).value = formula
            
            # For new rows, copy formulas from the last row with formulas
            if max_data_row > len(original_data) + 1:
                # Find the last row with formulas
                template_row = None
                for row in sorted(original_data.keys(), reverse=True):
                    has_formulas = False
                    for col in formula_columns:
                        if col in original_data[row] and original_data[row][col]['is_formula']:
                            has_formulas = True
                            break
                    if has_formulas:
                        template_row = row
                        break
                
                if template_row:
                    # Copy formulas to new rows
                    for new_row in range(len(original_data) + 2, max_data_row + 1):
                        for col in formula_columns:
                            if col in original_data[template_row] and original_data[template_row][col]['is_formula']:
                                formula = original_data[template_row][col]['formula_value']
                                # Adjust formula references
                                adjusted_formula = adjust_formula_row_references(formula, template_row, new_row)
                                sheet.cell(row=new_row, column=col).value = adjusted_formula
            
            # Preserve formulas for rows beyond the data
            for row in range(max_data_row + 1, sheet.max_row + 1):
                if row in original_data:
                    for col in formula_columns:
                        if col in original_data[row] and original_data[row][col]['is_formula']:
                            formula = original_data[row][col]['formula_value']
                            sheet.cell(row=row, column=col).value = formula
        
        # Save the workbook
        wb.save(EXCEL_FILE_PATH)
        logger.info(f"Successfully updated {sheet_name} with formula preservation")
        return True
    
    except Exception as e:
        logger.error(f"Error updating {sheet_name}: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def adjust_formula_row_references(formula, old_row, new_row):
    """
    Adjust row references in a formula when copying to a new row

    Parameters:
    formula (str): Formula to modify
    old_row (int): Original row index
    new_row(int): Target row index

    Returns:
    str: Adjusted formula with updated reference rows
    """
    import re
    
    row_diff = new_row - old_row
    
    # Pattern to match cell references (e.g., A1, $B$2, C$3)
    cell_pattern = r'(\$?)([A-Z]+)(\$?)(\d+)'
    
    def replace_ref(match):
        dollar1, col, dollar2, row = match.groups()
        
        # Only adjust if the row reference is not absolute (no $ before the number)
        if not dollar2:
            new_row_num = int(row) + row_diff
            return f"{dollar1}{col}{dollar2}{new_row_num}"
        else:
            return match.group(0)
    
    adjusted = re.sub(cell_pattern, replace_ref, formula)
    return adjusted

def process_data(process_func, reference_df=None, *args, **kwargs):
    """
    Data processing function that uses the same name standardization
    for all data sources
    
    Parameters:
    process_func (function): Original data processing function 
    reference_df (pandas DataFrame): Reference DataFrame with correct names
    *args, **kwargs: Arguments to pass to the original function
    
    Returns:
    pandas DataFrame: Processed data with standardized player names
    """
    # Call the original function to get data
    df = process_func(*args, **kwargs)
    
    # Standardize all player names using our comprehensive function
    if isinstance(df, pd.DataFrame) and not df.empty and 'Name' in df.columns:
        df = standardize_all_player_names(df, reference_df)
    
    return df

def get_player_handedness_from_desktop_files():
    """
    Load and process player handedness data from the PitcherHandedness and BatterHandedness files on the Desktop
    Handles both regular Desktop and OneDrive Desktop paths. Adjust as needed.
    
    Returns:
    pandas DataFrame: Processed handedness data with columns for Name, Team, Bats, and Throws
    """
    logger.info("Loading handedness data from Desktop files...")
    
    # Find potential Desktop folder paths (both regular and OneDrive)
    possible_paths = [
        os.path.join(os.path.expanduser("~"), "Desktop"),
        os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop"),
        os.path.join(os.path.expanduser("~"), "OneDrive - Personal", "Desktop"),
        os.path.join(os.path.expanduser("~"), "OneDrive - Business", "Desktop")
    ]
    
    # Also check the current directory where the script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))
    if script_dir not in possible_paths:
        possible_paths.append(script_dir)
    
    # Try to locate the handedness files
    pitcher_file = None
    batter_file = None
    
    for path in possible_paths:
        temp_pitcher_file = os.path.join(path, "PitcherHandedness.csv")
        temp_batter_file = os.path.join(path, "BatterHandedness.csv")
        
        if os.path.exists(temp_pitcher_file):
            pitcher_file = temp_pitcher_file
            logger.info(f"Found pitcher handedness file at: {pitcher_file}")
        
        if os.path.exists(temp_batter_file):
            batter_file = temp_batter_file
            logger.info(f"Found batter handedness file at: {batter_file}")
        
        # If both files found, break the loop
        if pitcher_file and batter_file:
            break
    
    pitcher_df = pd.DataFrame()
    batter_df = pd.DataFrame()
    
    # Load pitcher handedness data
    if pitcher_file and os.path.exists(pitcher_file):
        try:
            pitcher_df = pd.read_csv(pitcher_file)
            logger.info(f"Loaded pitcher handedness data: {len(pitcher_df)} records")
            
            # Make sure we have the required columns
            if 'Name' not in pitcher_df.columns or 'Throws' not in pitcher_df.columns:
                logger.warning("Pitcher handedness file is missing required columns")
                pitcher_df = pd.DataFrame()
        except Exception as e:
            logger.error(f"Error loading pitcher handedness data: {str(e)}")
            pitcher_df = pd.DataFrame()
    else:
        logger.warning(f"Pitcher handedness file not found in any of the checked locations")
    
    # Load batter handedness data
    if batter_file and os.path.exists(batter_file):
        try:
            batter_df = pd.read_csv(batter_file)
            logger.info(f"Loaded batter handedness data: {len(batter_df)} records")
            
            # Make sure we have the required columns
            if 'Name' not in batter_df.columns or 'Bats' not in batter_df.columns:
                logger.warning("Batter handedness file is missing required columns")
                batter_df = pd.DataFrame()
        except Exception as e:
            logger.error(f"Error loading batter handedness data: {str(e)}")
            batter_df = pd.DataFrame()
    else:
        logger.warning(f"Batter handedness file not found in any of the checked locations")
    
    # If both dataframes are empty, fall back to a default dataset
    if pitcher_df.empty and batter_df.empty:
        logger.warning("Could not load handedness data from files. Using default dataset.")
        return create_default_handedness_data()
    
    # The rest of the function remains the same as before
    # Process the data for the final combined dataset
    result_df = pd.DataFrame()
    
    # Process batters
    if not batter_df.empty:
        logger.info(f"Using data from {len(batter_df)} batters")
        result_df = batter_df.copy()
        
        # If 'Throws' column doesn't exist, add it
        if 'Throws' not in result_df.columns:
            result_df['Throws'] = None
        
        # Add 'Team' column if it doesn't exist
        if 'Team' not in result_df.columns:
            result_df['Team'] = ""
    
    # Add pitchers' throwing hand data
    if not pitcher_df.empty:
        logger.info(f"Adding data from {len(pitcher_df)} pitchers")
        
        # If we have batters data
        if not result_df.empty:
            # Create lookup dictionary for pitchers
            pitcher_data = {}
            for _, row in pitcher_df.iterrows():
                name = row['Name']
                throws = row.get('Throws', None)
                team = row.get('Team', "")
                pitcher_data[name] = {'Throws': throws, 'Team': team}
            
            # Shohei Rule: Update existing batters with pitcher data if names match
            for idx, row in result_df.iterrows():
                name = row['Name']
                if name in pitcher_data:
                    # Update Throws if it's None or empty
                    if pd.isna(result_df.at[idx, 'Throws']) or result_df.at[idx, 'Throws'] == '':
                        result_df.at[idx, 'Throws'] = pitcher_data[name]['Throws']
                    
                    # Update Team if it's empty
                    if result_df.at[idx, 'Team'] == '':
                        result_df.at[idx, 'Team'] = pitcher_data[name]['Team']
            
            # Add pitchers who aren't in the batters list
            new_pitchers = []
            for _, row in pitcher_df.iterrows():
                name = row['Name']
                if name not in result_df['Name'].values:
                    new_row = {'Name': name, 'Throws': row.get('Throws', None), 'Team': row.get('Team', "")}
                    if 'Bats' in result_df.columns:
                        new_row['Bats'] = None
                    new_pitchers.append(new_row)
            
            # Add new pitchers to the result
            if new_pitchers:
                new_df = pd.DataFrame(new_pitchers)
                result_df = pd.concat([result_df, new_df], ignore_index=True)
        else:
            # If we only have pitchers data
            result_df = pitcher_df.copy()
            
            # If 'Bats' column doesn't exist, add it
            if 'Bats' not in result_df.columns:
                result_df['Bats'] = None
            
            # Add 'Team' column if it doesn't exist
            if 'Team' not in result_df.columns:
                result_df['Team'] = ""
    
    # Fill in missing values with reasonable defaults
    # For throws, right-handed is the default
    result_df['Throws'] = result_df['Throws'].fillna('R')
    
    # For batting, match throwing hand if possible, otherwise right-handed is default
    if 'Bats' in result_df.columns:
        result_df['Bats'] = result_df.apply(
            lambda row: row['Throws'] if pd.isna(row['Bats']) and row['Throws'] == 'L' else ('R' if pd.isna(row['Bats']) else row['Bats']), 
            axis=1
        )
    
    # Make sure we have all necessary columns with the right names
    required_cols = ['Name', 'Team', 'Bats', 'Throws']
    for col in required_cols:
        if col not in result_df.columns:
            result_df[col] = "" if col == 'Team' else None
    
    # Keep only the columns we want in the right order
    result_df = result_df[required_cols]
    
    # Remove any duplicate rows
    result_df = result_df.drop_duplicates(subset=['Name'])
    
    logger.info(f"Successfully compiled handedness data for {len(result_df)} players")
    return result_df

def get_player_handedness_from_exports(batters_csv=None, pitchers_csv=None, delete_after=True):
    """
    Process manually exported FanGraphs CSV files for handedness data - DEPRECATED
    This function is kept for backward compatibility but is no longer used.
    Use get_player_handedness_from_desktop_files() instead.
    
    Parameters:
    batters_csv (str): Path to the exported batters CSV file. If None, will look for a file.
    pitchers_csv (str): Path to the exported pitchers CSV file. If None, will look for a file.
    delete_after (bool): Whether to delete the CSV files after processing
    
    Returns:
    pandas DataFrame: Player handedness data with columns for player name, team, and handedness
    """
    logger.warning("get_player_handedness_from_exports is deprecated, using desktop files instead")
    return get_player_handedness_from_desktop_files()

def get_player_handedness():
    """
    Main handedness function that reads data from local files
    
    Returns:
    pandas DataFrame: Player handedness data with columns for player name, team, and handedness
    """
    logger.info("Getting player handedness data...")
    
    # Use the desktop files for handedness data
    result = get_player_handedness_from_desktop_files()
    
    if result.empty:
        # Fall back to default data if desktop files couldn't be processed
        logger.warning("Using default handedness data as fallback")
        result = create_default_handedness_data()
        
    return result

def create_default_handedness_data():
    """Create a default handedness dataset with known MLB players"""
    logger.warning("Creating default handedness data for common MLB players")
    
    # Build a baseline with common players if things fail
    data = {
        'Name': ['Shohei Ohtani', 'Aaron Judge', 'Juan Soto', 'Bryce Harper', 'Mookie Betts',
                'Clayton Kershaw', 'Mike Trout', 'Freddie Freeman', 'Ronald Acuña Jr.', 'Gerrit Cole',
                'Manny Machado', 'Yordan Alvarez', 'Corbin Burnes', 'Vladimir Guerrero Jr.', 'Zack Wheeler',
                'Jose Altuve', 'Kyle Tucker', 'Fernando Tatis Jr.', 'Gunnar Henderson', 'Francisco Lindor',
                'Luis Robert Jr.', 'Marcus Semien', 'Rafael Devers', 'Bobby Witt Jr.', 'Max Scherzer',
                'Julio Rodríguez', 'Pete Alonso', 'Jazz Chisholm Jr.', 'Paul Goldschmidt', 'Nolan Arenado',
                'Zac Gallen', 'Tyler Glasnow', 'Shane Bieber', 'Corey Seager', 'Matt Olson',
                'Spencer Strider', 'Randy Arozarena', 'José Ramírez', 'Xander Bogaerts', 'Aaron Nola',
                'Bryan Reynolds', 'Cedric Mullins', 'Salvador Perez', 'Zack Greinke', 'Austin Riley',
                'Kyle Schwarber', 'Bo Bichette', 'Justin Verlander', 'Byron Buxton', 'Dylan Cease'],
        'Team': ['LAD', 'NYY', 'NYY', 'PHI', 'LAD', 
                'TEX', 'LAA', 'LAD', 'ATL', 'NYY', 
                'SDP', 'HOU', 'BAL', 'TOR', 'PHI', 
                'HOU', 'HOU', 'SDP', 'BAL', 'NYM',
                'CHW', 'TEX', 'BOS', 'KCR', 'TEX',
                'SEA', 'NYM', 'MIA', 'STL', 'STL',
                'ARI', 'LAD', 'CLE', 'TEX', 'ATL',
                'ATL', 'TBR', 'CLE', 'SDP', 'PHI',
                'PIT', 'BAL', 'KCR', 'KCR', 'ATL',
                'PHI', 'TOR', 'HOU', 'MIN', 'CHW'],
        'Bats': ['L', 'R', 'L', 'L', 'R', 'L', 'R', 'L', 'R', 'R', 
                 'R', 'L', 'R', 'R', 'R', 'R', 'L', 'R', 'L', 'B',
                 'R', 'R', 'L', 'R', 'R',
                 'R', 'R', 'B', 'R', 'R',
                 'R', 'R', 'R', 'L', 'L',
                 'R', 'R', 'B', 'R', 'R',
                 'B', 'L', 'R', 'R', 'R',
                 'L', 'R', 'R', 'R', 'R'],
        'Throws': ['L', 'R', 'L', 'R', 'R', 'L', 'R', 'R', 'R', 'R',
                  'R', 'R', 'R', 'R', 'R', 'R', 'R', 'R', 'R', 'R',
                  'R', 'R', 'R', 'R', 'R',
                  'R', 'R', 'R', 'R', 'R',
                  'R', 'R', 'R', 'R', 'R',
                  'R', 'R', 'R', 'R', 'R',
                  'R', 'R', 'R', 'R', 'R',
                  'R', 'R', 'R', 'R', 'R']
    }
    
    default_df = pd.DataFrame(data)
    logger.warning(f"Created default handedness data with {len(default_df)} players")
    return default_df

def get_salaries_for_reference():
    """
    Get DraftKings salaries to use as a reference for player names
    This will be used as the "source of truth" for player name formats
    
    Returns:
    pandas DataFrame: Salaries data with player names, or None if not available
    """
    try:
        # Try to read from the existing Excel file
        if os.path.exists(EXCEL_FILE_PATH):
            salaries_df = pd.read_excel(EXCEL_FILE_PATH, sheet_name='Salaries')
            if not salaries_df.empty and 'Name' in salaries_df.columns:
                logger.info(f"Using existing Salaries sheet as name reference with {len(salaries_df)} players")
                return salaries_df
        
        # If that fails, get fresh salary data
        logger.info("Getting fresh salary data as name reference")
        salaries_df = get_draftkings_salaries_csv()
        if not salaries_df.empty and 'Name' in salaries_df.columns:
            logger.info(f"Using fresh salary data as name reference with {len(salaries_df)} players")
            return salaries_df
            
    except Exception as e:
        logger.error(f"Error getting salary reference data: {str(e)}")
    
    logger.warning("No salary reference data available for name standardization")
    return None

def get_fangraphs_probables():
    """
    Get probable pitchers from FanGraphs probables grid for today's games only
    Using a fixed column index approach - column 1 is always today's games
    For doubleheaders, defaults to game 2
    """
    logger.info("Getting probable pitchers from FanGraphs with fixed column index...")

    try:
        # Get current date for filtering
        today = datetime.now().strftime("%Y-%m-%d")

        # Chrome options
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--ignore-certificate-errors")

        # URL to probables grid
        url = "https://www.fangraphs.com/roster-resource/probables-grid"

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.get(url)

        # Wait for content to load
        logger.info("Waiting for page to load...")
        time.sleep(10)

        today_col_idx = 1

        # Find rows with team codes in the first column (3 letters or less)
        all_matchups = []

        # Team code correction to match salaries
        fangraphs_team_code_map = {
            "WSN": "WSH",
            "TBR": "TB",
            "SDP": "SD",
            "KCR": "KC",
            "CHW": "CWS",
            "SFG": "SF"
        }

        # XPath to find team rows with valid team codes
        team_rows = driver.find_elements(By.XPATH, "//tr[td[1][string-length(text()) <= 3 and string-length(text()) > 0]]")
        logger.info(f"Found {len(team_rows)} potential team rows")

        for row in team_rows:
            try:
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) <= today_col_idx:
                    continue

                team_cell = cells[0]
                team_code_raw = team_cell.text.strip()

                # Skip if team code is empty
                if not team_code_raw:
                    continue

                # Apply mapping correction if necessary
                team_code = fangraphs_team_code_map.get(team_code_raw, team_code_raw)

                # Get today's game cell
                game_cell = cells[today_col_idx]
                game_text = game_cell.text.strip()

                # Skip if this is an OFF day or empty cell
                if not game_text or "OFF" in game_text:
                    continue

                # Log for debugging
                logger.info(f"Processing {team_code} game: {game_text}")

                # Parse the game cell to extract opponent and pitcher
                lines = game_text.split('\n')

                opponent_code_raw = None
                pitcher_name = "TBD"
                is_away = False

                # First line should have opponent - handle doubleheader marker
                if lines and lines[0]:
                    # Check for doubleheader and clean up
                    if "(2)" in lines[0]:
                        lines[0] = lines[0].replace("(2)", "").strip()
                    
                    # Away game pattern: "@ TEAM"
                    if lines[0].startswith("@"):
                        is_away = True
                        opponent_code_raw = lines[0].replace("@", "").strip()
                    else:
                        # Home game pattern: "TEAM"
                        opponent_code_raw = lines[0].strip()

                # Handle pitcher selection - for doubleheaders, use the second pitcher if available
                if len(lines) > 2 and lines[1] and lines[2]:
                    # This might be a doubleheader with multiple pitchers listed
                    pitcher_text_2 = lines[2]
                    
                    # Default to second game's pitcher for doubleheaders
                    pitcher_text = pitcher_text_2
                    
                    # Extract pitcher name
                    if "(" in pitcher_text:
                        pitcher_name = pitcher_text.split("(")[0].strip()
                    else:
                        pitcher_name = pitcher_text.strip()
                    
                    logger.info(f"Selected pitcher from doubleheader: {pitcher_name}")
                
                # Regular single game
                elif len(lines) > 1 and lines[1]:
                    pitcher_text = lines[1]
                    if "(" in pitcher_text:
                        pitcher_name = pitcher_text.split("(")[0].strip()
                    else:
                        pitcher_name = pitcher_text.strip()

                # Skip if no opponent found
                if not opponent_code_raw:
                    continue

                # Apply mapping correction to opponent code as well
                opponent_code = fangraphs_team_code_map.get(opponent_code_raw, opponent_code_raw)

                # Create matchup record
                if is_away:
                    matchup = {
                        "Date": today,
                        "AwayTeam": team_code,
                        "HomeTeam": opponent_code,
                        "AwayPitcher": pitcher_name,
                        "HomePitcher": "TBD",
                        "AwayTeamFull": team_code,
                        "HomeTeamFull": opponent_code
                    }
                else:
                    matchup = {
                        "Date": today,
                        "AwayTeam": opponent_code,
                        "HomeTeam": team_code,
                        "AwayPitcher": "TBD",
                        "HomePitcher": pitcher_name,
                        "AwayTeamFull": opponent_code,
                        "HomeTeamFull": team_code
                    }

                all_matchups.append(matchup)
                logger.info(f"Added matchup: {matchup['AwayTeam']} @ {matchup['HomeTeam']}, "
                            f"A: {matchup['AwayPitcher']}, H: {matchup['HomePitcher']}")

            except Exception as row_e:
                logger.warning(f"Error processing row: {str(row_e)}")

        # Process matchups to remove duplicates and combine pitcher info
        unique_games = {}

        for matchup in all_matchups:
            game_key = f"{matchup['AwayTeam']}@{matchup['HomeTeam']}"

            if game_key not in unique_games:
                unique_games[game_key] = matchup
            else:
                if matchup['AwayPitcher'] != "TBD" and unique_games[game_key]['AwayPitcher'] == "TBD":
                    unique_games[game_key]['AwayPitcher'] = matchup['AwayPitcher']
                if matchup['HomePitcher'] != "TBD" and unique_games[game_key]['HomePitcher'] == "TBD":
                    unique_games[game_key]['HomePitcher'] = matchup['HomePitcher']

        clean_matchups = list(unique_games.values())

        final_matchups = [m for m in clean_matchups
                          if m['AwayTeam'] and m['HomeTeam']
                          and len(m['AwayTeam']) >= 2 and len(m['HomeTeam']) >= 2]

        logger.info(f"Found {len(final_matchups)} matchups for today after cleanup")

        df = pd.DataFrame(final_matchups) if final_matchups else pd.DataFrame()

        if df.empty:
            logger.warning("No games found on FanGraphs, falling back to ESPN")
            driver.quit()
            return get_probable_pitchers_espn()

        df = df.sort_values('AwayTeam').reset_index(drop=True)

        driver.quit()
        return df

    except Exception as e:
        logger.error(f"Error in get_fangraphs_probables: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())

        if 'driver' in locals():
            driver.quit()

        logger.info("Falling back to ESPN for probable pitchers")
        return get_probable_pitchers_espn()

    
def get_probable_pitchers_espn():
    """
    Get probable pitchers from ESPN schedule page for today's games only, fallback method if fangraphs fails
    
    Returns:
    pandas DataFrame: Properly formatted probable pitchers data with columns:
        - Date: Current date
        - AwayTeam: 3-letter code for away team
        - HomeTeam: 3-letter code for home team
        - AwayPitcher: Name of away team pitcher
        - HomePitcher: Name of home team pitcher
        - AwayTeamFull: Full name of away team
        - HomeTeamFull: Full name of home team
    """
    logger.info("Getting probable pitchers from ESPN for today's games only...")
    
    try:
        # Get today's date for filtering
        today = datetime.now().strftime("%Y-%m-%d")
        today_date_path = datetime.now().strftime("%Y%m%d")
        
        # ESPN URL for today's schedule only - this is key to getting only today's games
        url = f"https://www.espn.com/mlb/schedule/_/date/{today_date_path}"
        
        logger.info(f"Fetching ESPN schedule for today only: {url}")
        
        # Setup headless Chrome
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.get(url)
        
        # Wait for content to load
        time.sleep(5)
        
        # Define standard team codes mapping
        team_codes = {
            'ARI': 'ARI', 'Diamondbacks': 'ARI', 'Arizona': 'ARI', 
            'ATL': 'ATL', 'Braves': 'ATL', 'Atlanta': 'ATL', 
            'BAL': 'BAL', 'Orioles': 'BAL', 'Baltimore': 'BAL',
            'BOS': 'BOS', 'Red Sox': 'BOS', 'Boston': 'BOS',
            'CHC': 'CHC', 'Cubs': 'CHC', 'Chicago Cubs': 'CHC',
            'CHW': 'CHW', 'White Sox': 'CHW', 'Chicago White Sox': 'CHW',
            'CIN': 'CIN', 'Reds': 'CIN', 'Cincinnati': 'CIN',
            'CLE': 'CLE', 'Guardians': 'CLE', 'Cleveland': 'CLE',
            'COL': 'COL', 'Rockies': 'COL', 'Colorado': 'COL',
            'DET': 'DET', 'Tigers': 'DET', 'Detroit': 'DET',
            'HOU': 'HOU', 'Astros': 'HOU', 'Houston': 'HOU',
            'KC': 'KC', 'Royals': 'KC', 'Kansas City': 'KC', 'KCR': 'KC',
            'LAA': 'LAA', 'Angels': 'LAA', 'Los Angeles Angels': 'LAA', 'LA Angels': 'LAA',
            'LAD': 'LAD', 'Dodgers': 'LAD', 'Los Angeles Dodgers': 'LAD', 'LA Dodgers': 'LAD',
            'MIA': 'MIA', 'Marlins': 'MIA', 'Miami': 'MIA',
            'MIL': 'MIL', 'Brewers': 'MIL', 'Milwaukee': 'MIL',
            'MIN': 'MIN', 'Twins': 'MIN', 'Minnesota': 'MIN',
            'NYM': 'NYM', 'Mets': 'NYM', 'New York Mets': 'NYM',
            'NYY': 'NYY', 'Yankees': 'NYY', 'New York Yankees': 'NYY',
            'OAK': 'OAK', 'Athletics': 'OAK', 'Oakland': 'OAK', 'A\'s': 'OAK',
            'PHI': 'PHI', 'Phillies': 'PHI', 'Philadelphia': 'PHI',
            'PIT': 'PIT', 'Pirates': 'PIT', 'Pittsburgh': 'PIT',
            'SD': 'SD', 'Padres': 'SD', 'San Diego': 'SD', 'SDP': 'SD',
            'SF': 'SF', 'Giants': 'SF', 'San Francisco': 'SF', 'SFG': 'SF',
            'SEA': 'SEA', 'Mariners': 'SEA', 'Seattle': 'SEA',
            'STL': 'STL', 'Cardinals': 'STL', 'St. Louis': 'STL',
            'TB': 'TB', 'Rays': 'TB', 'Tampa Bay': 'TB', 'TBR': 'TB',
            'TEX': 'TEX', 'Rangers': 'TEX', 'Texas': 'TEX',
            'TOR': 'TOR', 'Blue Jays': 'TOR', 'Toronto': 'TOR',
            'WSH': 'WSH', 'Nationals': 'WSH', 'Washington': 'WSH', 'WAS': 'WSH'
        }
        
        # Function to get team code
        def get_team_code(team_name):
            if not team_name or not isinstance(team_name, str):
                return ""
                
            # Check exact match first (case sensitive)
            if team_name in team_codes:
                return team_codes[team_name]
                
            # Check case-insensitive match
            for name, code in team_codes.items():
                if name.lower() == team_name.lower():
                    return code
            
            # Try partial match
            for name, code in team_codes.items():
                if name.lower() in team_name.lower() or team_name.lower() in name.lower():
                    return code
            
            # Return original if no match found
            return team_name
        
        # Collect all matchups
        all_matchups = []
        
        # Skip the CSS selectors and go directly to read_html
        try:
            # Parse HTML tables - this is much more reliable
            tables = pd.read_html(StringIO(driver.page_source))
            
            # Try to find the schedule table
            schedule_table = None
            for table in tables:
                # Look for columns that would indicate this is a schedule table
                cols = [str(col).lower() for col in table.columns]
                if any('matchup' in col for col in cols) or any('game' in col for col in cols):
                    schedule_table = table
                    break
            
            if schedule_table is not None:
                logger.info(f"Found schedule table with columns: {schedule_table.columns}")
                
                # Find the matchup column
                matchup_col = None
                for col in schedule_table.columns:
                    if 'MATCHUP' in str(col).upper() or 'GAME' in str(col).upper():
                        matchup_col = col
                        break
                
                if matchup_col is not None:
                    # Try to find pitcher column if it exists
                    pitcher_col = None
                    for col in schedule_table.columns:
                        if 'PITCHER' in str(col).upper() or 'PITCHING' in str(col).upper():
                            pitcher_col = col
                            break
                    
                    # Process each row in the table
                    for _, row in schedule_table.iterrows():
                        # Skip rows without matchup data
                        if matchup_col not in row or pd.isna(row[matchup_col]):
                            continue
                            
                        matchup = str(row[matchup_col])
                        
                        # Skip if completed game or not matchup data
                        if 'FINAL' in matchup.upper() or 'POSTPONED' in matchup.upper():
                            continue
                        
                        # Parse team names - need to handle possible formats
                        away_team = None
                        home_team = None
                        
                        # Format: "Away @ Home"
                        if ' @ ' in matchup:
                            parts = matchup.split(' @ ')
                            if len(parts) == 2:
                                away_team = parts[0].strip()
                                home_team = parts[1].strip()
                        # Format: "Away at Home"
                        elif ' at ' in matchup:
                            parts = matchup.split(' at ')
                            if len(parts) == 2:
                                away_team = parts[0].strip()
                                home_team = parts[1].strip()
                        # Format: "Away vs. Home" or "Away vs Home"
                        elif ' vs. ' in matchup or ' vs ' in matchup:
                            parts = re.split(r' vs\.? ', matchup)
                            if len(parts) == 2:
                                away_team = parts[0].strip()
                                home_team = parts[1].strip()
                        
                        # Skip if we didn't find both teams
                        if not away_team or not home_team:
                            continue
                        
                        # Default to TBD for pitchers
                        away_pitcher = "TBD"
                        home_pitcher = "TBD"
                        
                        # Try to get pitchers if column exists
                        if pitcher_col is not None and pitcher_col in row and not pd.isna(row[pitcher_col]):
                            pitching_text = str(row[pitcher_col])
                            
                            # Format: "Pitcher1 vs. Pitcher2"
                            if ' vs. ' in pitching_text:
                                parts = pitching_text.split(' vs. ')
                                if len(parts) == 2:
                                    away_pitcher = parts[0].strip()
                                    home_pitcher = parts[1].strip()
                            # Format: "Pitcher1 v Pitcher2"
                            elif ' v ' in pitching_text:
                                parts = pitching_text.split(' v ')
                                if len(parts) == 2:
                                    away_pitcher = parts[0].strip()
                                    home_pitcher = parts[1].strip()
                        
                        # Get team codes
                        away_code = get_team_code(away_team)
                        home_code = get_team_code(home_team)
                        
                        # Add to matchups
                        all_matchups.append({
                            'Date': today,
                            'AwayTeam': away_code,
                            'HomeTeam': home_code,
                            'AwayPitcher': away_pitcher,
                            'HomePitcher': home_pitcher,
                            'AwayTeamFull': away_team,
                            'HomeTeamFull': home_team
                        })
            
        except Exception as table_e:
            logger.warning(f"Error parsing HTML tables: {str(table_e)}")
        
        # If we found matchups, return as DataFrame
        if all_matchups:
            df = pd.DataFrame(all_matchups)
            logger.info(f"Successfully found {len(df)} games from ESPN for today")
            driver.quit()
            return df
        
        # Direct text parsing as a last resort
        try:
            # Get the page source and search for game patterns
            page_source = driver.page_source
            
            # Look for patterns like "Team @ Team" in the page text
            matchup_pattern = r'([A-Za-z\s.]+)\s+@\s+([A-Za-z\s.]+)'
            matchups = re.findall(matchup_pattern, page_source)
            
            logger.info(f"Found {len(matchups)} potential matchups via text pattern search")
            
            for away_team, home_team in matchups:
                # Skip if teams don't match team names format
                if len(away_team) < 3 or len(home_team) < 3:
                    continue
                    
                # Clean up team names
                away_team = away_team.strip()
                home_team = home_team.strip()
                
                # Get team codes
                away_code = get_team_code(away_team)
                home_code = get_team_code(home_team)
                
                # Only add if these look like valid team codes
                if len(away_code) == 2 or len(away_code) == 3:
                    if len(home_code) == 2 or len(home_code) == 3:
                        # Add to matchups with TBD pitchers
                        all_matchups.append({
                            'Date': today,
                            'AwayTeam': away_code,
                            'HomeTeam': home_code,
                            'AwayPitcher': "TBD",
                            'HomePitcher': "TBD",
                            'AwayTeamFull': away_team,
                            'HomeTeamFull': home_team
                        })
            
            # If we found matchups, return as DataFrame
            if all_matchups:
                df = pd.DataFrame(all_matchups)
                logger.info(f"Successfully found {len(df)} games via text pattern search")
                driver.quit()
                return df
                
        except Exception as pattern_e:
            logger.warning(f"Error in text pattern search: {str(pattern_e)}")
        
        # If we got here, no matchups were found
        logger.warning("No matchups found on ESPN for today, using dummy data")
        
        # Create dummy data for today only
        df = pd.DataFrame({
            "Date": [today],
            "AwayTeam": ["---"],
            "HomeTeam": ["---"],
            "AwayPitcher": ["No games scheduled for today"],
            "HomePitcher": ["No games scheduled for today"],
            "AwayTeamFull": ["---"],
            "HomeTeamFull": ["---"]
        })
        
        driver.quit()
        return df
        
    except Exception as e:
        logger.error(f"Error getting ESPN probables: {str(e)}")
        
        if 'driver' in locals():
            driver.quit()
        
        # Return dummy data on complete failure
        today = datetime.now().strftime("%Y-%m-%d")
        return pd.DataFrame({
            "Date": [today],
            "AwayTeam": ["---"],
            "HomeTeam": ["---"],
            "AwayPitcher": ["Error fetching data"],
            "HomePitcher": ["Error fetching data"],
            "AwayTeamFull": ["---"],
            "HomeTeamFull": ["---"]
        })

def synchronize_handedness_names_with_salaries(handedness_df, salaries_df, name_column='Name'):
    """
    Specifically synchronize handedness data names with the salary data names
    This is an extra step beyond standardization to ensure exact matches
    
    Parameters:
    handedness_df (pandas DataFrame): DataFrame containing handedness data
    salaries_df (pandas DataFrame): DataFrame containing salary data
    name_column (str): Name of the column containing player names
    
    Returns:
    pandas DataFrame: Handedness data with names matched to salary data
    """
    logger.info(f"Synchronizing handedness names with salaries for {len(handedness_df)} rows")
    
    if handedness_df.empty or salaries_df.empty:
        return handedness_df
    
    result_df = handedness_df.copy()
    
    # Create sets of names for lookups
    handedness_names = set(handedness_df[name_column].dropna())
    salary_names = set(salaries_df[name_column].dropna())
    
    # Create lowercase mappings for case-insensitive matching
    salary_names_lower = {name.lower(): name for name in salary_names if isinstance(name, str)}
    
    # Count matches and mismatches
    direct_matches = handedness_names.intersection(salary_names)
    logger.info(f"Initial direct matches: {len(direct_matches)} out of {len(handedness_names)}")
    
    # Track all changes for logging
    changes = []
    
    # For each handedness name, find its best match in salary names
    for idx, row in result_df.iterrows():
        name = row[name_column]
        if pd.isna(name) or not isinstance(name, str) or name in salary_names:
            continue  # Skip NaN, non-string, or exact matches
        
        # Try lowercase match first
        if name.lower() in salary_names_lower:
            salary_match = salary_names_lower[name.lower()]
            result_df.at[idx, name_column] = salary_match
            changes.append((name, salary_match))
            continue
        
        # Try removing accent alternatives (name with accents -> name without accents)
        unaccented_name = ''.join(c for c in unicodedata.normalize('NFD', name) 
                                 if unicodedata.category(c) != 'Mn')
        if unaccented_name in salary_names:
            result_df.at[idx, name_column] = unaccented_name
            changes.append((name, unaccented_name))
            continue
            
        # Try common Latino name patterns
        for salary_name in salary_names:
            if not isinstance(salary_name, str):
                continue
                
            # For José vs Jose patterns
            if name.replace('é', 'e') == salary_name or name.replace('e', 'é') == salary_name:
                result_df.at[idx, name_column] = salary_name
                changes.append((name, salary_name))
                break
                
            # For Muñoz vs Munoz patterns
            if name.replace('ñ', 'n') == salary_name or name.replace('n', 'ñ') == salary_name:
                result_df.at[idx, name_column] = salary_name
                changes.append((name, salary_name))
                break
    
    # Log results
    logger.info(f"Made {len(changes)} handedness-to-salary name adjustments")
    if changes:
        for old, new in changes[:20]:  # Show first 20 changes
            logger.info(f"  Adjusted: '{old}' -> '{new}'")
    
    # Calculate final match rate
    final_names = set(result_df[name_column].dropna())
    final_matches = final_names.intersection(salary_names)
    match_rate = len(final_matches) / len(final_names) if final_names else 0
    logger.info(f"Final direct matches: {len(final_matches)} out of {len(final_names)} ({match_rate:.1%})")
    
    # List remaining mismatches for debugging
    mismatches = final_names - salary_names
    if mismatches:
        logger.info(f"Sample of {min(10, len(mismatches))} remaining mismatches:")
        for name in list(mismatches)[:10]:
            logger.info(f"  Missing in salaries: '{name}'")
    
    return result_df

def normalize_name(name):
    return unicodedata.normalize("NFC", name.strip()) if name else ''

def synchronize_datasets_for_lookups():
    """
    Synchronization step to ensure all dataset names match for VLOOKUP compatibility
    Uses a dynamic approach rather than hardcoded mappings
    """
    logger.info("Performing final synchronization between datasets for VLOOKUP compatibility")
    
    try:
        # Helper to remove accented letters
        def remove_accents(name):
            if not name or not isinstance(name, str):
                return name
            import unicodedata
            normalized = unicodedata.normalize('NFKD', name)
            return ''.join([c for c in normalized if not unicodedata.combining(c)])
            
        # Helper function to standardize a name
        def standardize_name(name):
            if not name or not isinstance(name, str):
                return name
                
            # Clean the name for lookup
            clean_name = remove_accents(name.lower())
            
            # Special case for Luisangel/Luis Angel Acuna
            if "luis angel" in clean_name and "acuna" in clean_name:
                clean_name = "luisangel acuna"
                
            # Return the reference version if available
            if clean_name in reference_names:
                return reference_names[clean_name]
                
            # Otherwise, return a clean version without accents
            return remove_accents(name)
        
        # Read the latest data from the sheets
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH, keep_vba=True)
        
        # Get the Salaries sheet as reference if available
        reference_names = {}
        if "Salaries" in wb.sheetnames:
            ws_salaries = wb["Salaries"]
            name_col = None
            
            # Find name column in Salaries
            for i, cell in enumerate(ws_salaries[1]):
                if cell.value == "Name":
                    name_col = i + 1
                    break
            
            if name_col:
                # Build reference dictionary from Salaries
                for i in range(2, ws_salaries.max_row + 1):
                    name_cell = ws_salaries.cell(row=i, column=name_col)
                    if name_cell.value and isinstance(name_cell.value, str):
                        # Create a clean version as the key
                        clean_name = remove_accents(name_cell.value.lower())
                        reference_names[clean_name] = name_cell.value
        
        # Update all sheets
        sheets_to_update = ["Pitcher", "Hitter", "FGHitters", "FGPitchers", 
                           "FGHittersL7", "FGPitchersL30", "FGHittersL3Yrs", "FGPitchersL3Yrs"]
        
        # Get reference data for comprehensive standardization
        reference_df = None
        if "Salaries" in wb.sheetnames:
            # Convert the Salaries sheet to a DataFrame
            ws_salaries = wb["Salaries"]
            headers = [cell.value for cell in ws_salaries[1]]
            data = []
            for row in range(2, ws_salaries.max_row + 1):
                row_data = [ws_salaries.cell(row=row, column=col).value for col in range(1, len(headers) + 1)]
                data.append(row_data)
            reference_df = pd.DataFrame(data, columns=headers)
            logger.info(f"Created reference DataFrame from Salaries sheet with {len(reference_df)} rows")
        
        for sheet_name in sheets_to_update:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                name_col = None
                
                # Find name column
                for i, cell in enumerate(ws[1]):
                    if cell.value == "Name":
                        name_col = i + 1
                        break
                
                if name_col:
                    # Extract the sheet data to a DataFrame for better processing
                    headers = [cell.value for cell in ws[1]]
                    data = []
                    for row in range(2, ws.max_row + 1):
                        row_data = [ws.cell(row=row, column=col).value for col in range(1, len(headers) + 1)]
                        data.append(row_data)
                    
                    sheet_df = pd.DataFrame(data, columns=headers)
                    
                    # Apply comprehensive standardization if we have a reference
                    if reference_df is not None and 'Name' in sheet_df.columns:
                        original_names = sheet_df['Name'].copy()
                        sheet_df = standardize_all_player_names(sheet_df, reference_df)
                        changes = sum(original_names != sheet_df['Name'])
                        logger.info(f"Made {changes} name updates in {sheet_name} sheet using comprehensive standardization")
                        
                        # Write back the standardized names
                        for i, name in enumerate(sheet_df['Name']):
                            row_num = i + 2  # +2 because we start at row 2 (after header)
                            if row_num <= ws.max_row:
                                ws.cell(row=row_num, column=name_col).value = name
                    else:
                        # Fall back to simple standardization if no reference is available
                        changes = 0
                        for i in range(2, ws.max_row + 1):
                            name_cell = ws.cell(row=i, column=name_col)
                            original = name_cell.value
                            
                            if original and isinstance(original, str):
                                standardized = standardize_name(original)
                                if standardized != original:
                                    name_cell.value = standardized
                                    changes += 1
                        
                        logger.info(f"Made {changes} name updates in {sheet_name} sheet using simple standardization")
        
        # Save the workbook
        wb.save(EXCEL_FILE_PATH)
        logger.info("All sheets synchronized for name compatibility")
        
        return True
    
    except Exception as e:
        logger.error(f"Error synchronizing datasets: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False
    
def process_ranged_stats(days=7, player_type="hitters", reference_df=None):
    """
    Process stats retrieved using date range functions with unified name standardization
    
    Parameters:
    days (int): Number of days to look back
    player_type (str): "hitters" or "pitchers"
    reference_df (pandas DataFrame): Reference DataFrame with correct names
    
    Returns:
    pandas DataFrame: Stats data with standardized names
    """
    logger.info(f"Getting last {days} days {player_type} data from range functions...")
    
    end_date = datetime.now().strftime("%Y-%m-%d")
    start_date = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    
    try:
        if player_type == "hitters":
            df = batting_stats_range(start_date, end_date)
        else:
            df = pitching_stats_range(start_date, end_date)
        
        logger.info(f"Successfully retrieved {len(df)} {player_type} records for last {days} days")
        
        # Apply name standardization approach
        if not df.empty and 'Name' in df.columns:
            # Use the standardization function directly
            df = standardize_all_player_names(df, reference_df)
            logger.info(f"Successfully standardized {len(df)} player names using unified approach")
        
        return df
    
    except Exception as e:
        logger.error(f"Error getting last {days} days {player_type} data: {str(e)}")
        return pd.DataFrame()

def update_park_factors():
    """
    Production-ready function based on the exact approach that worked in diagnostic testing.
    Uses direct cell assignment with openpyxl exactly as in the diagnostic version.
    """
    logger.info("Starting park factors update with production version of working approach...")
    
    try:
        # First get the 3-year rolling park factors (for the 28 permanent stadiums)
        url_3yr = "https://baseballsavant.mlb.com/leaderboard/statcast-park-factors"
        
        # Then get the 1-year park factors for the 2 temporary stadiums, Rays and Athletics
        url_1yr = "https://baseballsavant.mlb.com/leaderboard/statcast-park-factors?type=year&year=2025&batSide=&stat=index_wOBA&condition=All&rolling=1&parks=all"
        
        # Setup headless Chrome browser
        logger.info("Setting up Chrome browser...")
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        logger.info("Chrome browser setup completed")
        
        # Define exact team code mapping - using direct, specific mappings
        team_codes = {
            # Team names to codes (exact matches)
            'Rockies': 'COL',
            'Red Sox': 'BOS',
            'Royals': 'KC', 
            'D-backs': 'ARI',
            'Reds': 'CIN',
            'Twins': 'MIN',
            'Marlins': 'MIA',
            'Yankees': 'NYY',
            'Nationals': 'WSH',
            'Braves': 'ATL',
            'Orioles': 'BAL',
            'Phillies': 'PHI',
            'Astros': 'HOU',
            'Pirates': 'PIT',
            'Angels': 'LAA',
            'Cardinals': 'STL',
            'Dodgers': 'LAD',
            'Rangers': 'TEX',
            'Blue Jays': 'TOR',
            'Padres': 'SD',
            'White Sox': 'CWS',
            'Tigers': 'DET',
            'Guardians': 'CLE',
            'Mets': 'NYM',
            'Cubs': 'CHC',
            'Brewers': 'MIL',
            'Giants': 'SF',
            'Mariners': 'SEA',
            'Athletics': 'OAK',
            'Rays': 'TB'
        }
        
        # Process 3-year data first
        logger.info(f"Navigating to 3-year rolling park factors URL")
        driver.get(url_3yr)
        logger.info("Waiting 5 seconds for page to load...")
        time.sleep(5)
        
        # Get 3-year data using the exact working approach
        logger.info("Starting to extract 3-year park factors data")
        
        # Create a simple dataframe with all the desired columns
        team_data = []
        
        try:
            # Use pandas to extract tables from HTML
            tables = pd.read_html(StringIO(driver.page_source))
            logger.info(f"Found {len(tables)} tables on the 3-year rolling page")
            
            # Find the correct table with park factors
            for i, table in enumerate(tables):
                # Check if this table has the park factors data we need
                cols = [str(col) for col in table.columns]
                if 'Team' in cols and 'Venue' in cols and any('Park Factor' in col for col in cols):
                    logger.info(f"Found park factors table at index {i}")
                    logger.info(f"Available columns: {cols}")
                    
                    # Process each row with desired columns
                    for _, row in table.iterrows():
                        team_name = row['Team']
                        team_code = team_codes.get(team_name, "---")
                        
                        # Create data dictionary with desired columns
                        team_info = {
                            'TeamCode': team_code,
                            'Team': team_name,
                            'Venue': row['Venue'],
                            'ParkFactor': row['Park Factor'],
                            'R': row['R'],
                            'HR': row['HR'],
                            'Source': '3-year rolling'
                        }
                        
                        # Add these columns if they exist in the table
                        if 'OBP' in cols:
                            team_info['OBP'] = row['OBP']
                        if 'wOBACon' in cols:
                            team_info['wOBACon'] = row['wOBACon']
                        if '2B' in cols:
                            team_info['2B'] = row['2B']
                        if '3B' in cols:
                            team_info['3B'] = row['3B']
                        if 'BB' in cols:
                            team_info['BB'] = row['BB']
                        if 'SO' in cols:
                            team_info['SO'] = row['SO']
                        if 'xwOBACon' in cols:
                            team_info['xwOBACon'] = row['xwOBACon']
                        if 'BACON' in cols:
                            team_info['BACON'] = row['BACON']
                        if 'xBACON' in cols:
                            team_info['xBACON'] = row['xBACON']
                        if 'HardHit' in cols:
                            team_info['HardHit'] = row['HardHit']
                        
                        team_data.append(team_info)
                    
                    logger.info(f"Extracted {len(team_data)} teams from 3-year data")
                    break
        
        except Exception as e:
            logger.error(f"Error in 3-year data extraction: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
        
        # Get temporary stadiums with same approach
        try:
            logger.info(f"Getting temporary stadiums data")
            driver.get(url_1yr)
            time.sleep(5)
            
            tables_1yr = pd.read_html(StringIO(driver.page_source))
            logger.info(f"Found {len(tables_1yr)} tables on 1-year page")
            
            for i, table in enumerate(tables_1yr):
                cols = [str(col) for col in table.columns]
                if 'Team' in cols and 'Venue' in cols and any('Park Factor' in col for col in cols):
                    logger.info(f"Found 1-year table at index {i}")
                    
                    # Find temporary stadiums
                    for idx, row in table.iterrows():
                        team_name = row['Team']
                        venue = row['Venue']
                        
                        # Check if this is a temporary stadium
                        is_temp = False
                        team_code = None
                        
                        if ('Oakland' in str(team_name) or 'Athletics' in str(team_name) or 
                            'Sutter' in str(venue) or 'Sacramento' in str(venue)):
                            team_code = 'OAK'
                            is_temp = True
                            logger.info(f"Found temporary stadium: {venue} for Athletics")
                        elif ('Tampa' in str(team_name) or 'Rays' in str(team_name) or 
                              'Steinbrenner' in str(venue)):
                            team_code = 'TB'
                            is_temp = True
                            logger.info(f"Found temporary stadium: {venue} for Rays")
                        
                        if is_temp and team_code:
                            # Remove existing team data for this code
                            team_data = [d for d in team_data if d['TeamCode'] != team_code]
                            
                            # Create data for temporary stadium
                            temp_info = {
                                'TeamCode': team_code,
                                'Team': team_name,
                                'Venue': venue,
                                'ParkFactor': row['Park Factor'],
                                'R': row['R'],
                                'HR': row['HR'],
                                'Source': '1-year temporary'
                            }
                            
                            # Add additional columns if they exist
                            if 'OBP' in cols:
                                temp_info['OBP'] = row['OBP']
                            if 'wOBACon' in cols:
                                temp_info['wOBACon'] = row['wOBACon']
                            if '2B' in cols:
                                temp_info['2B'] = row['2B']
                            if '3B' in cols:
                                temp_info['3B'] = row['3B']
                            if 'BB' in cols:
                                temp_info['BB'] = row['BB']
                            if 'SO' in cols:
                                temp_info['SO'] = row['SO']
                            if 'xwOBACon' in cols:
                                temp_info['xwOBACon'] = row['xwOBACon']
                            if 'BACON' in cols:
                                temp_info['BACON'] = row['BACON']
                            if 'xBACON' in cols:
                                temp_info['xBACON'] = row['xBACON']
                            if 'HardHit' in cols:
                                temp_info['HardHit'] = row['HardHit']
                            
                            team_data.append(temp_info)
                    break
        
        except Exception as e:
            logger.error(f"Error getting temporary stadiums: {str(e)}")
        
        finally:
            # Always close the driver
            if 'driver' in locals() and driver:
                logger.info("Closing Chrome driver")
                driver.quit()
        
        # Write directly to Excel using openpyxl
        if team_data:
            logger.info(f"Writing {len(team_data)} teams to Excel")
            
            # Create a backup first
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(BACKUP_FOLDER, f"MLBProjections_parks_backup_{timestamp}.xlsm")
                
            if os.path.exists(EXCEL_FILE_PATH):
                shutil.copy2(EXCEL_FILE_PATH, backup_path)
                logger.info(f"Created backup at {backup_path}")
            
            # Load the workbook
            wb = openpyxl.load_workbook(EXCEL_FILE_PATH, keep_vba=True)
            
            # Determine which columns to include
            # First get a set of all keys from all dictionaries
            all_keys = set()
            for team in team_data:
                all_keys.update(team.keys())
            
            # Define column order - ensure TeamCode is first and common stats follow
            column_order = ['TeamCode', 'Team', 'Venue', 'ParkFactor', 'R', 'HR']
            
            # Add any additional columns found in data
            for key in sorted(all_keys):
                if key not in column_order and key != 'Source':  # Add all except Source
                    column_order.append(key)
            
            # Add Source as the last column
            column_order.append('Source')
            
            logger.info(f"Columns to be written: {column_order}")
            
            # Remove existing Parks sheet if it exists
            if "Parks" in wb.sheetnames:
                logger.info("Removing existing Parks sheet")
                sheet = wb["Parks"]
                wb.remove(sheet)
                logger.info("Existing Parks sheet removed")
            
            # Create a new Parks sheet
            parks_sheet = wb.create_sheet("Parks")
            logger.info("Created new Parks sheet")
            
            # Write headers explicitly
            for col_idx, header in enumerate(column_order, start=1):
                parks_sheet.cell(row=1, column=col_idx).value = header
            logger.info(f"Wrote {len(column_order)} column headers")
            
            # Write data with explicit cell assignments - exactly as in the working diagnostic
            for row_idx, team in enumerate(team_data, start=2):
                for col_idx, column in enumerate(column_order, start=1):
                    # Get value with empty string as fallback
                    value = team.get(column, '')
                    parks_sheet.cell(row=row_idx, column=col_idx).value = value
            
            logger.info(f"Wrote data for {len(team_data)} teams")
            
            # Log some samples for verification
            logger.info("Sample of first 3 rows that will be written:")
            for i, team in enumerate(team_data[:3]):
                logger.info(f"Team {i+1}: {team['TeamCode']} - {team['Team']}")
                logger.info(f"  ParkFactor: {team.get('ParkFactor')}")
                logger.info(f"  R: {team.get('R')}")
                logger.info(f"  HR: {team.get('HR')}")
            
            # Save the workbook
            logger.info("Saving updated workbook")
            try:
                wb.save(EXCEL_FILE_PATH)
                logger.info("Workbook saved successfully")
                
                # Verify the saved data
                try:
                    verify_wb = openpyxl.load_workbook(EXCEL_FILE_PATH, read_only=True)
                    if "Parks" in verify_wb.sheetnames:
                        verify_sheet = verify_wb["Parks"]
                        # Get dimensions
                        max_row = verify_sheet.max_row
                        max_col = verify_sheet.max_column
                        logger.info(f"Verification: Parks sheet has {max_row} rows and {max_col} columns")
                        
                        # Read headers
                        headers = []
                        for col in range(1, max_col + 1):
                            cell_value = verify_sheet.cell(row=1, column=col).value
                            headers.append(cell_value)
                        
                        logger.info(f"Verification: Headers in Parks sheet: {headers}")
                    else:
                        logger.warning("Verification: Parks sheet not found after save")
                except Exception as verify_e:
                    logger.error(f"Error in verification: {str(verify_e)}")
                
            except Exception as save_e:
                logger.error(f"Error saving workbook: {str(save_e)}")
                import traceback
                logger.error(traceback.format_exc())
                return False
            
            logger.info("Parks sheet update completed successfully")
            return True
        else:
            logger.warning("No park factors data was collected, skipping update")
            return False
    
    except Exception as e:
        logger.error(f"Error in update_park_factors: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False
    
def get_team_batting_stats(year=None):
    """
    Get team batting stats from FanGraphs using the pybaseball library
    
    Parameters:
    year (int, optional): Year to get data for. If None, gets current year data.
    
    Returns:
    pandas DataFrame: Team batting stats with DK_Points calculations
    """
    logger = logging.getLogger("MLB Updater")
    logger.info(f"Getting team batting stats for year: {year or 'current'}")
    
    try:
        # If year is not provided, use current year
        if year is None:
            year = pd.Timestamp.now().year
        
        # Get team batting stats from FanGraphs
        teams_df = team_batting(year, year)
        
        # Log data fetch success
        logger.info(f"Successfully retrieved batting stats for {len(teams_df)} teams")
        
        # Standardize team codes (use the same mapping as in the main code)
        team_code_map = {
            "WSN": "WSH",
            "TBR": "TB",
            "SDP": "SD",
            "KCR": "KC",
            "CHW": "CWS",
            "SFG": "SF"
        }
        
        # Extract team code from the Team column - usually it's in the format "Team_name Team_code"
        if 'Team' in teams_df.columns:
            teams_df['TeamCode'] = teams_df['Team'].apply(lambda x: x.split()[-1] if isinstance(x, str) and len(x.split()) > 1 else x)
            # Apply the team code mapping
            teams_df['TeamCode'] = teams_df['TeamCode'].map(lambda x: team_code_map.get(x, x))
        
        # Calculate DK points for offense
        teams_df = calculate_team_dk_points(teams_df)
        
        return teams_df
    
    except Exception as e:
        logger.error(f"Error getting team batting stats: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return pd.DataFrame()

def calculate_team_dk_points(teams_df):
    """
    Calculate DraftKings fantasy points for team offense
    
    Parameters:
    teams_df (pandas DataFrame): Team batting stats
    
    Returns:
    pandas DataFrame: Original DataFrame with DK_Points column added
    """
    logger = logging.getLogger("MLB Updater")
    
    if teams_df.empty:
        return teams_df
    
    try:
        # Make a copy to avoid overwriting the original
        result_df = teams_df.copy()
        
        # Check if required columns exist
        required_cols = ['H', '2B', '3B', 'HR', 'RBI', 'R', 'BB', 'HBP', 'SB', 'G']
        missing_cols = [col for col in required_cols if col not in result_df.columns]
        
        if missing_cols:
            logger.warning(f"Missing columns for DK points calculation: {missing_cols}")
            # Try to find alternative column names
            col_mapping = {
                'H': ['H', 'Hits'],
                '2B': ['2B', 'Doubles', '2B'],
                '3B': ['3B', 'Triples', '3B'],
                'HR': ['HR', 'HomeRuns', 'Home Runs'],
                'RBI': ['RBI', 'Runs Batted In'],
                'R': ['R', 'Runs'],
                'BB': ['BB', 'Walks', 'Base on Balls'],
                'HBP': ['HBP', 'Hit By Pitch'],
                'SB': ['SB', 'Stolen Bases'],
                'G': ['G', 'Games', 'GP']
            }
            
            # Try to map columns
            for missing_col in missing_cols:
                for alt_col in col_mapping.get(missing_col, []):
                    if alt_col in result_df.columns:
                        result_df[missing_col] = result_df[alt_col]
                        logger.info(f"Mapped column {missing_col} to {alt_col}")
                        break
        
        # Calculate Singles (H - 2B - 3B - HR)
        if all(col in result_df.columns for col in ['H', '2B', '3B', 'HR']):
            result_df['Singles'] = result_df['H'] - result_df['2B'] - result_df['3B'] - result_df['HR']
        else:
            result_df['Singles'] = 0
            logger.warning("Could not calculate Singles due to missing columns")
        
        # DK points formula for offense
        # Singles = 3 pts, Doubles = 5 pts, Triples = 8 pts, HR = 10 pts, RBI = 2 pts, 
        # R = 2 pts, BB = 2 pts, HBP = 2 pts, SB = 5 pts
        
        # Safe get method to handle missing/renamed columns
        def safe_get(df, col, default=0):
            if col in df.columns:
                return df[col].fillna(0)
            return pd.Series(default, index=df.index)
        
        # Calculate total team points
        result_df['TeamPoints'] = (
            safe_get(result_df, 'Singles') * 3 +
            safe_get(result_df, '2B') * 5 +
            safe_get(result_df, '3B') * 8 +
            safe_get(result_df, 'HR') * 10 +
            safe_get(result_df, 'RBI') * 2 +
            safe_get(result_df, 'R') * 2 +
            safe_get(result_df, 'BB') * 2 +
            safe_get(result_df, 'HBP') * 2 +
            safe_get(result_df, 'SB') * 5
        )
        
        # Calculate points per game
        games = safe_get(result_df, 'G', 1)  # Default to 1 to avoid division by zero
        result_df['DK_Points_Per_Game'] = result_df['TeamPoints'] / games
        
        # Calculate standardized offense score (higher = better offense)
        mean_points = result_df['DK_Points_Per_Game'].mean()
        std_points = result_df['DK_Points_Per_Game'].std()
        
        if std_points > 0:
            result_df['OffenseScore'] = (result_df['DK_Points_Per_Game'] - mean_points) / std_points
        else:
            result_df['OffenseScore'] = 0
            
        logger.info(f"Calculated DK points for {len(result_df)} teams")
        
        return result_df
    
    except Exception as e:
        logger.error(f"Error calculating team DK points: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return teams_df


def update_team_hitting_sheet(excel_path, team_stats_df, sheet_name="FGTmHitting"):
    """
    Update the FGTmHitting sheet in the Excel file with team batting stats
    
    Parameters:
    excel_path (str): Path to the Excel file
    team_stats_df (pandas DataFrame): DataFrame with team offensive stats
    sheet_name (str): Name of the sheet to update
    
    Returns:
    bool: True if successful, False otherwise
    """
    logger = logging.getLogger("MLB Updater")
    logger.info(f"Updating {sheet_name} sheet with team batting stats")
    
    if team_stats_df.empty:
        logger.warning("Cannot update team batting stats: empty DataFrame")
        return False
    
    try:
        
        # Load workbook
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        
        # Check if sheet exists
        if sheet_name in wb.sheetnames:
            # Remove existing sheet
            wb.remove(wb[sheet_name])
            logger.info(f"Removed existing {sheet_name} sheet")
        
        # Create new sheet
        ws = wb.create_sheet(sheet_name)
        
        # Convert DataFrame to rows
        rows = dataframe_to_rows(team_stats_df, index=False, header=True)
        
        # Write rows to the sheet
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Save workbook
        wb.save(excel_path)
        
        logger.info(f"Successfully updated {sheet_name} sheet with {len(team_stats_df)} rows")
        return True
    
    except Exception as e:
        logger.error(f"Error updating {sheet_name} sheet: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def calculate_advanced_pitch_weight(pitchers_df, output_col="PitchWeight", training_df=None):
    """
    Calculate an advanced pitch weight metric using a machine learning approach
    
    This function analyzes pitcher performance metrics to create a more
    sophisticated pitch weight value that better predicts fantasy performance.
    
    Parameters:
    pitchers_df (pandas DataFrame): DataFrame containing all pitcher data
    output_col (str): Name of the output column for pitch weight
    training_df (pandas DataFrame, optional): DataFrame containing only starting pitchers for ML training
    
    Returns:
    pandas DataFrame: Original DataFrame with pitch weight column added
    """
    
    logger.info("Calculating advanced pitch weight using machine learning approach...")
    
    # Make a copy to avoid overwriting the original
    result_df = pitchers_df.copy()
    
    # Use filtered pitchers for ML if provided, otherwise use all pitchers
    if training_df is not None and not training_df.empty:
        logger.info(f"Using training dataset with {len(training_df)} pitchers")
    else:
        logger.info("No filtered training dataset provided, using all pitchers")
        training_df = result_df
    
    # log data structure
    logger.info(f"DataFrame shape: {result_df.shape}")
    logger.info(f"Training DataFrame shape: {training_df.shape}")
    logger.info(f"DataFrame dtypes: {result_df.dtypes}")
    
    # Sample first row for debugging
    if len(training_df) > 0:
        sample_row = training_df.iloc[0]
        logger.info(f"First training row Name: {sample_row.get('Name')}")
        logger.info(f"First training row DK_Points L3Yrs: {sample_row.get('DK_Points L3Yrs')}")
        logger.info(f"First training row K% L3yrs: {sample_row.get('K% L3yrs')}")
    
    # Check for column existence
    dk_points_cols = [col for col in training_df.columns if 'DK_Points' in str(col)]
    logger.info(f"DK_Points columns found: {dk_points_cols}")
    
    # Define target variables - fantasy points 
    target_cols = [col for col in dk_points_cols if col in training_df.columns]
    logger.info(f"Target columns to use: {target_cols}")
    
    # Only keep rows with sufficient data - more lenient criteria
    if 'Name' in training_df.columns:
        # Count rows with name and at least one DK_Points value
        valid_count = 0
        valid_rows_idx = []
        for idx, row in training_df.iterrows():
            if pd.notna(row.get('Name')):
                has_target = False
                for col in target_cols:
                    if pd.notna(row.get(col)) and row.get(col) != 0:
                        has_target = True
                        break
                if has_target:
                    valid_count += 1
                    valid_rows_idx.append(idx)
        
        logger.info(f"Found {valid_count} rows with Name and at least one valid target value")
        
        # Lower threshold for ML to 50 valid rows
        if valid_count >= 50:
            try:
                # Use only valid rows for ML
                valid_rows = training_df.loc[valid_rows_idx].copy()
                
                # Define features for ML - use what's available
                feature_cols = []
                for col_prefix in ['K%', 'BB%', 'xFIP', 'BABIP', 'GB%', 'FB%', 'Hard%', 'Soft%', 'DK/IP']:
                    for year_suffix in [' 2025', ' L3yrs', ' L30']:
                        col = f"{col_prefix}{year_suffix}"
                        if col in training_df.columns and training_df[col].count() > 50:
                            feature_cols.append(col)
                
                logger.info(f"Using {len(feature_cols)} feature columns: {feature_cols}")
                
                # Prepare features, handling missing values
                X = valid_rows[feature_cols].copy()
                X = X.fillna(X.mean())  # Fill NAs with mean
                
                # Create target - weighted average of available targets 
                y = pd.Series(0.0, index=valid_rows.index)
                weights = {'DK_Points L30': 0.2, 'DK_Points 2025': 0.4, 'DK_Points L3Yrs': 0.4}
                
                target_counts = {}
                for col in target_cols:
                    target_counts[col] = valid_rows[col].count()
                logger.info(f"Target column counts: {target_counts}")
                
                # Use weighted average of available targets
                for col, weight in weights.items():
                    if col in valid_rows.columns:
                        # Only add if column has values
                        if valid_rows[col].count() > 0:
                            y += valid_rows[col].fillna(0) * weight
                
                # Normalize features for better model performance
                scaler = StandardScaler()
                X_scaled = scaler.fit_transform(X)
                
                # Train a Ridge regression model (regularized to prevent overfitting)
                model = Ridge(alpha=1.0)
                model.fit(X_scaled, y)
                
                # Calculate feature importance
                importance = np.abs(model.coef_)
                importance_dict = dict(zip(feature_cols, importance))
                
                # Log top 10 most important features
                top_features = sorted(importance_dict.items(), key=lambda x: x[1], reverse=True)[:5]
                logger.info("Top 5 predictive features for pitch weight:")
                for feature, importance in top_features:
                    logger.info(f"  {feature}: {importance:.4f}")
                
                # Apply model to all rows in the full dataset (will use means for NA values)
                X_all = result_df[feature_cols].copy()
                X_all = X_all.fillna(X_all.mean())
                X_all_scaled = scaler.transform(X_all)
                
                # Generate pitch weight predictions
                raw_weights = model.predict(X_all_scaled)
                
                # Normalize to useful range (5-15 is a good baseball range)
                min_weight, max_weight = 5, 15
                normalized_weights = min_weight + (max_weight - min_weight) * (raw_weights - np.min(raw_weights)) / (np.max(raw_weights) - np.min(raw_weights))
                
                # Ensure no negative values
                normalized_weights = np.maximum(normalized_weights, 0)
                
                # For pitchers with ALL missing data, set weight to 7.5
                unknown_pitchers = []
                for idx, row in result_df.iterrows():
                    # Check if this pitcher has ALL NAs in the feature columns
                    all_missing = True
                    for col in feature_cols:
                        if pd.notna(row.get(col)):
                            all_missing = False
                            break
                    
                    if all_missing and pd.notna(row.get('Name')):
                        unknown_pitchers.append(idx)
                        normalized_weights[idx] = 7.5  # Set to our default for unknown pitchers

                if unknown_pitchers:
                    unknown_count = len(unknown_pitchers)
                    logger.info(f"Set weight to 7.5 for {unknown_count} pitchers with no stats")
                    if unknown_count <= 10:
                        for idx in unknown_pitchers:
                            logger.info(f"  Unknown pitcher: {result_df.iloc[idx]['Name']}")
                
                # Add to result dataframe
                result_df[output_col] = normalized_weights
                
                # Calculate additional stats for validation
                logger.info(f"ML-based pitch weight stats: min={np.min(normalized_weights):.2f}, max={np.max(normalized_weights):.2f}, mean={np.mean(normalized_weights):.2f}")
                
                # Show top 10 pitchers by weight
                top_indices = np.argsort(normalized_weights)[-5:][::-1]
                top_names = result_df.iloc[top_indices]['Name'].tolist()
                top_weights = normalized_weights[top_indices]
                logger.info("Top 5 pitchers by ML weight:")
                for name, weight in zip(top_names, top_weights):
                    logger.info(f"  {name}: {weight:.2f}")
                
                logger.info(f"Successfully calculated ML-based pitch weight for {len(result_df)} pitchers")
                return result_df
                
            except Exception as e:
                logger.error(f"Error in ML pitch weight calculation: {str(e)}")
                logger.info("Falling back to rules-based pitch weight calculation")
                import traceback
                logger.error(traceback.format_exc())
    
    # Fall back to rules-based approach if ML fails or not enough data
    logger.warning(f"Not enough valid data rows for ML approach. Using rules-based calculation instead.")
    result_df[output_col] = rules_based_pitch_weight(result_df)
    return result_df


def add_pitch_weight_to_excel(excel_path, sheet_name="Pitcher", filtered_pitcher_df=None):
    """
    Add the advanced pitch weight calculation to the Excel file
    
    Parameters:
    excel_path (str): Path to the Excel file
    sheet_name (str): Name of the sheet to update
    filtered_pitcher_df (pandas DataFrame, optional): Filtered DataFrame containing only starting pitchers
    
    Returns:
    bool: True if successful, False otherwise
    """
    logger.info(f"Adding pitch weight to {sheet_name} sheet in {excel_path}")
    
    try:
        # Use xlwings to read the pitcher data with formula evaluation
        logger.info("Using xlwings to read data with formula evaluation...")
        try:
            app = xw.App(visible=False)
            wb = app.books.open(excel_path)
            sheet = wb.sheets[sheet_name]
            
            # Get all data including headers
            data_range = sheet.used_range
            raw_data = data_range.value
            
            # Convert to pandas DataFrame
            headers = raw_data[0]
            data = raw_data[1:]
            pitcher_df = pd.DataFrame(data, columns=headers)
            
            logger.info(f"Successfully read {len(pitcher_df)} rows with xlwings (formulas evaluated)")
            
            # Close Excel
            wb.close()
            app.quit()
            
        except Exception as xlw_e:
            # Fallback to pandas if xlwings fails
            logger.warning(f"xlwings read failed: {str(xlw_e)}. Falling back to pandas...")
            pitcher_df = pd.read_excel(excel_path, sheet_name=sheet_name)
            logger.info(f"Read {len(pitcher_df)} pitchers using pandas (formulas not evaluated)")
        
        # Debug
        logger.info("Diagnosing data access issues:")
        sample_cols = ['Name', 'K% L3yrs', 'BB% L3Yrs', 'DK_Points L3Yrs']
        available_cols = [col for col in sample_cols if col in pitcher_df.columns]
        
        # Print all column headers to verify exact names
        logger.info(f"Actual column headers from Excel: {list(pitcher_df.columns)}")
        
        # Check for null vs missing values
        for col in available_cols:
            null_count = pitcher_df[col].isnull().sum()
            total = len(pitcher_df)
            logger.info(f"Column '{col}': {total-null_count} values, {null_count} nulls")
        
        # Show some sample data
        sample_rows = min(5, len(pitcher_df))
        if sample_rows > 0:
            logger.info(f"First {sample_rows} rows of key data:")
            for i in range(sample_rows):
                row_data = {}
                for col in available_cols:
                    row_data[col] = pitcher_df.iloc[i].get(col, 'N/A')
                logger.info(f"  Row {i+1}: {row_data}")
        
        # Create training DataFrame filtered to only starters
        if filtered_pitcher_df is not None and not filtered_pitcher_df.empty:
            logger.info(f"Using filtered dataset with {len(filtered_pitcher_df)} starting pitchers for ML training")
            
            # Get the set of starter names for faster lookups
            starter_names = set(filtered_pitcher_df['Name'].values)
            
            # Filter pitcher_df to include only starters for ML training
            training_df = pitcher_df[pitcher_df['Name'].isin(starter_names)].copy()
            
            logger.info(f"Created training dataset with {len(training_df)} starting pitchers")
            
            # Use starters for training but apply weights to ALL pitchers
            pitcher_df = calculate_advanced_pitch_weight(pitcher_df, "PitchWeight", training_df)
        else:
            logger.warning("No filtered pitchers provided, using all pitchers for training")
            pitcher_df = calculate_advanced_pitch_weight(pitcher_df, "PitchWeight")
        
        # Create backup
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(BACKUP_FOLDER, 
                             f"MLBProjections_backup_before_pitch_weight_{timestamp}.xlsm")
        shutil.copy2(excel_path, backup_path)
        logger.info(f"Created backup at {backup_path}")
        
        # Open the workbook with openpyxl to modify without losing formulas/formatting
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        ws = wb[sheet_name]
        
        # Check if PitchWeight column already exists
        pitch_weight_col = None
        for i, cell in enumerate(ws[1]):
            if cell.value == "PitchWeight":
                pitch_weight_col = i + 1
                logger.info(f"Found existing PitchWeight column at position {pitch_weight_col}")
                break
        
        # If PitchWeight column doesn't exist, add it at the end
        if pitch_weight_col is None:
            # Find last column and add header for PitchWeight
            last_col = 0
            for i, cell in enumerate(ws[1]):
                if cell.value:
                    last_col = i + 1
            
            # Add header for the new column
            pitch_weight_col = last_col + 1
            ws.cell(row=1, column=pitch_weight_col).value = "PitchWeight"
            logger.info(f"Created new PitchWeight column at position {pitch_weight_col}")
        
        # Add pitch weight values
        for i, (_, row) in enumerate(pitcher_df.iterrows(), start=2):
            if i <= ws.max_row:  # Make sure we don't go beyond existing rows
                ws.cell(row=i, column=pitch_weight_col).value = row['PitchWeight']
        
        # Save the workbook
        wb.save(excel_path)
        logger.info(f"Added/updated PitchWeight column in '{sheet_name}' sheet")
        
        # Log sample pitch weights
        samples = pitcher_df.sort_values('PitchWeight', ascending=False).head(5)
        logger.info("Sample top 5 pitch weights:")
        for _, row in samples.iterrows():
            logger.info(f"  {row['Name']}: {row['PitchWeight']:.2f}")
        
        return True
    
    except Exception as e:
        logger.error(f"Error adding pitch weight to Excel: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def rules_based_pitch_weight(df):
    """
    Calculate pitch weight using a hard coded approach
    Uses available statistics to determine pitcher quality without arbitrary lists
    
    Parameters:
    df (pandas DataFrame): DataFrame containing pitcher data
    
    Returns:
    pandas Series: Calculated pitch weights
    """
    logger.info("Using hard coded pitch weight calculation")
    
    # Start with base value
    weights = pd.Series(9.0, index=df.index)
    
    # Process each pitcher
    for idx, row in df.iterrows():
        # Skip rows without a name
        if pd.isna(row.get('Name')):
            continue
        
        # Base score - start at 9.0
        score = 9.0
        adjustments_made = False
        
        # Get all stats with careful error handling
        stats = {}
        
        # Key stats to look for
        stat_keys = {
            'k_pct': ['K% 2025', 'K% L3yrs'],
            'bb_pct': ['BB% 2025', 'BB% L3Yrs'],
            'xfip': ['xFIP 2025', 'XFIP L3Yrs'],
            'dk_ip': ['DK/IP 2025', 'DK/IP L3Yrs', 'DK/IP L30'],
            'dk_points': ['DK_Points 2025', 'DK_Points L3Yrs', 'DK_Points L30'],
            'gb_pct': ['GB% 2025', 'GB% L3Yrs']
        }
        
        # Try to get each stat
        for stat_name, columns in stat_keys.items():
            stats[stat_name] = None
            for col in columns:
                try:
                    if col in row and not pd.isna(row[col]):
                        value = row[col]
                        if isinstance(value, (int, float)) or (isinstance(value, str) and value.replace('.', '', 1).isdigit()):
                            stats[stat_name] = float(value)
                            break
                except Exception as e:
                    logger.debug(f"Error getting {col} for {row.get('Name', 'unknown')}: {str(e)}")
        
        # Check if this is an "unknown" pitcher with no data
        has_data = any(v is not None for v in stats.values())
        if not has_data:
            # No data available at all - likely a call-up with little experience
            score = 7.5  # Start substantially below average for completely unknown pitchers
            weights[idx] = score
            continue  # Skip the rest of processing for this pitcher
            
        # Log stats found
        if idx < 5:  # Log first 5 pitchers for debugging
            logger.info(f"Stats found for {row.get('Name')}: {stats}")
        
        # Apply adjustments based on stats (purely objective criteria)
        
        # 1. Strikeout rate - most predictive stats
        if stats['k_pct'] is not None:
            adjustments_made = True
            k_pct = stats['k_pct']
            if k_pct >= 30:       # Elite (top 1%)
                score += 3.0
            elif k_pct >= 27:     # Excellent (top 5%)
                score += 2.5
            elif k_pct >= 24:     # Very good (top 15%)
                score += 1.5
            elif k_pct >= 22:     # Above average (top 30%)
                score += 0.75
            elif k_pct <= 15:     # Poor (bottom 10%)
                score -= 1.5
            elif k_pct <= 18:     # Below average (bottom 25%)
                score -= 0.75
        
        # 2. Walk rate
        if stats['bb_pct'] is not None:
            adjustments_made = True
            bb_pct = stats['bb_pct']
            if bb_pct <= 4:       # Elite control (top 5%)
                score += 2.0
            elif bb_pct <= 6:     # Excellent control (top 15%)
                score += 1.5
            elif bb_pct <= 7.5:   # Good control (top 30%)
                score += 0.75
            elif bb_pct >= 12:    # Poor control (bottom 10%)
                score -= 1.5
            elif bb_pct >= 10:    # Below average control (bottom 20%)
                score -= 1.0
        
        # 3. xFIP - excellent predictor of ERA
        if stats['xfip'] is not None:
            adjustments_made = True
            xfip = stats['xfip']
            if xfip <= 3.0:       # Elite (top 5%)
                score += 2.5
            elif xfip <= 3.5:     # Excellent (top 15%)
                score += 2.0
            elif xfip <= 3.8:     # Very good (top 30%)
                score += 1.0
            elif xfip >= 5.0:     # Poor (bottom 10%)
                score -= 1.5
            elif xfip >= 4.5:     # Below average (bottom 25%)
                score -= 0.75
        
        # 4. DK points per inning - measure of fantasy value
        if stats['dk_ip'] is not None:
            adjustments_made = True
            dk_ip = stats['dk_ip']
            if dk_ip >= 4.2:      # Elite (top 5%)
                score += 2.5
            elif dk_ip >= 3.8:    # Excellent (top 15%)
                score += 2.0
            elif dk_ip >= 3.5:    # Very good (top 30%)
                score += 1.0
            elif dk_ip <= 2.5:    # Poor (bottom 10%)
                score -= 1.5
            elif dk_ip <= 3.0:    # Below average (bottom 25%)
                score -= 0.75
        
        # 5. Ground ball rate - prevent HRs
        if stats['gb_pct'] is not None:
            adjustments_made = True
            gb_pct = stats['gb_pct']
            if gb_pct >= 55:      # Elite ground ball rate (top 10%)
                score += 1.0
            elif gb_pct >= 50:    # Very good (top 25%)
                score += 0.5
            elif gb_pct <= 35:    # Extreme fly ball (bottom 10%)
                score -= 0.5
        
        # 6. Total fantasy points - volume matters
        if stats['dk_points'] is not None:
            adjustments_made = True
            dk_points = stats['dk_points']
            if dk_points >= 25:   # Ace workload
                score += 1.0
            elif dk_points >= 20: # Strong workload
                score += 0.5
            elif dk_points <= 10: # Limited workload
                score -= 0.5
        
        # If no adjustments were made but we have some data, add small random factor
        if not adjustments_made and has_data:
            import random
            random.seed(hash(str(row.get('Name', idx))))
            score += random.uniform(-0.5, 0.5)
        
        # Ensure score is within bounds
        score = max(5.0, min(15.0, score))
        
        # Store the final weight
        weights[idx] = score
    
    # Ensure we have a good distribution of values
    if len(weights) > 10:
        p10 = weights.quantile(0.1)
        p90 = weights.quantile(0.9)
        
        # If the distribution is too compressed, stretch it
        if p90 - p10 < 4:
            logger.info("Adjusting distribution to ensure better spread of values")
            median = weights.median()
            weights = median + (weights - median) * (6 / max(0.1, p90 - p10))
            weights = weights.clip(5, 15)
    
    # Report some stats about the weights
    logger.info(f"Pitch weight stats: min={weights.min():.2f}, max={weights.max():.2f}, mean={weights.mean():.2f}")
    
    return weights

def download_barrel_data_from_savant(output_folder="barrel_data", create_folder=True, overwrite=True):
    """
    Downloads barrel data CSV files from Baseball Savant using direct URL access
    
    Parameters:
    -----------
    output_folder : str
        Folder to save downloaded CSV files
    create_folder : bool
        Whether to create the output folder if it doesn't exist
    overwrite : bool
        Whether to overwrite existing files (True) or create date-stamped files (False)
    
    Returns:
    --------
    dict
        Dictionary containing paths to downloaded files and success status
    """
    
    logger.info("Starting Baseball Savant CSV download for barrel data")
    
    # Create output folder if needed
    if create_folder and not os.path.exists(output_folder):
        os.makedirs(output_folder)
        logger.info(f"Created output folder: {output_folder}")
    
    # URLs for different datasets - multi-year data (2022-2025)
    urls = {
        # Only fly balls - for specific FB barrel analysis
        "only_fb": "https://baseballsavant.mlb.com/statcast_search/csv?hfPT=&hfAB=&hfGT=R%7C&hfPR=&hfZ=&hfStadium=&hfBBL=&hfNewZones=&hfPull=&hfC=&hfSea=2025%7C2024%7C2023%7C2022%7C&hfSit=&player_type=batter&hfOuts=&hfOpponent=&pitcher_throws=&batter_stands=&hfSA=&game_date_gt=&game_date_lt=&hfMo=&hfTeam=&home_road=&hfRO=&position=&hfInfield=&hfOutfield=&hfInn=&hfBBT=fly%5C.%5C.ball%7C&hfFlag=&metric_1=&group_by=name&min_pitches=0&min_results=0&min_pas=0&sort_col=barrels_per_bbe_percent&player_event_sort=api_h_distance_projected&sort_order=desc",
        
        # Only line drives - for specific LD barrel analysis
        "only_ld": "https://baseballsavant.mlb.com/statcast_search/csv?hfPT=&hfAB=&hfGT=R%7C&hfPR=&hfZ=&hfStadium=&hfBBL=&hfNewZones=&hfPull=&hfC=&hfSea=2025%7C2024%7C2023%7C2022%7C&hfSit=&player_type=batter&hfOuts=&hfOpponent=&pitcher_throws=&batter_stands=&hfSA=&game_date_gt=&game_date_lt=&hfMo=&hfTeam=&home_road=&hfRO=&position=&hfInfield=&hfOutfield=&hfInn=&hfBBT=line%5C.%5C.drive%7C&hfFlag=&metric_1=&group_by=name&min_pitches=0&min_results=0&min_pas=0&sort_col=barrels_per_bbe_percent&player_event_sort=api_h_distance_projected&sort_order=desc",
        
        # Pull FB
        "pull_fb": "https://baseballsavant.mlb.com/statcast_search/csv?hfPT=&hfAB=&hfGT=R%7C&hfPR=&hfZ=&hfStadium=&hfBBL=&hfNewZones=&hfPull=Pull%7C&hfC=&hfSea=2025%7C2024%7C2023%7C2022%7C&hfSit=&player_type=batter&hfOuts=&hfOpponent=&pitcher_throws=&batter_stands=&hfSA=&game_date_gt=&game_date_lt=&hfMo=&hfTeam=&home_road=&hfRO=&position=&hfInfield=&hfOutfield=&hfInn=&hfBBT=fly%5C.%5C.ball%7C&hfFlag=&metric_1=&group_by=name&min_pitches=0&min_results=0&min_pas=0&sort_col=pitches&player_event_sort=api_h_distance_projected&sort_order=desc",
        
        # Pull LD
        "pull_ld": "https://baseballsavant.mlb.com/statcast_search/csv?hfPT=&hfAB=&hfGT=R%7C&hfPR=&hfZ=&hfStadium=&hfBBL=&hfNewZones=&hfPull=Pull%7C&hfC=&hfSea=2025%7C2024%7C2023%7C2022%7C&hfSit=&player_type=batter&hfOuts=&hfOpponent=&pitcher_throws=&batter_stands=&hfSA=&game_date_gt=&game_date_lt=&hfMo=&hfTeam=&home_road=&hfRO=&position=&hfInfield=&hfOutfield=&hfInn=&hfBBT=line%5C.%5C.drive%7C&hfFlag=&metric_1=&group_by=name&min_pitches=0&min_results=0&min_pas=0&sort_col=pitches&player_event_sort=api_h_distance_projected&sort_order=desc",
        
        # Oppo FB
        "oppo_fb": "https://baseballsavant.mlb.com/statcast_search/csv?hfPT=&hfAB=&hfGT=R%7C&hfPR=&hfZ=&hfStadium=&hfBBL=&hfNewZones=&hfPull=Opposite%7C&hfC=&hfSea=2025%7C2024%7C2023%7C2022%7C&hfSit=&player_type=batter&hfOuts=&hfOpponent=&pitcher_throws=&batter_stands=&hfSA=&game_date_gt=&game_date_lt=&hfMo=&hfTeam=&home_road=&hfRO=&position=&hfInfield=&hfOutfield=&hfInn=&hfBBT=fly%5C.%5C.ball%7C&hfFlag=&metric_1=&group_by=name&min_pitches=0&min_results=0&min_pas=0&sort_col=pitches&player_event_sort=api_h_distance_projected&sort_order=desc",
        
        # Oppo LD
        "oppo_ld": "https://baseballsavant.mlb.com/statcast_search/csv?hfPT=&hfAB=&hfGT=R%7C&hfPR=&hfZ=&hfStadium=&hfBBL=&hfNewZones=&hfPull=Opposite%7C&hfC=&hfSea=2025%7C2024%7C2023%7C2022%7C&hfSit=&player_type=batter&hfOuts=&hfOpponent=&pitcher_throws=&batter_stands=&hfSA=&game_date_gt=&game_date_lt=&hfMo=&hfTeam=&home_road=&hfRO=&position=&hfInfield=&hfOutfield=&hfInn=&hfBBT=line%5C.%5C.drive%7C&hfFlag=&metric_1=&group_by=name&min_pitches=0&min_results=0&min_pas=0&sort_col=pitches&player_event_sort=api_h_distance_projected&sort_order=desc",
        
        # All FB
        "all_fb": "https://baseballsavant.mlb.com/statcast_search/csv?hfPT=&hfAB=&hfGT=R%7C&hfPR=&hfZ=&hfStadium=&hfBBL=&hfNewZones=&hfPull=&hfC=&hfSea=2025%7C2024%7C2023%7C2022%7C&hfSit=&player_type=batter&hfOuts=&hfOpponent=&pitcher_throws=&batter_stands=&hfSA=&game_date_gt=&game_date_lt=&hfMo=&hfTeam=&home_road=&hfRO=&position=&hfInfield=&hfOutfield=&hfInn=&hfBBT=fly%5C.%5C.ball%7C&hfFlag=&metric_1=&group_by=name&min_pitches=0&min_results=0&min_pas=0&sort_col=pitches&player_event_sort=api_h_distance_projected&sort_order=desc",
        
        # All LD
        "all_ld": "https://baseballsavant.mlb.com/statcast_search/csv?hfPT=&hfAB=&hfGT=R%7C&hfPR=&hfZ=&hfStadium=&hfBBL=&hfNewZones=&hfPull=&hfC=&hfSea=2025%7C2024%7C2023%7C2022%7C&hfSit=&player_type=batter&hfOuts=&hfOpponent=&pitcher_throws=&batter_stands=&hfSA=&game_date_gt=&game_date_lt=&hfMo=&hfTeam=&home_road=&hfRO=&position=&hfInfield=&hfOutfield=&hfInn=&hfBBT=line%5C.%5C.drive%7C&hfFlag=&metric_1=&group_by=name&min_pitches=0&min_results=0&min_pas=0&sort_col=pitches&player_event_sort=api_h_distance_projected&sort_order=desc"
    }
    
    results = {}
    download_session = requests.Session()  # Use a session for better performance
    
    # Save the passed options
    results['options'] = {
        'overwrite': overwrite
    }
    
    # Download each dataset
    for dataset_name, url in urls.items():
        # Determine output file path based on overwrite parameter
        if overwrite:
            output_file = os.path.join(output_folder, f"{dataset_name}.csv")
            
            # Delete existing file if overwrite is True
            if os.path.exists(output_file):
                try:
                    os.remove(output_file)
                    logger.info(f"Deleted existing file: {output_file}")
                except Exception as del_e:
                    logger.warning(f"Could not delete existing file {output_file}: {str(del_e)}")
        else:
            # Use date-stamped filename
            current_date = datetime.now().strftime("%Y%m%d")
            output_file = os.path.join(output_folder, f"{dataset_name}_{current_date}.csv")
        
        try:
            logger.info(f"Downloading {dataset_name} dataset...")
            response = download_session.get(url, timeout=120)  # Longer timeout for multi-year data
            
            if response.status_code == 200:
                # Check if we got CSV data
                content_type = response.headers.get('Content-Type', '')
                if 'text/csv' in content_type or response.text.strip().startswith("player_name") or ',' in response.text:
                    # Write to file
                    with open(output_file, 'w', encoding='utf-8') as f:
                        f.write(response.text)
                    
                    # Check file size to ensure it's not empty or an error message
                    file_size = os.path.getsize(output_file)
                    if file_size > 100:  # Assume any real data is at least 100 bytes
                        logger.info(f"Successfully downloaded {dataset_name} dataset ({file_size} bytes)")
                        results[dataset_name] = {
                            "success": True,
                            "file_path": output_file,
                            "size": file_size,
                            "rows": response.text.count('\n') - 1  # Approximate row count
                        }
                    else:
                        # Small file might be an error message
                        logger.warning(f"Downloaded file is suspiciously small ({file_size} bytes). Possible error.")
                        results[dataset_name] = {
                            "success": False,
                            "file_path": output_file,
                            "error": "File too small, possible error response"
                        }
                else:
                    logger.error(f"Response is not a CSV: {content_type}")
                    results[dataset_name] = {
                        "success": False,
                        "error": f"Response is not a CSV: {content_type}"
                    }
            else:
                logger.error(f"Failed to download {dataset_name}: Status code {response.status_code}")
                results[dataset_name] = {
                    "success": False,
                    "error": f"HTTP error {response.status_code}"
                }
        
        except Exception as e:
            logger.error(f"Error downloading {dataset_name}: {str(e)}")
            results[dataset_name] = {
                "success": False, 
                "error": str(e)
            }
        
        # Add a delay between requests to avoid rate limiting
        time.sleep(3)
    
    # Process the downloaded data using the correct parameter - just pass the folder
    process_results = process_statcast_barrel_data(output_folder)
    
    return {**results, "processed": process_results}

def process_statcast_barrel_data(output_folder="barrel_data"):
    """
    Process aggregated Statcast data files to extract barrel percentages, distance metrics,
    and raw event counts for FB and LD without calculating percentages.
    
    Parameters:
    -----------
    output_folder : str
        Folder containing aggregated statcast data files
        
    Returns:
    --------
    dict
        Dictionary containing processed statistics
    """
    
    logger.info("Processing aggregated Statcast barrel data files...")
    
    try:
        # Define file paths for all required datasets
        required_files = {
            'only_fb': os.path.join(output_folder, "all_fb.csv"),
            'only_ld': os.path.join(output_folder, "all_ld.csv"),
            'pull_fb': os.path.join(output_folder, "pull_fb.csv"),
            'pull_ld': os.path.join(output_folder, "pull_ld.csv"),
            'oppo_fb': os.path.join(output_folder, "oppo_fb.csv"),
            'oppo_ld': os.path.join(output_folder, "oppo_ld.csv")
        }
        
        # Check if all required files exist
        missing_files = []
        for file_name, file_path in required_files.items():
            if not os.path.exists(file_path):
                missing_files.append(file_name)
        
        if missing_files:
            logger.error(f"Missing required files: {missing_files}")
            return {"success": False, "error": f"Missing required files: {missing_files}"}
        
        # Load all the datasets
        only_fb_file = required_files['only_fb']
        only_ld_file = required_files['only_ld']
        pull_fb_file = required_files['pull_fb']
        pull_ld_file = required_files['pull_ld']
        oppo_fb_file = required_files['oppo_fb']
        oppo_ld_file = required_files['oppo_ld']
        
        # Read all data
        fb_df = pd.read_csv(only_fb_file)
        ld_df = pd.read_csv(only_ld_file)
        pull_fb_df = pd.read_csv(pull_fb_file)
        pull_ld_df = pd.read_csv(pull_ld_file)
        oppo_fb_df = pd.read_csv(oppo_fb_file)
        oppo_ld_df = pd.read_csv(oppo_ld_file)
        
        # For debugging, check column names
        logger.info(f"FB dataset columns: {fb_df.columns[:10].tolist()}")
        logger.info(f"Pull FB dataset columns: {pull_fb_df.columns[:10].tolist()}")
        
        # Find player name column - should be 'player_name' in these datasets
        player_col = 'player_name'
        if player_col not in fb_df.columns:
            # Try to find alternative player column
            player_cols = [col for col in fb_df.columns if 'player' in col.lower()]
            if player_cols:
                player_col = player_cols[0]
            else:
                logger.error("Could not find player name column in the data")
                return {"success": False, "error": "Could not find player name column"}
        
        logger.info(f"Using {player_col} as player name column")
        
        # Find player ID column - NEW ADDITION
        player_id_col = None
        for potential_col in ['player_id', 'mlbam_id', 'mlbam', 'id']:
            if potential_col in fb_df.columns:
                player_id_col = potential_col
                break
                
        if player_id_col:
            logger.info(f"Found player ID column: {player_id_col}")
        else:
            logger.warning("No player ID column found. Duplicate player names may not be handled correctly.")
        
        # Find the barrel percentage column
        barrel_pct_col = None
        for col in fb_df.columns:
            if 'barrels_per_bbe_percent' in col:
                barrel_pct_col = col
                break
                
        if not barrel_pct_col:
            logger.warning("Could not find barrels_per_bbe_percent column in FB data")
            barrel_pct_col = 'barrels_per_bbe_percent'  # Default name
            
        logger.info(f"Using {barrel_pct_col} as barrel percentage column")
        
        # 1. Extract barrel percentages directly from files
        # Extract barrel percentages if the column exists, otherwise use zeros
        if barrel_pct_col in fb_df.columns:
            if player_id_col:
                # Include player ID for more accurate matching
                fb_barrels = fb_df[[player_col, player_id_col, barrel_pct_col]].copy()
            else:
                fb_barrels = fb_df[[player_col, barrel_pct_col]].copy()
            fb_barrels.rename(columns={barrel_pct_col: 'barrels/fb'}, inplace=True)
        else:
            if player_id_col:
                fb_barrels = pd.DataFrame({player_col: fb_df[player_col], player_id_col: fb_df[player_id_col], 'barrels/fb': 0})
            else:
                fb_barrels = pd.DataFrame({player_col: fb_df[player_col], 'barrels/fb': 0})
        
        if barrel_pct_col in ld_df.columns:
            if player_id_col:
                ld_barrels = ld_df[[player_col, player_id_col, barrel_pct_col]].copy()
            else:
                ld_barrels = ld_df[[player_col, barrel_pct_col]].copy()
            ld_barrels.rename(columns={barrel_pct_col: 'barrels/ld'}, inplace=True)
        else:
            if player_id_col:
                ld_barrels = pd.DataFrame({player_col: ld_df[player_col], player_id_col: ld_df[player_id_col], 'barrels/ld': 0})
            else:
                ld_barrels = pd.DataFrame({player_col: ld_df[player_col], 'barrels/ld': 0})
        
        # 2. Extract distance metrics
        # Look for bbdist column in aggregated files
        distance_col = 'bbdist'
        if distance_col in fb_df.columns:
            if player_id_col:
                fb_dist_df = fb_df[[player_col, player_id_col, distance_col]].copy()
            else:
                fb_dist_df = fb_df[[player_col, distance_col]].copy()
            fb_dist_df.rename(columns={distance_col: 'avg_distance_fb'}, inplace=True)
            # Add std_dist_fb - use 15% of avg distance as approximation
            fb_dist_df['std_distance_fb'] = fb_dist_df['avg_distance_fb'] * 0.15
        else:
            # Try to find alternative distance column
            alt_distance_cols = ['hit_distance_sc', 'hit_distance', 'distance']
            found_col = False
            for col in alt_distance_cols:
                if col in fb_df.columns:
                    if player_id_col:
                        fb_dist_df = fb_df[[player_col, player_id_col, col]].copy()
                    else:
                        fb_dist_df = fb_df[[player_col, col]].copy()
                    fb_dist_df.rename(columns={col: 'avg_distance_fb'}, inplace=True)
                    fb_dist_df['std_distance_fb'] = fb_dist_df['avg_distance_fb'] * 0.15
                    found_col = True
                    break
                    
            if not found_col:
                logger.warning("No distance column found in FB data, using zeros")
                if player_id_col:
                    fb_dist_df = pd.DataFrame({
                        player_col: fb_df[player_col],
                        player_id_col: fb_df[player_id_col],
                        'avg_distance_fb': 0,
                        'std_distance_fb': 0
                    })
                else:
                    fb_dist_df = pd.DataFrame({
                        player_col: fb_df[player_col],
                        'avg_distance_fb': 0,
                        'std_distance_fb': 0
                    })
        
        # Do the same for LD
        if distance_col in ld_df.columns:
            if player_id_col:
                ld_dist_df = ld_df[[player_col, player_id_col, distance_col]].copy()
            else:
                ld_dist_df = ld_df[[player_col, distance_col]].copy()
            ld_dist_df.rename(columns={distance_col: 'avg_distance_ld'}, inplace=True)
            # Add std_dist_ld
            ld_dist_df['std_distance_ld'] = ld_dist_df['avg_distance_ld'] * 0.10
        else:
            # Try to find alternative distance column
            alt_distance_cols = ['hit_distance_sc', 'hit_distance', 'distance']
            found_col = False
            for col in alt_distance_cols:
                if col in ld_df.columns:
                    if player_id_col:
                        ld_dist_df = ld_df[[player_col, player_id_col, col]].copy()
                    else:
                        ld_dist_df = ld_df[[player_col, col]].copy()
                    ld_dist_df.rename(columns={col: 'avg_distance_ld'}, inplace=True)
                    ld_dist_df['std_distance_ld'] = ld_dist_df['avg_distance_ld'] * 0.10
                    found_col = True
                    break
                    
            if not found_col:
                logger.warning("No distance column found in LD data, using zeros")
                if player_id_col:
                    ld_dist_df = pd.DataFrame({
                        player_col: ld_df[player_col],
                        player_id_col: ld_df[player_id_col],
                        'avg_distance_ld': 0,
                        'std_distance_ld': 0
                    })
                else:
                    ld_dist_df = pd.DataFrame({
                        player_col: ld_df[player_col],
                        'avg_distance_ld': 0,
                        'std_distance_ld': 0
                    })
        
        # 3. Extract event counts - find the 'pitches' or equivalent column for each file
        event_count_cols = ['pitches', 'bip', 'abs', 'pa']
        
        # Get counts for FB
        fb_count_col = None
        for col in event_count_cols:
            if col in fb_df.columns:
                fb_count_col = col
                break
        
        # Extract counts
        if fb_count_col:
            if player_id_col:
                fb_counts = fb_df[[player_col, player_id_col, fb_count_col]].copy()
            else:
                fb_counts = fb_df[[player_col, fb_count_col]].copy()
            fb_counts.rename(columns={fb_count_col: 'FB_Events'}, inplace=True)
        else:
            logger.warning(f"No count column found in FB data")
            if player_id_col:
                fb_counts = pd.DataFrame({
                    player_col: fb_df[player_col], 
                    player_id_col: fb_df[player_id_col], 
                    'FB_Events': 0
                })
            else:
                fb_counts = pd.DataFrame({player_col: fb_df[player_col], 'FB_Events': 0})
        
        # Get counts for Pull FB
        pull_fb_count_col = None
        for col in event_count_cols:
            if col in pull_fb_df.columns:
                pull_fb_count_col = col
                break
        
        if pull_fb_count_col:
            if player_id_col:
                pull_fb_counts = pull_fb_df[[player_col, player_id_col, pull_fb_count_col]].copy()
            else:
                pull_fb_counts = pull_fb_df[[player_col, pull_fb_count_col]].copy()
            pull_fb_counts.rename(columns={pull_fb_count_col: 'Pull_FB_Events'}, inplace=True)
        else:
            logger.warning(f"No count column found in Pull FB data")
            if player_id_col:
                pull_fb_counts = pd.DataFrame({
                    player_col: pull_fb_df[player_col], 
                    player_id_col: pull_fb_df[player_id_col],
                    'Pull_FB_Events': 0
                })
            else:
                pull_fb_counts = pd.DataFrame({player_col: pull_fb_df[player_col], 'Pull_FB_Events': 0})
        
        # Get counts for Oppo FB
        oppo_fb_count_col = None
        for col in event_count_cols:
            if col in oppo_fb_df.columns:
                oppo_fb_count_col = col
                break
        
        if oppo_fb_count_col:
            if player_id_col:
                oppo_fb_counts = oppo_fb_df[[player_col, player_id_col, oppo_fb_count_col]].copy()
            else:
                oppo_fb_counts = oppo_fb_df[[player_col, oppo_fb_count_col]].copy()
            oppo_fb_counts.rename(columns={oppo_fb_count_col: 'Oppo_FB_Events'}, inplace=True)
        else:
            logger.warning(f"No count column found in Oppo FB data")
            if player_id_col:
                oppo_fb_counts = pd.DataFrame({
                    player_col: oppo_fb_df[player_col], 
                    player_id_col: oppo_fb_df[player_id_col],
                    'Oppo_FB_Events': 0
                })
            else:
                oppo_fb_counts = pd.DataFrame({player_col: oppo_fb_df[player_col], 'Oppo_FB_Events': 0})
        
        # Get counts for LD
        ld_count_col = None
        for col in event_count_cols:
            if col in ld_df.columns:
                ld_count_col = col
                break
        
        if ld_count_col:
            if player_id_col:
                ld_counts = ld_df[[player_col, player_id_col, ld_count_col]].copy()
            else:
                ld_counts = ld_df[[player_col, ld_count_col]].copy()
            ld_counts.rename(columns={ld_count_col: 'LD_Events'}, inplace=True)
        else:
            logger.warning(f"No count column found in LD data")
            if player_id_col:
                ld_counts = pd.DataFrame({
                    player_col: ld_df[player_col], 
                    player_id_col: ld_df[player_id_col], 
                    'LD_Events': 0
                })
            else:
                ld_counts = pd.DataFrame({player_col: ld_df[player_col], 'LD_Events': 0})
        
        # Get counts for Pull LD
        pull_ld_count_col = None
        for col in event_count_cols:
            if col in pull_ld_df.columns:
                pull_ld_count_col = col
                break
        
        if pull_ld_count_col:
            if player_id_col:
                pull_ld_counts = pull_ld_df[[player_col, player_id_col, pull_ld_count_col]].copy()
            else:
                pull_ld_counts = pull_ld_df[[player_col, pull_ld_count_col]].copy()
            pull_ld_counts.rename(columns={pull_ld_count_col: 'Pull_LD_Events'}, inplace=True)
        else:
            logger.warning(f"No count column found in Pull LD data")
            if player_id_col:
                pull_ld_counts = pd.DataFrame({
                    player_col: pull_ld_df[player_col], 
                    player_id_col: pull_ld_df[player_id_col],
                    'Pull_LD_Events': 0
                })
            else:
                pull_ld_counts = pd.DataFrame({player_col: pull_ld_df[player_col], 'Pull_LD_Events': 0})
        
        # Get counts for Oppo LD
        oppo_ld_count_col = None
        for col in event_count_cols:
            if col in oppo_ld_df.columns:
                oppo_ld_count_col = col
                break
        
        if oppo_ld_count_col:
            if player_id_col:
                oppo_ld_counts = oppo_ld_df[[player_col, player_id_col, oppo_ld_count_col]].copy()
            else:
                oppo_ld_counts = oppo_ld_df[[player_col, oppo_ld_count_col]].copy()
            oppo_ld_counts.rename(columns={oppo_ld_count_col: 'Oppo_LD_Events'}, inplace=True)
        else:
            logger.warning(f"No count column found in Oppo LD data")
            if player_id_col:
                oppo_ld_counts = pd.DataFrame({
                    player_col: oppo_ld_df[player_col], 
                    player_id_col: oppo_ld_df[player_id_col],
                    'Oppo_LD_Events': 0
                })
            else:
                oppo_ld_counts = pd.DataFrame({player_col: oppo_ld_df[player_col], 'Oppo_LD_Events': 0})
        
        # Merge all components into one dataframe
        # If we have player IDs, use them for more accurate merging
        if player_id_col:
            # Merge using both player name and ID for more accurate matching
            barrel_df = pd.merge(fb_barrels, ld_barrels, on=[player_col, player_id_col], how='outer')
            barrel_df = pd.merge(barrel_df, fb_dist_df, on=[player_col, player_id_col], how='left')
            barrel_df = pd.merge(barrel_df, ld_dist_df, on=[player_col, player_id_col], how='left')
            barrel_df = pd.merge(barrel_df, fb_counts, on=[player_col, player_id_col], how='left')
            barrel_df = pd.merge(barrel_df, pull_fb_counts, on=[player_col, player_id_col], how='left')
            barrel_df = pd.merge(barrel_df, oppo_fb_counts, on=[player_col, player_id_col], how='left')
            barrel_df = pd.merge(barrel_df, ld_counts, on=[player_col, player_id_col], how='left')
            barrel_df = pd.merge(barrel_df, pull_ld_counts, on=[player_col, player_id_col], how='left')
            barrel_df = pd.merge(barrel_df, oppo_ld_counts, on=[player_col, player_id_col], how='left')
            
            # Rename player_id_col to PlayerID for consistency
            barrel_df.rename(columns={player_id_col: 'PlayerID'}, inplace=True)
        else:
            # Original merge without player ID
            barrel_df = pd.merge(fb_barrels, ld_barrels, on=player_col, how='outer')
            barrel_df = pd.merge(barrel_df, fb_dist_df, on=player_col, how='left')
            barrel_df = pd.merge(barrel_df, ld_dist_df, on=player_col, how='left')
            barrel_df = pd.merge(barrel_df, fb_counts, on=player_col, how='left')
            barrel_df = pd.merge(barrel_df, pull_fb_counts, on=player_col, how='left')
            barrel_df = pd.merge(barrel_df, oppo_fb_counts, on=player_col, how='left')
            barrel_df = pd.merge(barrel_df, ld_counts, on=player_col, how='left')
            barrel_df = pd.merge(barrel_df, pull_ld_counts, on=player_col, how='left')
            barrel_df = pd.merge(barrel_df, oppo_ld_counts, on=player_col, how='left')
        
        # Fill na values with 0
        for col in barrel_df.columns:
            if col != player_col and col != 'PlayerID':
                barrel_df[col] = barrel_df[col].fillna(0).round(2)
        
        # Calculate combined metrics for distance
        barrel_df['avg_dist_fb_ld'] = ((barrel_df['avg_distance_fb'] + barrel_df['avg_distance_ld']) / 2).round(2)
        barrel_df['std_dist_fb_ld'] = ((barrel_df['std_distance_fb'] + barrel_df['std_distance_ld']) / 2).round(2)
        
        # Added pull_fb_pct, pull_ld_pct, oppo_fb_pct, oppo_ld_pct columns with zeros
        # to maintain column order compatibility
        barrel_df['pull_fb_pct'] = 0
        barrel_df['pull_ld_pct'] = 0
        barrel_df['oppo_fb_pct'] = 0
        barrel_df['oppo_ld_pct'] = 0
        
        # Clean up player names for better Excel compatibility
        if player_col == 'player_name' and ',' in str(barrel_df[player_col].iloc[0]):
            # Format is "Last, First" - convert to "First Last"
            barrel_df['Name'] = barrel_df[player_col].apply(
                lambda x: f"{x.split(', ')[1]} {x.split(', ')[0]}" if isinstance(x, str) and ', ' in x else x
            )
            logger.info("Converted player names from 'Last, First' to 'First Last' format")
        else:
            barrel_df['Name'] = barrel_df[player_col]
        
        # Get reference data from salaries for standardization
        reference_df = get_salaries_for_reference()
        
        # Standardize player names to match other sheets
        if reference_df is not None and not reference_df.empty:
            logger.info("Standardizing barrel data player names to match salary data...")
            
            # Create a copy with standardized names
            if 'PlayerID' in barrel_df.columns:
                # Use the enhanced standardization that considers player IDs
                std_barrel_df = standardize_all_player_names(barrel_df, reference_df, 'Name', 'PlayerID')
            else:
                # Use the original standardization
                std_barrel_df = standardize_all_player_names(barrel_df, reference_df, 'Name')
            
            if 'Name' in std_barrel_df.columns:
                # Keep track of how many names were changed
                changed_count = sum(barrel_df['Name'] != std_barrel_df['Name'])
                if changed_count > 0:
                    logger.info(f"Standardized {changed_count} player names in barrel data")
                    
                    # Show examples of standardized names
                    if changed_count > 0:
                        examples = []
                        for idx, (old, new) in enumerate(zip(barrel_df['Name'], std_barrel_df['Name'])):
                            if old != new and len(examples) < 10:
                                examples.append((old, new))
                        
                        if examples:
                            logger.info("Examples of standardized barrel data names:")
                            for old, new in examples:
                                logger.info(f"  '{old}' -> '{new}'")
                
                # Use the standardized dataframe
                barrel_df = std_barrel_df
            else:
                logger.warning("Name column not found after standardization, using original names")
        else:
            logger.warning("No reference data available for name standardization")
        
        # Create alternate name format for better VLOOKUP matching
        def create_alt_name(name):
            if not isinstance(name, str):
                return name
                
            # Handle name variants (accents, etc.)
            variants = {
                'Acuna': 'Acuña',
                'Munoz': 'Muñoz',
                'Pena': 'Peña',
                'Jimenez': 'Jiménez',
                'Suarez': 'Suárez',
                'Sanchez': 'Sánchez',
                'Hernandez': 'Hernández',
                'Gonzalez': 'González',
                'Martinez': 'Martínez',
                'Ramirez': 'Ramírez',
                'Rodriguez': 'Rodríguez',
            }
            
            alt_name = name
            for name1, name2 in variants.items():
                if name1 in alt_name:
                    alt_name = alt_name.replace(name1, name2)
                    return alt_name
                elif name2 in alt_name:
                    alt_name = alt_name.replace(name2, name1)
                    return alt_name
            
            return name
            
        barrel_df['NameAlt'] = barrel_df['Name'].apply(create_alt_name)
        
        # Add timestamp
        current_date = datetime.now().strftime("%Y-%m-%d")
        barrel_df['LastUpdated'] = current_date
        
        # Reorder columns for VLOOKUP with 'Name' first
        column_order = [
            'Name', 
        ]
        
        # Add PlayerID right after Name if it exists
        if 'PlayerID' in barrel_df.columns:
            column_order.append('PlayerID')
            
        # Add the rest of the columns
        column_order.extend([
            'barrels/fb',  # Barrel FB%
            'barrels/ld',  # Barrel LD%
            'avg_distance_fb',  # Avg Dist FB
            'avg_distance_ld',  # Avg Dist LD
            'std_distance_fb',  # Std Dev of Dist FB
            'std_distance_ld',  # Std Dev of Dist LD
            'pull_fb_pct',  # Pull FB% (zero placeholder)
            'pull_ld_pct',  # Pull LD% (zero placeholder)
            'oppo_fb_pct',  # Oppo FB% (zero placeholder)
            'oppo_ld_pct',  # Oppo LD% (zero placeholder)
            'avg_dist_fb_ld',  # Combined for convenience 
            'std_dist_fb_ld',  # Combined for convenience
            'FB_Events',  # Raw event counts
            'Pull_FB_Events',  # Raw event counts
            'Oppo_FB_Events',  # Raw event counts
            'LD_Events',  # Raw event counts
            'Pull_LD_Events',  # Raw event counts
            'Oppo_LD_Events',  # Raw event counts
            'NameAlt',  # For alternative lookup
            'LastUpdated',  # Timestamp
            player_col  # Original player name
        ])
        
        # Ensure all columns exist in the result
        final_cols = [col for col in column_order if col in barrel_df.columns]
        
        # Add any columns that exist in the result but aren't in our order
        for col in barrel_df.columns:
            if col not in final_cols:
                final_cols.append(col)
        
        # Reorder columns
        barrel_df = barrel_df[final_cols]
        
        # Save to CSV with both timestamped and fixed filenames
        timestamp = datetime.now().strftime("%Y%m%d")
        result_file = os.path.join(output_folder, f"barrel_percentages_{timestamp}.csv")
        fixed_file = os.path.join(output_folder, "barrel_percentages.csv")
        
        # Remove existing files if they exist
        for file_path in [result_file, fixed_file]:
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    logger.info(f"Deleted existing file: {file_path}")
                except Exception as del_e:
                    logger.warning(f"Could not delete existing file {file_path}: {str(del_e)}")
        
        # Save the files
        barrel_df.to_csv(result_file, index=False)
        barrel_df.to_csv(fixed_file, index=False)
        
        logger.info(f"Successfully processed barrel data components for {len(barrel_df)} players")
        
        # Show top 10 players for logging
        if not barrel_df.empty:
            top_players = barrel_df.head(10)
            top_cols = ['Name']
            if 'PlayerID' in barrel_df.columns:
                top_cols.append('PlayerID')
            top_cols.extend(['barrels/fb', 'barrels/ld', 'FB_Events', 'LD_Events'])
            logger.info(f"Top 10 players by barrel percentage:\n{top_players[top_cols]}")
        
        # Check for duplicate names and report them
        if 'Name' in barrel_df.columns:
            dupes = barrel_df['Name'].value_counts()
            dupes = dupes[dupes > 1]
            if not dupes.empty:
                logger.warning(f"Found {len(dupes)} names with duplicates:")
                for name, count in dupes.items():
                    logger.warning(f"  {name}: {count} entries")
                    
                    # If we have PlayerID, show the details of the duplicates
                    if 'PlayerID' in barrel_df.columns:
                        dupe_entries = barrel_df[barrel_df['Name'] == name]
                        for _, entry in dupe_entries.iterrows():
                            logger.warning(f"    PlayerID: {entry['PlayerID']}, FB Events: {entry['FB_Events']}, LD Events: {entry['LD_Events']}")
        
        return {
            "success": True,
            "file_path": fixed_file,
            "player_count": len(barrel_df),
            "datestamp": timestamp
        }
    
    except Exception as e:
        logger.error(f"Error processing barrel data: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return {"success": False, "error": str(e)}


def update_excel_with_barrel_data(barrel_csv_path, excel_path=EXCEL_FILE_PATH, sheet_name="BarrelStats"):
    """
    Update the Excel file with barrel data from CSV
    
    Parameters:
    -----------
    barrel_csv_path : str
        Path to the CSV file with barrel data
    excel_path : str
        Path to the Excel file
    sheet_name : str
        Name of the sheet to update
    
    Returns:
    --------
    bool
        True if successful, False otherwise
    """
    
    logger.info(f"Updating Excel file with barrel data from {barrel_csv_path}")
    
    try:
        # Read the CSV file
        barrel_df = pd.read_csv(barrel_csv_path)
        
        if barrel_df.empty:
            logger.warning("Barrel data CSV is empty")
            return False
        
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        
        # Check if sheet exists, remove if it does
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
            logger.info(f"Removed existing {sheet_name} sheet")
        
        # Create new sheet
        ws = wb.create_sheet(sheet_name)
        
        # Write column headers
        for col_idx, col_name in enumerate(barrel_df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        # Write data rows
        for row_idx, row in enumerate(barrel_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save the workbook
        wb.save(excel_path)
        
        logger.info(f"Successfully updated {sheet_name} sheet with {len(barrel_df)} rows of barrel data")
        return True
        
    except Exception as e:
        logger.error(f"Error updating Excel with barrel data: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False
    
def find_most_recent_barrel_data(barrel_data_folder="barrel_data", pattern="barrel_percentages_*.csv"):
    """
    Find the most recent barrel percentages file in the specified folder
    
    Parameters:
    -----------
    barrel_data_folder : str
        Folder to search for barrel data files
    pattern : str
        Glob pattern to match barrel data files
    
    Returns:
    --------
    str or None
        Path to most recent barrel data file, or None if not found
    """
    logger.info(f"Searching for most recent barrel data in {barrel_data_folder}")
    
    try:
        # Check if folder exists
        if not os.path.exists(barrel_data_folder):
            logger.warning(f"Barrel data folder {barrel_data_folder} does not exist")
            return None
        
        # Check for fixed name file first (if using overwrite=True)
        fixed_file = os.path.join(barrel_data_folder, "barrel_percentages.csv")
        if os.path.exists(fixed_file):
            return fixed_file
        
        # Search for pattern files
        import glob
        pattern_path = os.path.join(barrel_data_folder, pattern)
        matching_files = glob.glob(pattern_path)
        
        # If we also want to check for the barrel_stats direct file
        barrel_stats_file = os.path.join(barrel_data_folder, "barrel_stats.csv")
        if os.path.exists(barrel_stats_file):
            matching_files.append(barrel_stats_file)
        
        if not matching_files:
            logger.warning(f"No barrel data files found matching {pattern}")
            return None
        
        # Sort by modification time (newest first)
        matching_files.sort(key=os.path.getmtime, reverse=True)
        
        most_recent = matching_files[0]
        logger.info(f"Found most recent barrel data file: {most_recent}")
        return most_recent
        
    except Exception as e:
        logger.error(f"Error finding most recent barrel data: {str(e)}")
        return None

def update_statcast_sheet_in_excel(excel_path=EXCEL_FILE_PATH, barrel_data_path=None, sheet_name="Statcast"):
    """
    Update or create a Statcast sheet in the Excel file with barrel data
    
    Parameters:
    -----------
    excel_path : str
        Path to the Excel file to update
    barrel_data_path : str or None
        Path to the barrel data CSV file, or None to auto-find most recent
    sheet_name : str
        Name of the sheet to update or create
    
    Returns:
    --------
    bool
        True if successful, False otherwise
    """
    logger.info(f"Updating {sheet_name} sheet in {excel_path} with barrel data")
    
    try:
        # Find most recent barrel data if not specified
        if barrel_data_path is None:
            # Look for fixed name file first
            fixed_file = os.path.join("barrel_data", "barrel_percentages.csv")
            
            if os.path.exists(fixed_file):
                barrel_data_path = fixed_file
                logger.info(f"Using fixed barrel data file: {fixed_file}")
            else:
                # Fall back to most recent file
                import glob
                pattern = os.path.join("barrel_data", "barrel_percentages_*.csv")
                matching_files = glob.glob(pattern)
                
                if not matching_files:
                    logger.error("No barrel data file found")
                    return False
                
                # Sort by modification time (newest first)
                matching_files.sort(key=os.path.getmtime, reverse=True)
                barrel_data_path = matching_files[0]
                logger.info(f"Using most recent barrel data file: {barrel_data_path}")
            
        # Read the barrel data
        barrel_df = pd.read_csv(barrel_data_path)
        
        if barrel_df.empty:
            logger.warning("Barrel data CSV is empty")
            return False
        
        # Make a backup before modifying the Excel file
        # Use the BACKUP_FOLDER instead of os.path.dirname
        backup_path = os.path.join(BACKUP_FOLDER, f"MLBProjections_backup_statcast_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsm")
        shutil.copy2(excel_path, backup_path)
        logger.info(f"Created backup before updating Statcast sheet: {backup_path}")
        
        # Ensure we have the expected columns
        expected_columns = ['Name', 'fb_count', 'fb_barrel_count', 'barrels/fb', 
                           'ld_count', 'ld_barrel_count', 'barrels/ld']
        
        missing_columns = [col for col in expected_columns if col not in barrel_df.columns]
        if missing_columns:
            logger.warning(f"Missing expected columns in barrel data: {missing_columns}")
            # Continue anyway with available columns
        
        # Check for player ID column
        has_player_id = 'PlayerID' in barrel_df.columns
        if has_player_id:
            logger.info("Found PlayerID column in barrel data, which will help with duplicate player names")
        else:
            logger.warning("No PlayerID column found in barrel data. Duplicate player names may not be handled correctly.")
        
        # Standardize player names to match Excel format
        excel_name_format = detect_excel_name_format(excel_path)
        logger.info(f"Detected name format in Excel: {excel_name_format}")
        
        # Make sure the Name column is properly formatted
        if 'Name' in barrel_df.columns:
            # Check for "Last, First" format that needs to be converted to "First Last"
            sample_name = barrel_df['Name'].iloc[0] if not barrel_df.empty else ""
            
            if isinstance(sample_name, str) and ',' in sample_name:
                if excel_name_format == "first_last":
                    # Convert "Last, First" to "First Last"
                    barrel_df['Name'] = barrel_df['Name'].apply(
                        lambda x: f"{x.split(', ')[1]} {x.split(', ')[0]}" if isinstance(x, str) and ', ' in x else x
                    )
                    logger.info("Converted player names from 'Last, First' to 'First Last' format")
            elif ',' not in sample_name and excel_name_format == "last_first":
                # Convert "First Last" to "Last, First"
                def convert_to_last_first(name):
                    if not isinstance(name, str) or ',' in name:
                        return name
                    
                    parts = name.strip().split()
                    if len(parts) < 2:
                        return name
                    
                    last_name = parts[-1]
                    first_name = ' '.join(parts[:-1])
                    return f"{last_name}, {first_name}"
                
                barrel_df['Name'] = barrel_df['Name'].apply(convert_to_last_first)
                logger.info("Converted player names from 'First Last' to 'Last, First' format")
        
        # Add NameAlt column with variations to help with lookups
        def enhance_name_matching(name):
            """Create an alternative version of the name to help with lookups"""
            if not isinstance(name, str):
                return name
                
            # Handle common name variants
            variants = {
                'Acuna': 'Acuña',
                'Munoz': 'Muñoz',
                'Pena': 'Peña',
                'Jiminez': 'Jiménez',
                'Jimenez': 'Jiménez',
                'Suarez': 'Suárez',
                'Sanchez': 'Sánchez',
                'Hernandez': 'Hernández',
                'Gonzalez': 'González',
                'Martinez': 'Martínez',
                'Ramirez': 'Ramírez',
                'Rodriguez': 'Rodríguez',
            }
            
            alt_name = name
            
            # Try variants in both directions
            for name1, name2 in variants.items():
                if name1 in alt_name:
                    alt_name = alt_name.replace(name1, name2)
                elif name2 in alt_name:
                    alt_name = alt_name.replace(name2, name1)
            
            return alt_name
        
        if 'NameAlt' not in barrel_df.columns:
            barrel_df['NameAlt'] = barrel_df['Name'].apply(enhance_name_matching)
        
        # Create a new column with the date when this data was last updated
        current_date = datetime.now().strftime("%Y-%m-%d")
        barrel_df['LastUpdated'] = current_date
        
        # Check for duplicate names and log them
        if 'Name' in barrel_df.columns:
            dupes = barrel_df['Name'].value_counts()
            dupes = dupes[dupes > 1]
            if not dupes.empty:
                logger.warning(f"Found {len(dupes)} names with duplicates before writing to Excel:")
                for name, count in dupes.items():
                    logger.warning(f"  {name}: {count} entries")
                    
                    # If we have PlayerID, show the details of the duplicates
                    if 'PlayerID' in barrel_df.columns:
                        dupe_entries = barrel_df[barrel_df['Name'] == name]
                        for _, entry in dupe_entries.iterrows():
                            logger.warning(f"    PlayerID: {entry['PlayerID']}, FB Events: {entry.get('FB_Events', 'N/A')}, LD Events: {entry.get('LD_Events', 'N/A')}")
        
        # Log data summary before updating Excel
        logger.info(f"Prepared {len(barrel_df)} player records with barrel data")
        if not barrel_df.empty:
            logger.info(f"Columns in final data: {barrel_df.columns.tolist()}")
            
            # Show top 5 barrel pct players for both FB and LD
            if 'barrels/fb' in barrel_df.columns:
                logger.info("Top 5 players by FB barrel percentage:")
                top_fb = barrel_df.sort_values('barrels/fb', ascending=False).head(5)
                for _, row in top_fb.iterrows():
                    name = row.get('Name', 'Unknown')
                    pct = row.get('barrels/fb', 0)
                    fb_events = row.get('FB_Events', 0)
                    fb_barrel_count = row.get('fb_barrel_count', 0)
                    pid = row.get('PlayerID', 'N/A') if has_player_id else 'N/A'
                    logger.info(f"  {name} (ID: {pid}): {pct:.1f}% ({fb_barrel_count}/{fb_events})")
                    
            if 'barrels/ld' in barrel_df.columns:
                logger.info("Top 5 players by LD barrel percentage:")
                top_ld = barrel_df.sort_values('barrels/ld', ascending=False).head(5)
                for _, row in top_ld.iterrows():
                    name = row.get('Name', 'Unknown')
                    pct = row.get('barrels/ld', 0)
                    ld_events = row.get('LD_Events', 0)
                    ld_barrel_count = row.get('ld_barrel_count', 0)
                    pid = row.get('PlayerID', 'N/A') if has_player_id else 'N/A'
                    logger.info(f"  {name} (ID: {pid}): {pct:.1f}% ({ld_barrel_count}/{ld_events})")
        
        # Open the Excel file and update/create the sheet
        workbook = openpyxl.load_workbook(excel_path, keep_vba=True)
        
        # Remove existing sheet if it exists
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
            logger.info(f"Removed existing {sheet_name} sheet")
        
        # Create new sheet
        worksheet = workbook.create_sheet(sheet_name)
        
        # Write column headers
        for col_idx, col_name in enumerate(barrel_df.columns, 1):
            worksheet.cell(row=1, column=col_idx, value=col_name)
        
        # Write data rows
        for row_idx, data_row in enumerate(barrel_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(data_row, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Save the workbook
        workbook.save(excel_path)
        
        logger.info(f"Successfully updated {sheet_name} sheet with {len(barrel_df)} rows of barrel data")
        
        # Verify data after saving
        if has_player_id:
            logger.info("PlayerID column was included in the sheet to help with duplicate player names")
        
        return True
        
    except Exception as e:
        logger.error(f"Error updating {sheet_name} sheet: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False
    
def download_fb_ld_event_data(output_folder="statcast_data", create_folder=True):
    """
    Downloads event-level data for fly balls and line drives from Baseball Savant
    by clicking the "Download Data as CSV" button on the page.
    
    Parameters:
    -----------
    output_folder : str
        Folder to save downloaded CSV files
    create_folder : bool
        Whether to create the output folder if it doesn't exist
    
    Returns:
    --------
    dict
        Dictionary containing path to downloaded file and processing status
    """
    
    logger.info("Starting Baseball Savant event-level FB+LD data download")
    start_time = time.time()
    
    # Create output folder if needed
    if create_folder and not os.path.exists(output_folder):
        os.makedirs(output_folder)
        logger.info(f"Created output folder: {output_folder}")
    
    # The specific URL for the FB+LD events search page
    url = "https://baseballsavant.mlb.com/statcast_search?hfPT=&hfAB=&hfGT=R%7C&hfPR=&hfZ=&hfStadium=&hfBBL=&hfNewZones=&hfPull=&hfC=&hfSea=2025%7C&hfSit=&player_type=batter&hfOuts=&hfOpponent=&pitcher_throws=&batter_stands=&hfSA=&game_date_gt=&game_date_lt=&hfMo=&hfTeam=&home_road=&hfRO=&position=&hfInfield=&hfOutfield=&hfInn=&hfBBT=fly%5C.%5C.ball%7Cline%5C.%5C.drive%7C&hfFlag=&metric_1=&group_by=name-event&min_pitches=0&min_results=0&min_pas=0&sort_col=pitches&player_event_sort=api_h_distance_projected&sort_order=desc&chk_event_release_speed=on&chk_event_hit_distance_sc=on#results"
    
    # Output filename - accept multiple possible filenames
    final_output_file = os.path.join(output_folder, "fb_ld_events.csv")
    
    # Delete existing final file if it exists
    if os.path.exists(final_output_file):
        try:
            os.remove(final_output_file)
            logger.info(f"Deleted existing file: {final_output_file}")
        except Exception as del_e:
            logger.warning(f"Could not delete existing file {final_output_file}: {str(del_e)}")
    
    try:
        logger.info("Setting up Chrome browser for downloading FB+LD event data...")
        
        # Setup Chrome options with download preferences
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1920,1080")  # Ensure elements are visible
        
        # Set download preferences
        prefs = {
            "download.default_directory": os.path.abspath(output_folder),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": False
        }
        chrome_options.add_experimental_option("prefs", prefs)
        
        # Create and configure the Chrome WebDriver
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.set_page_load_timeout(120)  # 2 minute timeout for page load
        
        page_load_start = time.time()
        logger.info(f"Navigating to FB+LD events page: {url}")
        driver.get(url)
        
        # Wait for the page to load
        logger.info("Waiting for page to load and results to appear...")
        try:
            # Wait for the table with results to be visible
            WebDriverWait(driver, 180).until(
                EC.presence_of_element_located((By.ID, "search_results"))
            )
            page_load_end = time.time()
            page_load_time = page_load_end - page_load_start
            logger.info(f"Search results table found - page took {page_load_time:.2f} seconds to load")
            
            # Find and click the CSV download button - specifically look for the database_link.png image
            logger.info("Looking for 'Download Data as CSV' button (database_link.png)...")
            button_search_start = time.time()
            
            # First try to find by ID
            csv_button = None
            try:
                # Try finding by ID
                csv_button = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((By.ID, "csv_all_pid_"))
                )
                logger.info("Found CSV button by ID (csv_all_pid_)")
            except:
                try:
                    # Try finding by class
                    csv_button = WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.CLASS_NAME, "csv"))
                    )
                    logger.info("Found CSV button by class name")
                except:
                    try:
                        # Try finding the img element with database_link.png
                        csv_button = WebDriverWait(driver, 30).until(
                            EC.element_to_be_clickable((By.XPATH, "//img[contains(@src, 'database_link.png')]"))
                        )
                        logger.info("Found CSV button by database_link.png image src")
                    except:
                        # One more try - look for any csv span that's not csv_table
                        csv_buttons = driver.find_elements(By.XPATH, "//*[contains(@class, 'csv') and not(contains(@class, 'csv_table')) or contains(@id, 'csv')]")
                        if csv_buttons:
                            csv_button = csv_buttons[0]
                            logger.info(f"Found {len(csv_buttons)} potential CSV buttons by partial match")
            
            button_search_end = time.time()
            button_search_time = button_search_end - button_search_start
            logger.info(f"Button search took {button_search_time:.2f} seconds")
            
            if csv_button:
                # Click the CSV download button
                logger.info("Clicking 'Download Data as CSV' button...")
                download_start = time.time()
                driver.execute_script("arguments[0].click();", csv_button)
                logger.info("CSV download initiated")
                
                # Wait for download to complete
                logger.info("Waiting for download to complete (up to 5 minutes)...")
                
                # Look for downloaded file with different possible names
                download_complete = False
                downloaded_file = None
                
                # List of possible downloaded filenames
                possible_filenames = [
                    "fb_ld_events.csv",
                    "statcast_search.csv",
                    "savant_data.csv",
                    "statcast_data.csv",
                    "search_event_data.csv",
                    "search_results.csv"
                ]
                
                # Wait for any of the possible files to appear
                while not download_complete and (time.time() - download_start) < 300:  # 5 minute timeout
                    for filename in possible_filenames:
                        file_path = os.path.join(output_folder, filename)
                        if os.path.exists(file_path):
                            # File exists, check if download is complete
                            try:
                                # Try to open the file - if it's still being written, this will fail
                                with open(file_path, 'r') as f:
                                    # Read a small amount to check if file is accessible
                                    _ = f.read(10)
                                    
                                # If we get here, file is accessible and download is complete
                                download_end = time.time()
                                download_time = download_end - download_start
                                logger.info(f"Download completed in {download_time:.2f} seconds: {file_path}")
                                
                                # Save the downloaded file path
                                downloaded_file = file_path
                                download_complete = True
                                break
                            except Exception:
                                # File exists but is still being written
                                time.sleep(1)
                    
                    if not download_complete:
                        time.sleep(2)  # Wait before checking again
                
                if download_complete and downloaded_file:
                    # Verify the downloaded file
                    file_size = os.path.getsize(downloaded_file)
                    logger.info(f"Downloaded file size: {file_size} bytes")
                    
                    if file_size > 1000:  # Ensure we got meaningful data
                        total_time = time.time() - start_time
                        logger.info(f"Successfully downloaded FB+LD events data ({file_size} bytes)")
                        logger.info(f"Total process took {total_time:.2f} seconds")
                        
                        # Try to copy the downloaded file to our standard name instead of renaming
                        # This avoids file access issues
                        try:
                            # If it's not already the final output file
                            if downloaded_file != final_output_file:
                                # Copy file content instead of renaming
                                with open(downloaded_file, 'r') as src_file:
                                    with open(final_output_file, 'w') as dest_file:
                                        dest_file.write(src_file.read())
                                logger.info(f"Copied {downloaded_file} to {final_output_file}")
                        except Exception as copy_e:
                            logger.warning(f"Could not copy file: {str(copy_e)}")
                            # If copy fails, just use the downloaded file directly
                            final_output_file = downloaded_file
                            logger.info(f"Using downloaded file directly: {final_output_file}")
                        
                        # Check if the file contains the expected raw event data
                        with open(final_output_file, 'r', encoding='utf-8') as f:
                            header = f.readline().strip()
                            
                        if 'hit_distance_sc' in header and 'launch_speed' in header:
                            logger.info("File contains raw event data with required columns")
                            result = {
                                "success": True,
                                "file_path": final_output_file,
                                "size": file_size,
                                "timing": {
                                    "page_load": page_load_time,
                                    "button_search": button_search_time,
                                    "download": download_time,
                                    "total": total_time
                                }
                            }
                        else:
                            logger.warning("Downloaded file doesn't contain expected raw event data columns")
                            logger.info(f"Header columns: {header[:200]}...")
                            result = {
                                "success": False,
                                "file_path": final_output_file,
                                "error": "File doesn't contain expected raw event data columns"
                            }
                    else:
                        logger.warning(f"Downloaded file is suspiciously small ({file_size} bytes)")
                        result = {
                            "success": False,
                            "file_path": downloaded_file,
                            "error": "File too small, possible error in download"
                        }
                else:
                    logger.error("Download timed out after 5 minutes")
                    result = {
                        "success": False,
                        "error": "Download timed out after 5 minutes"
                    }
            else:
                logger.error("Could not find 'Download Data as CSV' button")
                result = {
                    "success": False, 
                    "error": "Could not find 'Download Data as CSV' button"
                }
                
        except Exception as wait_e:
            logger.error(f"Error waiting for page elements: {str(wait_e)}")
            result = {
                "success": False, 
                "error": f"Error waiting for page elements: {str(wait_e)}"
            }
        
        # Always close the driver
        driver.quit()
        
    except Exception as e:
        logger.error(f"Error downloading FB+LD events: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        
        result = {
            "success": False, 
            "error": str(e)
        }
        
        if 'driver' in locals():
            driver.quit()
    
    # Process the downloaded data if successful
    if result.get("success", False) and os.path.exists(result["file_path"]):
        process_start = time.time()
        processed_results = process_fb_ld_event_data(result["file_path"])
        process_end = time.time()
        process_time = process_end - process_start
        
        logger.info(f"Data processing took {process_time:.2f} seconds")
        
        # Add processing time to result timing
        if "timing" in result:
            result["timing"]["processing"] = process_time
            result["timing"]["total"] += process_time
        
        result["processed"] = processed_results
    
    return result

def process_fb_ld_event_data(csv_file_path):
    """
    Process raw FB+LD event data to calculate distance standard deviation
    and pull/oppo percentages from individual events.
    
    Parameters:
    -----------
    csv_file_path : str
        Path to the downloaded CSV file with raw FB+LD event data
    
    Returns:
    --------
    dict
        Dictionary containing processed statistics
    """
    
    logger.info("Processing raw FB+LD event data...")
    
    try:
        # Read the CSV file
        events_df = pd.read_csv(csv_file_path)
        logger.info(f"Read {len(events_df)} raw FB+LD events")
        
        # Check for necessary columns
        required_columns = ['player_name', 'hit_distance_sc', 'bb_type']
        missing_columns = [col for col in required_columns if col not in events_df.columns]
        
        if missing_columns:
            logger.warning(f"Missing required columns: {missing_columns}")
            return {
                "success": False,
                "error": f"Missing required columns: {missing_columns}"
            }
        
        # Separate FB and LD events
        fb_events = events_df[events_df['bb_type'] == 'fly_ball']
        ld_events = events_df[events_df['bb_type'] == 'line_drive']
        
        logger.info(f"Separated into {len(fb_events)} fly balls and {len(ld_events)} line drives")
        
        # Create DataFrame for final results
        final_results = pd.DataFrame()
        
        # Add player names from both types of events
        all_players = set(events_df['player_name'].unique())
        final_results['player_name'] = list(all_players)
        
        # Process FB events
        if not fb_events.empty:
            # Group by player to calculate statistics
            fb_stats = fb_events.groupby('player_name').agg(
                FB_Events=('hit_distance_sc', 'count'),
                avg_distance_fb=('hit_distance_sc', 'mean'),
                std_distance_fb=('hit_distance_sc', 'std')
            ).reset_index()
            
            # Replace NaN std with 0 for players with only one event
            fb_stats['std_distance_fb'] = fb_stats['std_distance_fb'].fillna(0)
            
            # Count pull/oppo events if the data includes hit direction
            if 'hit_direction' in events_df.columns:
                # Pull events
                pull_fb = fb_events[fb_events['hit_direction'] == 'pull']
                pull_fb_counts = pull_fb.groupby('player_name').size().reset_index(name='Pull_FB_Events')
                
                # Oppo events
                oppo_fb = fb_events[fb_events['hit_direction'] == 'opposite']
                oppo_fb_counts = oppo_fb.groupby('player_name').size().reset_index(name='Oppo_FB_Events')
                
                # Calculate percentages
                pull_fb_pct = pd.merge(pull_fb_counts, fb_stats[['player_name', 'FB_Events']], on='player_name', how='right')
                pull_fb_pct['pull_fb_pct'] = (pull_fb_pct['Pull_FB_Events'].fillna(0) / pull_fb_pct['FB_Events']) * 100
                pull_fb_pct = pull_fb_pct[['player_name', 'pull_fb_pct']]
                
                oppo_fb_pct = pd.merge(oppo_fb_counts, fb_stats[['player_name', 'FB_Events']], on='player_name', how='right')
                oppo_fb_pct['oppo_fb_pct'] = (oppo_fb_pct['Oppo_FB_Events'].fillna(0) / oppo_fb_pct['FB_Events']) * 100
                oppo_fb_pct = oppo_fb_pct[['player_name', 'oppo_fb_pct']]
                
                # Merge with fb_stats
                fb_stats = pd.merge(fb_stats, pull_fb_pct, on='player_name', how='left')
                fb_stats = pd.merge(fb_stats, oppo_fb_pct, on='player_name', how='left')
            else:
                # Add empty columns if hit direction not available
                fb_stats['pull_fb_pct'] = 0
                fb_stats['oppo_fb_pct'] = 0
            
            # Calculate barrel percentages if available
            if all(col in events_df.columns for col in ['launch_speed', 'launch_angle']):
                # Define barrel criteria (launch speed ≥ 98 mph and launch angle between 26-30°)
                def is_barrel(row):
                    ls = row.get('launch_speed', 0)
                    la = row.get('launch_angle', 0)
                    
                    if pd.isna(ls) or pd.isna(la):
                        return False
                    
                    # Simplified barrel definition
                    if ls >= 98:
                        if 26 <= la <= 30:
                            return True
                        elif ls >= 100 and 25 <= la <= 31:
                            return True
                        elif ls >= 102 and 24 <= la <= 32:
                            return True
                        elif ls >= 104 and 23 <= la <= 33:
                            return True
                        elif ls >= 106 and 22 <= la <= 34:
                            return True
                        elif ls >= 108 and 21 <= la <= 35:
                            return True
                    return False
                
                # Apply barrel criteria to FB events
                fb_events['is_barrel'] = fb_events.apply(is_barrel, axis=1)
                
                # Count barrels by player
                fb_barrels = fb_events.groupby('player_name')['is_barrel'].sum().reset_index(name='fb_barrel_count')
                
                # Calculate barrel percentage
                fb_barrel_pct = pd.merge(fb_barrels, fb_stats[['player_name', 'FB_Events']], on='player_name', how='right')
                fb_barrel_pct['barrels/fb'] = (fb_barrel_pct['fb_barrel_count'].fillna(0) / fb_barrel_pct['FB_Events']) * 100
                fb_barrel_pct = fb_barrel_pct[['player_name', 'barrels/fb']]
                
                # Merge with fb_stats
                fb_stats = pd.merge(fb_stats, fb_barrel_pct, on='player_name', how='left')
            else:
                fb_stats['barrels/fb'] = 0
                
            # Merge with final results
            final_results = pd.merge(final_results, fb_stats.drop('FB_Events', axis=1), on='player_name', how='left')
        
        # Process LD events
        if not ld_events.empty:
            # Group by player to calculate statistics
            ld_stats = ld_events.groupby('player_name').agg(
                LD_Events=('hit_distance_sc', 'count'),
                avg_distance_ld=('hit_distance_sc', 'mean'),
                std_distance_ld=('hit_distance_sc', 'std')
            ).reset_index()
            
            # Replace na std with 0 for players with only one event
            ld_stats['std_distance_ld'] = ld_stats['std_distance_ld'].fillna(0)
            
            # Count pull/oppo events if the data includes hit direction
            if 'hit_direction' in events_df.columns:
                # Pull events
                pull_ld = ld_events[ld_events['hit_direction'] == 'pull']
                pull_ld_counts = pull_ld.groupby('player_name').size().reset_index(name='Pull_LD_Events')
                
                # Oppo events
                oppo_ld = ld_events[ld_events['hit_direction'] == 'opposite']
                oppo_ld_counts = oppo_ld.groupby('player_name').size().reset_index(name='Oppo_LD_Events')
                
                # Calculate percentages
                pull_ld_pct = pd.merge(pull_ld_counts, ld_stats[['player_name', 'LD_Events']], on='player_name', how='right')
                pull_ld_pct['pull_ld_pct'] = (pull_ld_pct['Pull_LD_Events'].fillna(0) / pull_ld_pct['LD_Events']) * 100
                pull_ld_pct = pull_ld_pct[['player_name', 'pull_ld_pct']]
                
                oppo_ld_pct = pd.merge(oppo_ld_counts, ld_stats[['player_name', 'LD_Events']], on='player_name', how='right')
                oppo_ld_pct['oppo_ld_pct'] = (oppo_ld_pct['Oppo_LD_Events'].fillna(0) / oppo_ld_pct['LD_Events']) * 100
                oppo_ld_pct = oppo_ld_pct[['player_name', 'oppo_ld_pct']]
                
                # Merge with ld_stats
                ld_stats = pd.merge(ld_stats, pull_ld_pct, on='player_name', how='left')
                ld_stats = pd.merge(ld_stats, oppo_ld_pct, on='player_name', how='left')
            else:
                # Add empty columns if hit direction not available
                ld_stats['pull_ld_pct'] = 0
                ld_stats['oppo_ld_pct'] = 0
            
            # Calculate barrel percentages if available
            if all(col in events_df.columns for col in ['launch_speed', 'launch_angle']):
                # Apply barrel criteria to LD events
                ld_events['is_barrel'] = ld_events.apply(is_barrel, axis=1)
                
                # Count barrels by player
                ld_barrels = ld_events.groupby('player_name')['is_barrel'].sum().reset_index(name='ld_barrel_count')
                
                # Calculate barrel percentage
                ld_barrel_pct = pd.merge(ld_barrels, ld_stats[['player_name', 'LD_Events']], on='player_name', how='right')
                ld_barrel_pct['barrels/ld'] = (ld_barrel_pct['ld_barrel_count'].fillna(0) / ld_barrel_pct['LD_Events']) * 100
                ld_barrel_pct = ld_barrel_pct[['player_name', 'barrels/ld']]
                
                # Merge with ld_stats
                ld_stats = pd.merge(ld_stats, ld_barrel_pct, on='player_name', how='left')
            else:
                ld_stats['barrels/ld'] = 0
                
            # Merge with final results
            final_results = pd.merge(final_results, ld_stats.drop('LD_Events', axis=1), on='player_name', how='left')
        
        # Fill na values with 0
        columns_to_fill = ['avg_distance_fb', 'std_distance_fb', 'pull_fb_pct', 'oppo_fb_pct', 'barrels/fb',
                           'avg_distance_ld', 'std_distance_ld', 'pull_ld_pct', 'oppo_ld_pct', 'barrels/ld']
        for col in columns_to_fill:
            if col in final_results.columns:
                final_results[col] = final_results[col].fillna(0)
        
        # Calculate combined metrics (for convenience)
        if all(col in final_results.columns for col in ['avg_distance_fb', 'avg_distance_ld']):
            final_results['avg_dist_fb_ld'] = ((final_results['avg_distance_fb'] + final_results['avg_distance_ld']) / 2).round(2)
            
        if all(col in final_results.columns for col in ['std_distance_fb', 'std_distance_ld']):
            final_results['std_dist_fb_ld'] = ((final_results['std_distance_fb'] + final_results['std_distance_ld']) / 2).round(2)
        
        # Convert player names from "Last, First" to "First Last" format
        final_results['Name'] = final_results['player_name'].apply(
            lambda x: f"{x.split(', ')[1]} {x.split(', ')[0]}" if isinstance(x, str) and ', ' in x else x
        )
        
        # Get reference data from salaries for name standardization
        reference_df = get_salaries_for_reference()
        
        # Standardize player names to match other sheets
        if reference_df is not None and not reference_df.empty:
            logger.info("Standardizing event-level data player names to match salary data...")
            final_results = standardize_all_player_names(final_results, reference_df, 'Name')
        
        # Create variant names for better matching
        def create_alt_name(name):
            if not isinstance(name, str):
                return name
                
            # Handle name variants (accents, etc.)
            variants = {
                'Acuna': 'Acuña',
                'Munoz': 'Muñoz',
                'Pena': 'Peña',
                'Jimenez': 'Jiménez',
                'Suarez': 'Suárez',
                'Sanchez': 'Sánchez',
                'Hernandez': 'Hernández',
                'Gonzalez': 'González',
                'Martinez': 'Martínez',
                'Ramirez': 'Ramírez',
                'Rodriguez': 'Rodríguez',
            }
            
            alt_name = name
            for name1, name2 in variants.items():
                if name1 in alt_name:
                    alt_name = alt_name.replace(name1, name2)
                    return alt_name
                elif name2 in alt_name:
                    alt_name = alt_name.replace(name2, name1)
                    return alt_name
            
            return name
            
        final_results['NameAlt'] = final_results['Name'].apply(create_alt_name)
        
        # Add timestamp
        current_date = datetime.now().strftime("%Y-%m-%d")
        final_results['LastUpdated'] = current_date
        
        # Reorder columns with Name first for VLOOKUP
        column_order = ['Name', 'barrels/fb', 'barrels/ld', 'avg_distance_fb', 'avg_distance_ld', 
                        'std_distance_fb', 'std_distance_ld', 'pull_fb_pct', 'pull_ld_pct', 
                        'oppo_fb_pct', 'oppo_ld_pct', 'avg_dist_fb_ld', 'std_dist_fb_ld', 
                        'NameAlt', 'LastUpdated', 'player_name']
        
        # Ensure all columns exist in the result
        final_columns = [col for col in column_order if col in final_results.columns]
        
        # Add any columns that exist in the result but aren't in our order
        for col in final_results.columns:
            if col not in final_columns:
                final_columns.append(col)
        
        # Reorder columns
        final_results = final_results[final_columns]
        
        # Save the processed results
        timestamp = datetime.now().strftime("%Y%m%d")
        output_file = os.path.join(os.path.dirname(csv_file_path), f"fb_ld_stats_{timestamp}.csv")
        final_results.to_csv(output_file, index=False)
        
        logger.info(f"Saved processed FB+LD statistics for {len(final_results)} players to {output_file}")
        
        return {
            "success": True,
            "file_path": output_file,
            "player_count": len(final_results),
            "columns": final_results.columns.tolist()
        }
    
    except Exception as e:
        logger.error(f"Error processing FB+LD event data: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return {
            "success": False,
            "error": str(e)
        }

def update_excel_with_fb_ld_data(stats_csv_path, excel_path=EXCEL_FILE_PATH, sheet_name="StatcastEvents"):
    """
    Update the Excel file with FB+LD event statistics
    
    Parameters:
    -----------
    stats_csv_path : str
        Path to the CSV file with processed FB+LD statistics
    excel_path : str
        Path to the Excel file
    sheet_name : str
        Name of the sheet to update
    
    Returns:
    --------
    bool
        True if successful, False otherwise
    """
    
    logger.info(f"Updating Excel with FB+LD event statistics from {stats_csv_path}")
    
    try:
        # Read the CSV file
        stats_df = pd.read_csv(stats_csv_path)
        
        if stats_df.empty:
            logger.warning("FB+LD stats CSV is empty")
            return False
        
        # Create backup
        if not create_backup():
            logger.error("Update aborted due to backup failure")
            return False
        
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        
        # Remove existing sheet if it exists
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
            logger.info(f"Removed existing {sheet_name} sheet")
        
        # Create new sheet
        ws = wb.create_sheet(sheet_name)
        
        # Write column headers
        for col_idx, col_name in enumerate(stats_df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        # Write data rows
        for row_idx, row in enumerate(stats_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save the workbook
        wb.save(excel_path)
        
        logger.info(f"Successfully updated {sheet_name} sheet with {len(stats_df)} rows of FB+LD statistics")
        return True
        
    except Exception as e:
        logger.error(f"Error updating Excel with FB+LD data: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def run_fb_ld_update():
    """Run FB+LD event data update process"""
    logger.info("Starting FB+LD event data update")
    
    try:
        # Download FB+LD event data
        logger.info("Downloading FB+LD event data...")
        download_result = download_fb_ld_event_data()
        
        if download_result.get("success", False):
            logger.info("Successfully downloaded FB+LD event data")
            
            # Check if processing was successful
            if download_result.get("processed", {}).get("success", False):
                stats_file = download_result["processed"]["file_path"]
                
                # Update Excel with FB+LD statistics
                update_result = update_excel_with_fb_ld_data(
                    stats_csv_path=stats_file,
                    excel_path=EXCEL_FILE_PATH,
                    sheet_name="StatcastEvents"
                )
                
                if update_result:
                    logger.info("Successfully updated Excel with FB+LD event statistics")
                    return True
                else:
                    logger.warning("Failed to update Excel with FB+LD statistics")
            else:
                logger.warning("Failed to process FB+LD event data")
                logger.error(download_result.get("processed", {}).get("error", "Unknown processing error"))
        else:
            logger.warning("Failed to download FB+LD event data")
            logger.error(download_result.get("error", "Unknown download error"))
        
        return False
        
    except Exception as e:
        logger.error(f"Error in FB+LD update process: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False
    
def detect_excel_name_format(excel_path):
    """
    Detect the player name format used in the Excel file (first_last vs last_first)
    
    Parameters:
    -----------
    excel_path : str
        Path to the Excel file
    
    Returns:
    --------
    str
        "first_last" or "last_first"
    """
    logger.info(f"Detecting player name format in {excel_path}")
    
    try:
        # Try to read a sample of names from the Hitter or Pitcher sheets
        sample_sheet = None
        
        # Load workbook without parsing all data
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        
        # Try to find sheets with player names
        candidate_sheets = ['Hitter', 'Pitcher', 'Salaries', 'Handedness', 'FGHitters', 'FGPitchers']
        
        for sheet_name in candidate_sheets:
            if sheet_name in wb.sheetnames:
                sample_sheet = sheet_name
                break
                
        if not sample_sheet:
            logger.warning("Could not find a sheet with player names")
            return "first_last"  # Default to first_last if can't determine
        
        # Read the sheet
        ws = wb[sample_sheet]
        
        # Find the name column
        name_col_idx = None
        for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), 1):
            if cell.value == 'Name':
                name_col_idx = col_idx
                break
                
        if not name_col_idx:
            logger.warning(f"No 'Name' column found in {sample_sheet} sheet")
            return "first_last"  # Default to first_last if can't determine
            
        # Sample some names to determine format
        names = []
        for row in ws.iter_rows(min_row=2, min_col=name_col_idx, max_col=name_col_idx, max_row=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    names.append(cell.value)
        
        if not names:
            logger.warning("No valid names found in the sheet")
            return "first_last"  # Default to first_last if can't determine
            
        # Check if names typically contain commas
        comma_count = sum(1 for name in names if ',' in name)
        
        # If more than half the names have commas, assume last_first format
        if comma_count > len(names) / 2:
            logger.info(f"Detected 'last_first' format ({comma_count}/{len(names)} names have commas)")
            return "last_first"
        else:
            logger.info(f"Detected 'first_last' format ({len(names) - comma_count}/{len(names)} names without commas)")
            return "first_last"
            
    except Exception as e:
        logger.error(f"Error detecting name format: {str(e)}")
        return "first_last"  # Default to first_last format on error
    
def enhance_name_matching(statcast_df, excel_path):
    """
    Enhance name matching between Statcast data and existing data in the Excel file
    by creating additional lookup variants for problematic names
    
    Parameters:
    -----------
    statcast_df : pandas DataFrame
        DataFrame with Statcast data
    excel_path : str
        Path to the Excel file with existing player data
    
    Returns:
    --------
    pandas DataFrame
        Enhanced DataFrame with better name matching
    """
    logger.info("Enhancing name matching for Statcast data")
    
    if statcast_df.empty:
        return statcast_df
        
    result_df = statcast_df.copy()
    
    # Find name column
    name_col = None
    for col in ['Name', 'player_name']:
        if col in result_df.columns:
            name_col = col
            break
            
    if not name_col:
        logger.warning("No name column found for enhancement")
        return result_df
        
    # Add a column with alternative name formats for lookup
    result_df['NameAlt'] = result_df[name_col].copy()
    
    # Common name mapping issues to fix
    name_fixes = {
        # Accent/no accent variations
        'Acuna': 'Acuña',
        'Munoz': 'Muñoz',
        'Pena': 'Peña',
        'Valdez': 'Valdéz',
        'Jiminez': 'Jiménez',
        'Jimenez': 'Jiménez',
        'Suarez': 'Suárez',
        'Sanchez': 'Sánchez',
        'Hernandez': 'Hernández',
        'Gonzalez': 'González',
        'Martinez': 'Martínez',
        'Ramirez': 'Ramírez',
        'Rodriguez': 'Rodríguez',
        
        # Special cases
        'Luisangel Acuna': 'Luisangel Acuña',
        'Ronald Acuna': 'Ronald Acuña',
        'Ronald Acuna Jr.': 'Ronald Acuña Jr.',
    }
    
    # Apply fixes in both directions
    for name1, name2 in name_fixes.items():
        # Create bidirectional mapping
        result_df['NameAlt'] = result_df['NameAlt'].str.replace(
            name1, name2, regex=False
        )
        result_df['NameAlt'] = result_df['NameAlt'].str.replace(
            name2, name1, regex=False
        )
    
    # Try to load existing names from Excel to check for specific cases
    try:
        excel_names = set()
        
        # Try to read from Hitter, Pitcher, or Salaries sheet
        for sheet_name in ['Hitter', 'Pitcher', 'Salaries']:
            try:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
                if 'Name' in df.columns:
                    excel_names.update(df['Name'].dropna().astype(str).tolist())
            except Exception:
                continue
                
        if excel_names:
            logger.info(f"Collected {len(excel_names)} names from Excel for matching")
            
            # Find closest matches for each Statcast name
            from difflib import get_close_matches
            
            for idx, row in result_df.iterrows():
                name = row[name_col]
                if isinstance(name, str):
                    # Look for close matches in Excel names
                    matches = get_close_matches(name, excel_names, n=1, cutoff=0.8)
                    if matches:
                        excel_match = matches[0]
                        # If there's a difference but they're close, add as alternate
                        if excel_match != name:
                            result_df.at[idx, 'NameAlt'] = excel_match
                            logger.debug(f"Matched '{name}' to '{excel_match}'")
                            
    except Exception as match_e:
        logger.warning(f"Error in enhanced name matching: {str(match_e)}")
    
    logger.info("Name matching enhancement completed")
    return result_df
    
def update_barrel_data_only():
    """
    Update only the barrel data in MLBProjections without running the full update
    """
    logger.info("Starting barrel data-only update")
    
    try:
        # Create a backup
        if not create_backup():
            logger.error("Update aborted due to backup failure")
            return False
            
        # Download fresh barrel data (overwriting existing files)
        barrel_results = download_barrel_data_from_savant(overwrite=True)
        
        if barrel_results.get('processed', {}).get('success'):
            logger.info("Successfully downloaded and processed barrel data")
            
            # Update the Excel file with new barrel data
            barrel_file = barrel_results['processed'].get('file_path')
            
            # Update the Statcast sheet
            update_result = update_statcast_sheet_in_excel(
                excel_path=EXCEL_FILE_PATH,
                barrel_data_path=barrel_file,
                sheet_name="Statcast"
            )
            
            if update_result:
                logger.info("Successfully completed barrel data-only update")
                return True
            else:
                logger.warning("Failed to update Statcast sheet")
                return False
        else:
            # Try using the raw data
            if barrel_results.get('barrel_stats', {}).get('success'):
                barrel_file = barrel_results['barrel_stats'].get('file_path')
                update_result = update_statcast_sheet_in_excel(
                    excel_path=EXCEL_FILE_PATH,
                    barrel_data_path=barrel_file,
                    sheet_name="Statcast"
                )
                
                if update_result:
                    logger.info("Successfully completed barrel data-only update with raw data")
                    return True
            
            logger.warning("Failed to download or process barrel data")
            return False
            
    except Exception as e:
        logger.error(f"Error in barrel data-only update: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False
    
def standardize_statcast_player_names(df, format_type="first_last"):
    """
    Standardize player names in Statcast data to match the desired format
    
    Parameters:
    -----------
    df : pandas DataFrame
        DataFrame containing Statcast data with player names
    format_type : str
        Desired format: "first_last" (Kyle Tucker) or "last_first" (Tucker, Kyle)
    
    Returns:
    --------
    pandas DataFrame
        DataFrame with standardized player names
    """
    logger.info(f"Standardizing player names to {format_type} format")
    
    if df.empty:
        return df
        
    result_df = df.copy()
    
    # Find the appropriate name column
    name_col = None
    for col in ['Name', 'player_name', 'full_name', 'name']:
        if col in result_df.columns:
            name_col = col
            break
            
    if not name_col:
        logger.warning("No recognized name column found in the data")
        return result_df
        
    # Check the current format of a sample name
    if len(result_df) > 0:
        sample_name = result_df[name_col].iloc[0]
        
        # We'll check if the name has a comma to determine format
        if isinstance(sample_name, str):
            has_comma = ',' in sample_name
            
            # Convert from "Last, First" to "First Last"
            if has_comma and format_type == "first_last":
                logger.info("Converting names from 'Last, First' to 'First Last' format")
                result_df[name_col] = result_df[name_col].apply(
                    lambda x: f"{x.split(', ')[1]} {x.split(', ')[0]}" if isinstance(x, str) and ', ' in x else x
                )
                
            # Convert from "First Last" to "Last, First"
            elif not has_comma and format_type == "last_first":
                logger.info("Converting names from 'First Last' to 'Last, First' format")
                
                def convert_to_last_first(name):
                    if not isinstance(name, str) or ',' in name:
                        return name
                    
                    # Handle multi-word first names and last names
                    parts = name.strip().split()
                    if len(parts) < 2:
                        return name
                        
                    # Special cases with known suffixes
                    suffixes = ['Jr.', 'Sr.', 'III', 'II', 'IV']
                    suffix = ""
                    
                    # Check for suffix
                    if parts[-1] in suffixes:
                        suffix = parts.pop()
                    
                    # Assume last part is the last name
                    last_name = parts.pop()
                    first_name = ' '.join(parts)
                    
                    # Format with suffix if present
                    if suffix:
                        return f"{last_name}, {first_name} {suffix}"
                    else:
                        return f"{last_name}, {first_name}"
                
                result_df[name_col] = result_df[name_col].apply(convert_to_last_first)
    
    # Add additional handling for special cases
    special_cases = {
        # Examples of special case handling
        "Luisangel Acuna": "Acuña, Luisangel",
        "Ronald Acuna Jr.": "Acuña Jr., Ronald",
        "Acuna, Luisangel": "Luisangel Acuña",
        "Acuna Jr., Ronald": "Ronald Acuña Jr."
    }
    
    # Apply special case handling in the appropriate direction
    if format_type == "first_last":
        # Find and convert special last_first cases to first_last
        for last_first, first_last in special_cases.items():
            if ',' in last_first:  # Only process last_first format entries
                result_df[name_col] = result_df[name_col].replace(last_first, first_last)
    else:
        # Find and convert special first_last cases to last_first
        for first_last, last_first in special_cases.items():
            if ',' not in first_last:  # Only process first_last format entries
                result_df[name_col] = result_df[name_col].replace(first_last, last_first)
    
    # Handle accent marks and special characters
    def normalize_name(name):
        if not isinstance(name, str):
            return name
            
        # Import here to ensure it's available
        import unicodedata
        
        # Function to remove accents while preserving the base characters
        def remove_accents(text):
            return ''.join(c for c in unicodedata.normalize('NFKD', text) if not unicodedata.combining(c))
            
        # Create accent-normalized version (preserves base letters)
        normalized = remove_accents(name)
        
        return normalized
    
    # Create a secondary normalized name column for lookups
    norm_col = f"{name_col}_normalized"
    result_df[norm_col] = result_df[name_col].apply(normalize_name)
    
    logger.info(f"Standardized {len(result_df)} player names to {format_type} format")
    return result_df

def get_fangraphs_splits_data(url, split_type="Batter vs LHP", retry_count=2, timeout=120):
    """
    Scrape handedness splits data directly from FanGraphs with retry logic
    
    Parameters:
    url (str): FanGraphs splits leaderboard URL
    split_type (str): Type of split for logging ("Batter vs LHP", "Batter vs RHP", etc.)
    retry_count (int): Number of times to retry if initial attempt fails
    timeout (int): Timeout for page loading in seconds
    
    Returns:
    pandas DataFrame: DataFrame with splits data
    """
    logger.info(f"Scraping {split_type} splits data from FanGraphs")
    
    for attempt in range(retry_count + 1):
        driver = None
        try:
            # Setup headless Chrome browser
            chrome_options = Options()
            chrome_options.add_argument("--headless")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_argument("--ignore-certificate-errors")
            chrome_options.add_argument("--disable-extensions")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--log-level=3")
            chrome_options.add_argument("--disable-logging")
            
            # Initialize the driver
            driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
            driver.set_page_load_timeout(timeout)  # Extend timeout
            
            # Load the page
            logger.info(f"Loading page: {url}")
            driver.get(url)
            
            # Wait for the data table to load
            logger.info("Waiting for data table to load...")
            WebDriverWait(driver, timeout).until(
                EC.presence_of_element_located((By.CLASS_NAME, "table-fixed"))
            )
            
            # Additional wait to ensure full data loading
            time.sleep(5)
            
            # Extract the table data using pandas read_html
            logger.info("Extracting table data...")
            tables = pd.read_html(StringIO(driver.page_source))
            
            # FanGraphs typically has the main data table as the largest one
            if not tables:
                if attempt < retry_count:
                    logger.warning(f"No tables found, retrying ({attempt+1}/{retry_count+1})...")
                    driver.quit()
                    continue
                else:
                    logger.error("No tables found on the page after retries")
                    driver.quit()
                    return pd.DataFrame()
            
            # Find the main splits table (typically the largest)
            main_table = None
            max_rows = 0
            for i, table in enumerate(tables):
                rows = len(table)
                if rows > max_rows:
                    max_rows = rows
                    main_table = table
                    
            logger.info(f"Found main splits table with {max_rows} rows")
            
            # Process the data
            if main_table is not None:
                # Clean up the table - handle multilevel columns if present
                if isinstance(main_table.columns, pd.MultiIndex):
                    # FanGraphs sometimes has multi-level columns
                    main_table.columns = [' '.join(col).strip() for col in main_table.columns.values]
                
                # Standardize column names
                main_table.columns = [col.replace('%', 'Pct').replace('/', '_').replace(' ', '') for col in main_table.columns]
                
                # Add a split type column
                main_table['SplitType'] = split_type
                
                # Add prefix to all stat columns
                prefix = ""
                if split_type == "Batter vs LHP":
                    prefix = "vsL_"
                elif split_type == "Batter vs RHP":
                    prefix = "vsR_"
                elif split_type == "Pitcher vs LHH":
                    prefix = "vsL_"
                elif split_type == "Pitcher vs RHH":
                    prefix = "vsR_"
                    
                # Don't prefix certain columns
                non_prefixed_cols = ['Name', 'Team', 'SplitType', 'Season', 'playerid']
                
                # Apply prefixes
                for col in main_table.columns:
                    if col not in non_prefixed_cols and not col.startswith(prefix):
                        main_table.rename(columns={col: f"{prefix}{col}"}, inplace=True)
                
                logger.info(f"Successfully processed {len(main_table)} rows of {split_type} data")
                
                # Clean up
                driver.quit()
                return main_table
            else:
                if attempt < retry_count:
                    logger.warning(f"Could not identify the main data table, retrying ({attempt+1}/{retry_count+1})...")
                    driver.quit()
                    continue
                else:
                    logger.error("Could not identify the main data table after retries")
                    driver.quit()
                    return pd.DataFrame()
                
        except Exception as e:
            if attempt < retry_count:
                logger.warning(f"Error scraping {split_type} data (attempt {attempt+1}/{retry_count+1}): {str(e)}")
                if driver:
                    driver.quit()
                time.sleep(5)  # Wait before retrying
                continue
            else:
                logger.error(f"Error scraping {split_type} data: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                
                if driver:
                    driver.quit()
                    
                return pd.DataFrame()
    
    # This should only be reached if all retries fail
    return pd.DataFrame()

def get_handedness_splits(retry_attempts=3):
    """
    Get all handedness splits data from FanGraphs with improved fallback options
    
    Parameters:
    retry_attempts (int): Number of retry attempts for each split
    
    Returns:
    dict: Dictionary with DataFrames for different splits
    """
    logger.info("Getting handedness splits data from FanGraphs")
    
    # URLs for different splits - UPDATED with correct split IDs
    urls = {
        "Batter vs LHP": "https://www.fangraphs.com/leaders/splits-leaderboards?splitArr=1&splitArrPitch=&autoPt=false&splitTeams=false&statType=player&statgroup=1&startDate=2022-03-01&endDate=2025-11-01&players=&filter=G%7Cgt%7C1&groupBy=career&wxTemperature=&wxPressure=&wxAirDensity=&wxElevation=&wxWindSpeed=&position=B&sort=22,1&pageitems=2000000000&pg=0",
        "Batter vs RHP": "https://www.fangraphs.com/leaders/splits-leaderboards?splitArr=2&splitArrPitch=&autoPt=false&splitTeams=false&statType=player&statgroup=1&startDate=2022-03-01&endDate=2025-11-01&players=&filter=G%7Cgt%7C1&groupBy=career&wxTemperature=&wxPressure=&wxAirDensity=&wxElevation=&wxWindSpeed=&position=B&sort=22,1&pageitems=2000000000&pg=0",
        "Pitcher vs LHB": "https://www.fangraphs.com/leaders/splits-leaderboards?splitArr=5&splitArrPitch=&autoPt=false&splitTeams=false&statType=player&statgroup=1&startDate=2022-03-01&endDate=2025-11-01&players=&filter=G%7Cgt%7C1&groupBy=career&wxTemperature=&wxPressure=&wxAirDensity=&wxElevation=&wxWindSpeed=&position=P&sort=22,1&pageitems=2000000000&pg=0",
        "Pitcher vs RHB": "https://www.fangraphs.com/leaders/splits-leaderboards?splitArr=6&splitArrPitch=&autoPt=false&splitTeams=false&statType=player&statgroup=1&startDate=2022-03-01&endDate=2025-11-01&players=&filter=G%7Cgt%7C1&groupBy=career&wxTemperature=&wxPressure=&wxAirDensity=&wxElevation=&wxWindSpeed=&position=P&sort=22,1&pageitems=2000000000&pg=0"
    }
    
    results = {}
    
    # Get data for each split in a specific order:
    # Try batter data first (usually more reliable), then pitcher data
    split_order = ["Batter vs LHP", "Batter vs RHP", "Pitcher vs LHB", "Pitcher vs RHB"]
    
    for split_type in split_order:
        logger.info(f"Getting {split_type} data...")
        url = urls[split_type]
        
        # Special handling for potentially problematic data
        if split_type == "Pitcher vs RHB":
            # Try main URL first
            df = get_fangraphs_splits_data(url, split_type, retry_count=retry_attempts, timeout=180)
            
            if df.empty:
                # First fallback: Try alternative URL with smaller page size
                logger.warning(f"Failed to retrieve {split_type} data with main URL, trying alternative URL...")
                alt_url = alt_urls[split_type]
                df = get_fangraphs_splits_data(alt_url, split_type, retry_count=retry_attempts, timeout=120)
                
                if df.empty:
                    # Second fallback: Try mirror approach by modifying LHB data
                    logger.warning(f"Failed to retrieve {split_type} data with alternative URL, trying to use LHB data as template...")
                    if "Pitcher vs LHB" in results:
                        # Create synthetic RHB data based on LHB data
                        lhb_df = results["Pitcher vs LHB"].copy()
                        
                        # Change the prefix from vsL_ to vsR_
                        for col in lhb_df.columns:
                            if col.startswith("vsL_"):
                                new_col = col.replace("vsL_", "vsR_")
                                lhb_df.rename(columns={col: new_col}, inplace=True)
                        
                        # Change the split type
                        lhb_df["SplitType"] = "Pitcher vs RHB"
                        
                        # Use this as a placeholder - clearly mark it as synthetic
                        lhb_df["DataSource"] = "FanGraphs (Synthetic from LHB data)"
                        df = lhb_df
                        
                        logger.warning(f"Created synthetic {split_type} data as placeholder")
                    else:
                        logger.error(f"Unable to create synthetic data for {split_type}, no LHB data available")
        else:
            # Normal handling for other split types
            df = get_fangraphs_splits_data(url, split_type, retry_count=retry_attempts, timeout=120)
        
        if not df.empty:
            results[split_type] = df
            logger.info(f"Successfully retrieved {len(df)} rows of {split_type} data")
        else:
            logger.warning(f"Failed to retrieve {split_type} data after all attempts")
    
    # Make sure we have at least some data
    if not results:
        logger.error("Failed to retrieve any handedness splits data")
        return {}
    
    return results

def process_splits_data(splits_dict, reference_df=None):
    """
    Process and standardize handedness splits data
    
    Parameters:
    splits_dict (dict): Dictionary with splits DataFrames
    reference_df (pandas DataFrame): Reference DataFrame for name standardization
    
    Returns:
    dict: Dictionary with processed DataFrames
    """
    logger.info("Processing handedness splits data")
    
    processed = {}
    
    for split_type, df in splits_dict.items():
        if df.empty:
            processed[split_type] = df
            continue
            
        logger.info(f"Processing {split_type} data ({len(df)} rows)")
        
        # Make a copy to avoid modifying the original
        processed_df = df.copy()
        
        # Check for player ID column
        player_id_col = None
        for col in processed_df.columns:
            if col.lower() in ['playerid', 'mlbamid', 'id']:
                player_id_col = col
                break
        
        # Extract team information from the Team column
        if 'Team' in processed_df.columns:
            # FanGraphs often has the team in a format like "Team Totals (LAD)"
            def extract_team_code(team_text):
                if pd.isna(team_text) or not isinstance(team_text, str):
                    return ""
                
                # Try to extract team code from parentheses
                if '(' in team_text and ')' in team_text:
                    # Format: "Team Totals (LAD)" or similar
                    team_code = team_text.split('(')[-1].split(')')[0].strip()
                    if len(team_code) <= 3 and team_code.isalpha():
                        return team_code
                
                # If no team code in parentheses, just return the original
                return team_text
            
            processed_df['TeamCode'] = processed_df['Team'].apply(extract_team_code)
        
        # Standardize player names if reference data is available
        if reference_df is not None and 'Name' in processed_df.columns:
            processed_df = standardize_all_player_names(
                processed_df, 
                reference_df, 
                'Name', 
                player_id_col
            )
            logger.info(f"Standardized player names in {split_type} data")
        
        # Convert percentage columns to decimal format for consistency
        for col in processed_df.columns:
            # Look for percentage columns that might be in percent format (0-100)
            if any(suffix in col for suffix in ['Pct', 'PCT', '%']) or any(col.endswith(rate) for rate in ['rate', 'Rate']):
                if col in processed_df.columns:
                    # Convert column to numeric first to handle any string values
                    try:
                        # Convert to numeric, coercing errors to NaN
                        processed_df[col] = pd.to_numeric(processed_df[col], errors='coerce')
                        
                        # Check if values are already in decimal format (0-1)
                        # Get non-null values for checking
                        non_null_values = processed_df[col].dropna()
                        
                        if len(non_null_values) > 0:
                            max_value = non_null_values.max()
                            
                            if max_value <= 1.0:
                                # Already in decimal format
                                logger.info(f"Column {col} is already in decimal format")
                            else:
                                # Convert from percentage (0-100) to decimal (0-1)
                                processed_df[col] = processed_df[col] / 100
                                logger.info(f"Converted {col} from percentage to decimal format")
                        else:
                            logger.warning(f"Column {col} has no valid numeric values")
                            
                    except Exception as e:
                        logger.error(f"Error converting {col} to numeric: {str(e)}")
                        # If conversion fails, skip this column
                        continue
        
        # Calculate fantasy points
        if split_type.startswith("Batter"):
            # First check if we have the necessary columns for DK points
            required_cols = ['H', '2B', '3B', 'HR', 'RBI', 'R', 'BB', 'HBP', 'SB', 'G']
            
            # Check which required columns are missing and available
            missing_cols = []
            for col in required_cols:
                vsL_col = f"vsL_{col}"
                vsR_col = f"vsR_{col}"
                if col not in processed_df.columns and vsL_col not in processed_df.columns and vsR_col not in processed_df.columns:
                    missing_cols.append(col)
            
            if missing_cols:
                logger.warning(f"Missing columns for DK points calculation: {missing_cols}")
                # Try to find alternative column names
                col_mapping = {
                    'H': ['H', 'Hits', 'hits'],
                    '2B': ['2B', 'Doubles', 'doubles'],
                    '3B': ['3B', 'Triples', 'triples'],
                    'HR': ['HR', 'HomeRuns', 'Home Runs', 'homers'],
                    'RBI': ['RBI', 'Runs Batted In', 'runsBattedIn'],
                    'R': ['R', 'Runs', 'runs'],
                    'BB': ['BB', 'Walks', 'Base on Balls', 'walks'],
                    'HBP': ['HBP', 'Hit By Pitch', 'hitByPitch'],
                    'SB': ['SB', 'Stolen Bases', 'stolenBases'],
                    'G': ['G', 'Games', 'GP', 'gamesPlayed']
                }
                
                # Look for columns in both vsL_ and vsR_ prefixes
                for prefix in ['vsL_', 'vsR_', '']:
                    # Try to map columns
                    for missing_col in missing_cols:
                        for alt_col in col_mapping.get(missing_col, []):
                            prefixed_alt = f"{prefix}{alt_col}"
                            if prefixed_alt in processed_df.columns:
                                if prefix:
                                    # Found a prefixed alternative
                                    processed_df[f"{prefix}{missing_col}"] = processed_df[prefixed_alt]
                                    logger.info(f"Mapped {prefixed_alt} to {prefix}{missing_col}")
                                else:
                                    # Found an unprefixed alternative
                                    processed_df[missing_col] = processed_df[prefixed_alt]
                                    logger.info(f"Mapped {prefixed_alt} to {missing_col}")
                                break
            
            # Calculate DK points for both vsL and vsR splits
            for prefix in ['vsL_', 'vsR_']:
                prefix_cols = [col for col in processed_df.columns if col.startswith(prefix)]
                
                if prefix_cols:
                    # Create a temporary DataFrame with renamed columns for the DK points calculation
                    temp_df = processed_df.copy()
                    
                    # Rename columns to remove prefix for calculation
                    for col in prefix_cols:
                        unprefixed = col[len(prefix):]
                        temp_df[unprefixed] = temp_df[col]
                    
                    # Check if we have all necessary columns for DK points
                    has_required = all(col in temp_df.columns for col in required_cols)
                    
                    if has_required:
                        # Calculate DK points using the standard function
                        try:
                            temp_df = add_dk_points_to_dataframe(temp_df, "hitters")
                            
                            # Add the result back to the original DataFrame with the prefix
                            processed_df[f"{prefix}DK_Points"] = temp_df['DK_Points']
                            logger.info(f"Calculated DK points for {prefix} splits")
                        except Exception as dk_e:
                            logger.warning(f"Error calculating {prefix}DK_Points: {str(dk_e)}")
                    else:
                        logger.warning(f"Missing required columns for DK points calculation for {prefix} splits")
                        # Create a placeholder DK_Points column with zeros
                        processed_df[f"{prefix}DK_Points"] = 0
        
        elif split_type.startswith("Pitcher"):
            # First check if we have the necessary columns for DK points
            required_cols = ['IP', 'ER', 'H', 'BB', 'SO', 'W', 'G']
            
            # Check which required columns are missing and available
            missing_cols = []
            for col in required_cols:
                vsL_col = f"vsL_{col}"
                vsR_col = f"vsR_{col}"
                if col not in processed_df.columns and vsL_col not in processed_df.columns and vsR_col not in processed_df.columns:
                    missing_cols.append(col)
            
            if missing_cols:
                logger.warning(f"Missing columns for pitcher DK points calculation: {missing_cols}")
                # Try to find alternative column names
                col_mapping = {
                    'IP': ['IP', 'InningsPitched', 'Innings'],
                    'ER': ['ER', 'EarnedRuns', 'earnedRuns'],
                    'H': ['H', 'Hits', 'hits', 'HitsAllowed'],
                    'BB': ['BB', 'Walks', 'BasesOnBalls', 'walks'],
                    'SO': ['SO', 'K', 'Strikeouts', 'strikeouts'],
                    'W': ['W', 'Wins', 'wins'],
                    'G': ['G', 'Games', 'appearances']
                }
                
                # Look for columns in both vsL_ and vsR_ prefixes
                for prefix in ['vsL_', 'vsR_', '']:
                    # Try to map columns
                    for missing_col in missing_cols:
                        for alt_col in col_mapping.get(missing_col, []):
                            prefixed_alt = f"{prefix}{alt_col}"
                            if prefixed_alt in processed_df.columns:
                                if prefix:
                                    # Found a prefixed alternative
                                    processed_df[f"{prefix}{missing_col}"] = processed_df[prefixed_alt]
                                    logger.info(f"Mapped {prefixed_alt} to {prefix}{missing_col}")
                                else:
                                    # Found an unprefixed alternative
                                    processed_df[missing_col] = processed_df[prefixed_alt]
                                    logger.info(f"Mapped {prefixed_alt} to {missing_col}")
                                break
            
            # Calculate DK points for both vsL and vsR splits
            for prefix in ['vsL_', 'vsR_']:
                prefix_cols = [col for col in processed_df.columns if col.startswith(prefix)]
                
                if prefix_cols:
                    # Create a temporary DataFrame with renamed columns for the DK points calculation
                    temp_df = processed_df.copy()
                    
                    # Rename columns to remove prefix for calculation
                    for col in prefix_cols:
                        unprefixed = col[len(prefix):]
                        temp_df[unprefixed] = temp_df[col]
                    
                    # Check if we have all necessary columns for DK points
                    has_required = all(col in temp_df.columns for col in required_cols)
                    
                    if has_required:
                        # Calculate DK points using the standard function
                        try:
                            temp_df = add_dk_points_to_dataframe(temp_df, "pitchers")
                            
                            # Add the result back to the original DataFrame with the prefix
                            processed_df[f"{prefix}DK_Points"] = temp_df['DK_Points']
                            logger.info(f"Calculated DK points for {prefix} splits")
                        except Exception as dk_e:
                            logger.warning(f"Error calculating {prefix}DK_Points: {str(dk_e)}")
                    else:
                        logger.warning(f"Missing required columns for DK points calculation for {prefix} splits")
                        # Create a placeholder DK_Points column with zeros
                        processed_df[f"{prefix}DK_Points"] = 0
        
        # Add timestamp for when this data was updated
        processed_df['LastUpdated'] = datetime.now().strftime("%Y-%m-%d")
        
        # Add some metadata about the data source and date range
        processed_df['DataSource'] = 'FanGraphs'
        processed_df['YearRange'] = '2022-2025'
        processed_df['SplitType'] = split_type
        
        # Calculate sample size metrics for reference
        if split_type.startswith("Batter"):
            # Use PA or AB for batter sample size
            for prefix in ['vsL_', 'vsR_']:
                pa_col = f"{prefix}PA"
                ab_col = f"{prefix}AB"
                
                if pa_col in processed_df.columns:
                    processed_df[f"{prefix}SampleSize"] = processed_df[pa_col]
                elif ab_col in processed_df.columns:
                    processed_df[f"{prefix}SampleSize"] = processed_df[ab_col]
        else:
            # Use TBF or IP for pitcher sample size
            for prefix in ['vsL_', 'vsR_']:
                tbf_col = f"{prefix}TBF"
                ip_col = f"{prefix}IP"
                
                if tbf_col in processed_df.columns:
                    processed_df[f"{prefix}SampleSize"] = processed_df[tbf_col]
                elif ip_col in processed_df.columns:
                    # IP * 3 is a rough estimate of batters faced
                    processed_df[f"{prefix}SampleSize"] = processed_df[ip_col] * 3
        
        # Log stats about the processed data
        logger.info(f"Processed {len(processed_df)} rows for {split_type}")
        logger.info(f"Final columns: {len(processed_df.columns)} columns")
        
        processed[split_type] = processed_df
    
    return processed

def update_splits_sheets(excel_path, processed_splits):
    """
    Update Excel sheets with handedness splits data, preserving formulas
    
    Parameters:
    excel_path (str): Path to the Excel file
    processed_splits (dict): Dictionary with processed splits DataFrames
    
    Returns:
    bool: True if successful, False otherwise
    """
    logger.info(f"Updating Excel file {excel_path} with handedness splits data")
    
    try:
        # Define sheet names for each split type
        sheet_names = {
            "Batter vs LHP": "BatterVsLHP",
            "Batter vs RHP": "BatterVsRHP",
            "Pitcher vs LHB": "PitcherVsLHB",
            "Pitcher vs RHB": "PitcherVsRHB"
        }
        
        # Update each sheet
        for split_type, df in processed_splits.items():
            if df.empty:
                logger.warning(f"Skipping update for {split_type} - empty DataFrame")
                continue
                
            sheet_name = sheet_names.get(split_type)
            if not sheet_name:
                logger.warning(f"No sheet name defined for {split_type}")
                continue
                
            logger.info(f"Updating {sheet_name} sheet with {len(df)} rows")
            
            # Use formula-preserving function
            success = update_splits_data_with_formulas(sheet_name, df)
            
            if success:
                logger.info(f"Successfully updated {sheet_name} sheet with formula preservation")
            else:
                logger.warning(f"Failed to update {sheet_name} sheet")
        
        return True
        
    except Exception as e:
        logger.error(f"Error updating splits sheets: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def run_handedness_splits_update():
    """
    Main function to run the handedness splits update
    
    Returns:
    bool: True if successful, False otherwise
    """
    logger.info("Starting handedness splits update")
    
    try:
        # Create a backup first
        if not create_backup():
            logger.error("Update aborted due to backup failure")
            return False
        
        # Get reference data for name standardization
        reference_df = get_salaries_for_reference()
        
        # Get splits data
        splits_data = get_handedness_splits()
        
        # Check if we got any data
        if not splits_data:
            logger.error("Failed to retrieve any splits data")
            return False
        
        # Process the data
        processed_splits = process_splits_data(splits_data, reference_df)
        
        # Update Excel sheets
        success = update_splits_sheets(EXCEL_FILE_PATH, processed_splits)
        
        if success:
            logger.info("Successfully completed handedness splits update")
        else:
            logger.warning("Failed to update Excel sheets with splits data")
        
        return success
        
    except Exception as e:
        logger.error(f"Error in handedness splits update: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False
    
def calculate_adjusted_projections(excel_path=EXCEL_FILE_PATH, sheet_name="Hitter"):
    """
    Calculate adjusted projections based on various stats factors
    Uses dynamic league averages calculated from the current slate
    Uses xlwings to properly evaluate formulas in sheets
    
    Parameters:
    excel_path (str): Path to the Excel file
    sheet_name (str): Name of the sheet to update
    
    Returns:
    bool: True if successful, False otherwise
    """
    logger.info(f"Calculating adjusted projections for {sheet_name}")
    
    try:
        # Use xlwings to read the hitter data with formula evaluation
        logger.info("Using xlwings to read data with formula evaluation...")
        app = xw.App(visible=False)
        wb = app.books.open(excel_path)
        sheet = wb.sheets[sheet_name]
        
        # Get all data including headers
        data_range = sheet.used_range
        raw_data = data_range.value
        
        # Convert to pandas DataFrame
        headers = raw_data[0]
        data = raw_data[1:]
        hitter_df = pd.DataFrame(data, columns=headers)
        
        logger.info(f"Successfully read {len(hitter_df)} rows with xlwings (formulas evaluated)")
        
        # Close Excel
        wb.close()
        app.quit()
        
        
        # Check if we have the necessary columns for adjustments
        required_cols = ['xHR/FB', 'HR/FB', 'PitcherHR/FB%', 'Hard%', 'wOBA v P', 'wOBA v Batter', 
                        'ISO v P', 'Brls/FB', 'Brls/LD', 'Proj Pts', 'Salary']
        
        missing_cols = [col for col in required_cols if col not in hitter_df.columns]
        if missing_cols:
            logger.warning(f"Missing required columns: {missing_cols}")
            
            # look for alternative column names
            col_mapping = {
                'xHR/FB': ['xHR/FB', 'xHR_FB', 'xHR-FB'],
                'HR/FB': ['HR/FB', 'HR_FB', 'HR-FB'],
                'PitcherHR/FB%': ['PitcherHR/FB%', 'Pitcher HR/FB%', 'PitcherHR_FB%'],
                'Hard%': ['Hard%', 'HardPct', 'Hard Pct'],
                'wOBA v P': ['wOBA v P', 'wOBA_v_P', 'wOBA vs P'],
                'wOBA v Batter': ['wOBA v Batter', 'wOBA_v_Batter', 'wOBA vs Batter'],
                'ISO v P': ['ISO v P', 'ISO_v_P', 'ISO vs P'],
                'Brls/FB': ['Brls/FB', 'Barrels/FB', 'Brls_FB'],
                'Brls/LD': ['Brls/LD', 'Barrels/LD', 'Brls_LD'],
                'Proj Pts': ['Proj Pts', 'Projected Pts', 'Projected Points'],
                'Salary': ['Salary', 'DK Salary', 'DraftKings Salary']
            }
            
            # Create a refined column map
            actual_cols = {}
            for target_col, alternatives in col_mapping.items():
                for alt_col in alternatives:
                    if alt_col in hitter_df.columns:
                        actual_cols[target_col] = alt_col
                        logger.info(f"Found '{alt_col}' for '{target_col}'")
                        break
            
            # If we still can't find required columns, return False
            essential_cols = ['Proj Pts', 'Salary']
            missing_essential = [col for col in essential_cols if col not in actual_cols and col not in hitter_df.columns]
            if missing_essential:
                logger.error(f"Cannot proceed without essential columns: {missing_essential}")
                return False
        else:
            actual_cols = {col: col for col in required_cols}
        
        # Calculate league averages from the current slate
        league_averages = calculate_slate_averages(hitter_df)
        
        logger.info(f"Calculated league averages for slate:")
        for stat, avg in league_averages.items():
            logger.info(f"  {stat}: {avg:.3f}")
        
        # Define weights, tweak to your liking
        weights = {
            'xHR_FB_quality': 0.12,      # Quality of contact that should be HRs
            'HR_FB_actual': 0.08,        # Actual HR/FB performance
            'pitcher_hr_factor': 0.15,   # Pitcher's HR tendency
            'hard_hit_factor': 0.20,     # Quality of contact
            'woba_matchup_factor': 0.25, # Overall hitting ability matchup
            'iso_factor': 0.12,          # Power matchup
            'barrel_factor': 0.08        # Barrel consistency
        }
        
        # Calculate adjustment factors for each row
        adjustments = []
        adjustment_components = []  # Store individual components for analysis
        
        for idx, row in hitter_df.iterrows():
            # Initialize adjustment to 1.0 (no change)
            adjustment = 1.0
            components = {}
            
            # Check if we have enough data to make adjustments
            has_data = False
            
            # 1. xHR/FB as a quality of contact indicator
            xhr_fb_col = actual_cols.get('xHR/FB')
            if xhr_fb_col and pd.notna(row.get(xhr_fb_col)):
                try:
                    xhr_fb = float(row[xhr_fb_col])
                    if xhr_fb > 0:
                        has_data = True
                        xhr_fb_factor = xhr_fb / league_averages['xHR/FB']
                        components['xHR_FB_quality'] = xhr_fb_factor
                        adjustment *= (1 + weights['xHR_FB_quality'] * (xhr_fb_factor - 1))
                except (ValueError, TypeError):
                    pass
            
            # 2. Actual HR/FB as a separate performance factor
            hr_fb_col = actual_cols.get('HR/FB')
            if hr_fb_col and pd.notna(row.get(hr_fb_col)):
                try:
                    hr_fb = float(row[hr_fb_col])
                    if hr_fb > 0:
                        has_data = True
                        hr_fb_factor = hr_fb / league_averages['HR/FB']
                        components['HR_FB_actual'] = hr_fb_factor
                        adjustment *= (1 + weights['HR_FB_actual'] * (hr_fb_factor - 1))
                except (ValueError, TypeError):
                    pass
            
            # 3. Pitcher HR/FB% (higher is better for batters)
            pitcher_hr_fb_col = actual_cols.get('PitcherHR/FB%')
            if pitcher_hr_fb_col and pd.notna(row.get(pitcher_hr_fb_col)):
                try:
                    pitcher_hr_fb = float(row[pitcher_hr_fb_col])
                    league_avg = league_averages['PitcherHR/FB%']
                    
                    if pitcher_hr_fb > 0 and league_avg > 0:
                        has_data = True
                        # Higher pitcher HR/FB% is better for batters
                        pitcher_hr_factor = pitcher_hr_fb / league_avg
                    elif pitcher_hr_fb == 0:
                        pitcher_hr_factor = 0.3  # Strong negative adjustment
                    else:
                        pitcher_hr_factor = 1.0  # Neutral if we can't calculate
                    
                    components['pitcher_hr'] = pitcher_hr_factor
                    adjustment *= (1 + weights['pitcher_hr_factor'] * (pitcher_hr_factor - 1))
                except (ValueError, TypeError):
                    pass
            
            # 4. Hard% (higher is better)
            hard_pct_col = actual_cols.get('Hard%')
            if hard_pct_col and pd.notna(row.get(hard_pct_col)):
                try:
                    hard_pct = float(row[hard_pct_col])
                    if hard_pct > 0:
                        has_data = True
                        hard_hit_factor = hard_pct / league_averages['Hard%']
                        hard_hit_factor = max(0.7, min(1.3, hard_hit_factor))  # Constrain range
                        components['hard_hit'] = hard_hit_factor
                        adjustment *= (1 + weights['hard_hit_factor'] * (hard_hit_factor - 1))
                except (ValueError, TypeError):
                    pass
            
            # 5. wOBA matchup
            woba_p_col = actual_cols.get('wOBA v P')
            woba_b_col = actual_cols.get('wOBA v Batter')
            
            if woba_p_col and woba_b_col:
                try:
                    woba_p = float(row.get(woba_p_col, 0))
                    woba_b = float(row.get(woba_b_col, 0))
                    
                    if pd.notna(woba_p) and pd.notna(woba_b) and woba_p > 0 and woba_b > 0:
                        has_data = True
                        batter_woba_factor = woba_p / league_averages['wOBA_v_P']
                        pitcher_woba_factor = league_averages['wOBA_v_Batter'] / woba_b
                        woba_matchup_factor = (batter_woba_factor * 0.6 + pitcher_woba_factor * 0.4)
                        woba_matchup_factor = max(0.7, min(1.3, woba_matchup_factor))
                        components['woba_matchup'] = woba_matchup_factor
                        adjustment *= (1 + weights['woba_matchup_factor'] * (woba_matchup_factor - 1))
                except (ValueError, TypeError):
                    pass
            
            # 6. ISO (Isolated Power) matchup
            iso_col = actual_cols.get('ISO v P')
            if iso_col and pd.notna(row.get(iso_col)):
                try:
                    iso = float(row[iso_col])
                    if iso >= 0:  # ISO can be 0
                        has_data = True
                        iso_factor = iso / league_averages['ISO_v_P'] if league_averages['ISO_v_P'] > 0 else 1.0
                        iso_factor = max(0.5, min(1.5, iso_factor))
                        components['iso'] = iso_factor
                        adjustment *= (1 + weights['iso_factor'] * (iso_factor - 1))
                except (ValueError, TypeError):
                    pass
            
            # 7. Barrel rates
            brls_fb_col = actual_cols.get('Brls/FB')
            brls_ld_col = actual_cols.get('Brls/LD')
            
            if brls_fb_col and brls_ld_col:
                try:
                    brls_fb = float(row.get(brls_fb_col, 0))
                    brls_ld = float(row.get(brls_ld_col, 0))
                    
                    if pd.notna(brls_fb) and pd.notna(brls_ld):
                        has_data = True
                        combined_barrel_rate = (brls_fb * 0.7 + brls_ld * 0.3)
                        if combined_barrel_rate > 0 and league_averages['Brls/FB'] > 0:
                            barrel_factor = combined_barrel_rate / league_averages['Brls/FB']
                            barrel_factor = max(0.5, min(1.5, barrel_factor))
                            components['barrel'] = barrel_factor
                            adjustment *= (1 + weights['barrel_factor'] * (barrel_factor - 1))
                except (ValueError, TypeError):
                    pass
            
            # If no data was found for adjustment, keep at 1.0
            if not has_data:
                logger.debug(f"No adjustment data found for player at index {idx} ({row.get('Name', 'Unknown')})")
            else:
                # Cap the total adjustment to reasonable bounds
                adjustment = max(0.6, min(1.6, adjustment))
            
            adjustments.append(adjustment)
            adjustment_components.append(components)
        
        # Add adjustment column to DataFrame
        hitter_df['Adjustment'] = adjustments
        
        # Calculate adjusted points
        proj_pts_col = actual_cols.get('Proj Pts', 'Proj Pts')
        if proj_pts_col in hitter_df.columns:
            hitter_df['AdjPoints'] = hitter_df.apply(
                lambda row: float(row[proj_pts_col]) * row['Adjustment'] if pd.notna(row[proj_pts_col]) else 0,
                axis=1
            )
        else:
            logger.error(f"Could not find projection points column")
            return False
        
        # Calculate adjusted value
        salary_col = actual_cols.get('Salary', 'Salary')
        if salary_col in hitter_df.columns:
            # Prevent division by zero
            hitter_df['AdjValue'] = hitter_df.apply(
                lambda row: (row['AdjPoints'] * 1000) / float(row[salary_col]) if float(row[salary_col]) > 0 else 0,
                axis=1
            )
        else:
            logger.error(f"Could not find salary column")
            return False
        
        # Update the Excel file with new columns
        update_hitter_sheet_with_adjustments(excel_path, hitter_df, sheet_name)
        
        # Log statistics
        valid_adjustments = [adj for adj in adjustments if adj != 1.0]
        if valid_adjustments:
            logger.info(f"Adjustment statistics (excluding unchanged):")
            logger.info(f"  Mean adjustment: {np.mean(valid_adjustments):.3f}")
            logger.info(f"  Median adjustment: {np.median(valid_adjustments):.3f}")
            logger.info(f"  Std dev: {np.std(valid_adjustments):.3f}")
            logger.info(f"  Min adjustment: {np.min(valid_adjustments):.3f}")
            logger.info(f"  Max adjustment: {np.max(valid_adjustments):.3f}")
            
            # Log top adjustments
            sorted_idx = np.argsort(adjustments)
            logger.info("Top 5 positive adjustments:")
            for idx in sorted_idx[-5:][::-1]:
                if adjustments[idx] != 1.0:
                    logger.info(f"  {hitter_df.iloc[idx]['Name']}: {adjustments[idx]:.3f}")
            
            logger.info("Bottom 5 negative adjustments:")
            for idx in sorted_idx[:5]:
                if adjustments[idx] != 1.0:
                    logger.info(f"  {hitter_df.iloc[idx]['Name']}: {adjustments[idx]:.3f}")
        else:
            logger.warning("No adjustments were made - all players kept at 1.0")
            logger.info(f"Number of players processed: {len(hitter_df)}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error calculating adjusted projections: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def update_hitter_sheet_with_adjustments(excel_path, hitter_df, sheet_name):
    """
    Update the hitter sheet with adjustment calculations while preserving formulas
    Uses xlwings for better Excel integration
    
    Parameters:
    excel_path (str): Path to the Excel file
    hitter_df (pandas DataFrame): DataFrame with adjustments
    sheet_name (str): Name of the sheet to update
    
    Returns:
    bool: True if successful, False otherwise
    """
    logger.info(f"Updating {sheet_name} with adjustment calculations")
    
    app = None
    wb = None
    
    try:
        # Use xlwings to update the data
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        
        wb = app.books.open(excel_path)
        
        # Check if the sheet exists
        if sheet_name not in [sheet.name for sheet in wb.sheets]:
            logger.error(f"Sheet '{sheet_name}' not found in workbook")
            return False
            
        sheet = wb.sheets[sheet_name]
        
        # Find the last column with data
        try:
            # Try to get the used range
            used_range = sheet.used_range
            if used_range is not None:
                last_col = used_range.last_cell.column
            else:
                # Fallback: check first row for headers
                last_col = 1
                for i in range(1, 100):  # Check up to column 100
                    if sheet.range((1, i)).value is None:
                        break
                    last_col = i
        except:
            # If all else fails, use a conservative estimate
            last_col = 50
            logger.warning(f"Could not determine last column, using {last_col}")
        
        # Check if adjustment columns already exist
        headers = []
        try:
            for i in range(1, last_col + 1):
                header = sheet.range((1, i)).value
                if header is not None:
                    headers.append(header)
                else:
                    break
        except:
            logger.warning("Error reading headers, continuing with empty list")
            headers = []
        
        adj_col = None
        adj_pts_col = None  
        adj_val_col = None
        
        # Find existing columns
        for i, header in enumerate(headers):
            if header == 'Adjustment':
                adj_col = i + 1
            elif header == 'AdjPoints':
                adj_pts_col = i + 1
            elif header == 'AdjValue':
                adj_val_col = i + 1
        
        # Add new columns if they don't exist
        if adj_col is None:
            adj_col = len(headers) + 1
            sheet.range((1, adj_col)).value = 'Adjustment'
            
        if adj_pts_col is None:
            adj_pts_col = adj_col + 1
            sheet.range((1, adj_pts_col)).value = 'AdjPoints'
            
        if adj_val_col is None:
            adj_val_col = adj_pts_col + 1
            sheet.range((1, adj_val_col)).value = 'AdjValue'
        
        # Write the data
        data_rows = len(hitter_df)
        logger.info(f"Writing {data_rows} rows of adjustment data to columns {adj_col}, {adj_pts_col}, {adj_val_col}")
        
        # Write the adjustment values using proper range specification
        if 'Adjustment' in hitter_df.columns:
            # Convert to list of lists for proper column writing
            adj_values = [[value] for value in hitter_df['Adjustment'].tolist()]
            sheet.range((2, adj_col)).value = adj_values
            
            # Log first few values for debugging
            for i in range(min(5, len(adj_values))):
                logger.info(f"Row {i+2}: Adjustment = {adj_values[i][0]}")
        
        if 'AdjPoints' in hitter_df.columns:
            adj_points_values = [[value] for value in hitter_df['AdjPoints'].tolist()]
            sheet.range((2, adj_pts_col)).value = adj_points_values
            
        if 'AdjValue' in hitter_df.columns:
            adj_value_values = [[value] for value in hitter_df['AdjValue'].tolist()]
            sheet.range((2, adj_val_col)).value = adj_value_values
        
        # Save the workbook
        wb.save()
        
        logger.info(f"Successfully updated {sheet_name} with adjustments")
        return True
        
    except Exception as e:
        logger.error(f"Error updating hitter sheet with adjustments: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False
        
    finally:
        # Clean up
        try:
            if wb:
                wb.close()
            if app:
                app.quit()
        except:
            pass

def calculate_slate_averages(hitter_df):
    """
    Calculate league averages from the current slate of games
    Handles data types from xlwings properly
    
    Parameters:
    hitter_df (pandas DataFrame): DataFrame with hitter data
    
    Returns:
    dict: Dictionary of league averages for different stats
    """
    averages = {}
    
    # For hitter stats - average all values
    hitter_stats = ['xHR/FB', 'HR/FB', 'Hard%', 'wOBA v P', 'ISO v P', 'Brls/FB', 'Brls/LD']
    for stat in hitter_stats:
        if stat in hitter_df.columns:
            # Convert to numeric and filter out NaN and zero values
            valid_values = []
            for value in hitter_df[stat]:
                try:
                    # Convert to float, handling string values that might come from Excel
                    if pd.notna(value) and value != "":
                        numeric_value = float(value)
                        if stat != 'ISO v P' and numeric_value > 0:  # ISO can be 0
                            valid_values.append(numeric_value)
                        elif stat == 'ISO v P' and numeric_value >= 0:
                            valid_values.append(numeric_value)
                except (ValueError, TypeError):
                    continue
            
            if valid_values:
                averages[stat] = np.mean(valid_values)
            else:
                # Fallback to typical MLB averages if no valid values
                fallbacks = {
                    'xHR/FB': 12.0,
                    'HR/FB': 14.0,
                    'Hard%': 35.0,
                    'wOBA v P': 0.320,
                    'ISO v P': 0.165,
                    'Brls/FB': 7.0,
                    'Brls/LD': 25.0
                }
                averages[stat] = fallbacks.get(stat, 1.0)
                logger.warning(f"No valid values for {stat}, using fallback: {averages[stat]}")
    
    # For pitcher stats - get unique values and average them
    pitcher_stats = ['PitcherHR/FB%', 'wOBA v Batter', 'PitcherFB%']
    for stat in pitcher_stats:
        if stat in hitter_df.columns:
            # Get unique values (one per pitcher)
            unique_values = []
            seen_values = set()
            
            for value in hitter_df[stat]:
                try:
                    if pd.notna(value) and value != "":
                        numeric_value = float(value)
                        if numeric_value > 0 and numeric_value not in seen_values:
                            unique_values.append(numeric_value)
                            seen_values.add(numeric_value)
                except (ValueError, TypeError):
                    continue
            
            if unique_values:
                averages[stat] = np.mean(unique_values)
            else:
                # Fallback to typical MLB averages
                fallbacks = {
                    'PitcherHR/FB%': 14.0,
                    'wOBA v Batter': 0.320,
                    'PitcherFB%': 40.0
                }
                averages[stat] = fallbacks.get(stat, 1.0)
                logger.warning(f"No valid values for {stat}, using fallback: {averages[stat]}")
    
    # Ensure all needed stats have values (add fallbacks with approx league averages if missing)
    required_stats = {
        'xHR/FB': 12.0,
        'HR/FB': 14.0,
        'PitcherHR/FB%': 14.0,
        'Hard%': 35.0,
        'wOBA v P': 0.320,
        'wOBA v Batter': 0.320,
        'ISO v P': 0.165,
        'Brls/FB': 7.0,
        'Brls/LD': 25.0,
        'PitcherFB%': 40.0
    }
    
    for stat, fallback in required_stats.items():
        if stat not in averages:
            averages[stat] = fallback
            logger.info(f"Using fallback value for missing stat {stat}: {fallback}")
    
    # Standardize key names for internal use
    averages['wOBA_v_P'] = averages.get('wOBA v P', averages.get('wOBA_v_P', 0.320))
    averages['wOBA_v_Batter'] = averages.get('wOBA v Batter', averages.get('wOBA_v_Batter', 0.320))
    averages['ISO_v_P'] = averages.get('ISO v P', averages.get('ISO_v_P', 0.165))
    
    # Log the calculated averages
    logger.info("Calculated slate averages:")
    for stat, value in averages.items():
        logger.info(f"  {stat}: {value:.3f}")
    
    return averages

def scrape_mlb_starting_lineups(date_str=None):
    """
    Scrape MLB starting lineups from MLB.com
    
    Parameters:
    date_str (str): Date in YYYY-MM-DD format. If None, uses today's date
    
    Returns:
    pandas DataFrame: Starting lineups data
    """
    logger.info("Scraping MLB starting lineups from MLB.com")
    
    # Use today's date if none provided
    if date_str is None:
        date_str = datetime.now().strftime("%Y-%m-%d")
    
    url = f"https://www.mlb.com/starting-lineups/{date_str}"
    logger.info(f"Scraping URL: {url}")
    
    try:
        # Setup selenium
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
        
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.get(url)
        
        # Wait for page to load
        time.sleep(5)
        
        # Wait for the lineups to load
        try:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "starting-lineups__matchup"))
            )
            logger.info("Found starting-lineups__matchup elements")
        except:
            logger.warning("Timeout waiting for lineups to load")
            # Try to save a screenshot for debugging
            try:
                driver.save_screenshot("lineup_debug.png")
                logger.info("Saved screenshot for debugging")
            except:
                pass
        
        # Get all matchup containers
        matchups = driver.find_elements(By.CLASS_NAME, "starting-lineups__matchup")
        logger.info(f"Found {len(matchups)} matchups")
        
        all_lineups = []
        
        for i, matchup in enumerate(matchups):
            try:
                logger.info(f"Processing matchup {i+1}/{len(matchups)}")
                
                # Get team names from the new structure
                team_names_container = matchup.find_element(By.CLASS_NAME, "starting-lineups__team-names")
                
                # Get away team
                away_team_element = team_names_container.find_element(By.CLASS_NAME, "starting-lineups__team-name--away")
                away_team_link = away_team_element.find_element(By.TAG_NAME, "a")
                away_team = away_team_link.text.strip()
                away_team_code = away_team_link.get_attribute("data-tri-code")
                
                # Get home team
                home_team_element = team_names_container.find_element(By.CLASS_NAME, "starting-lineups__team-name--home")
                home_team_link = home_team_element.find_element(By.TAG_NAME, "a")
                home_team = home_team_link.text.strip()
                home_team_code = home_team_link.get_attribute("data-tri-code")
                
                logger.info(f"Found game: {away_team} ({away_team_code}) @ {home_team} ({home_team_code})")
                
                # Get game time/status
                try:
                    game_time_element = matchup.find_element(By.CLASS_NAME, "starting-lineups__game-date-time")
                    game_status = game_time_element.text.strip()
                except:
                    game_status = "TBD"
                
                # Get starting pitchers
                pitchers_container = matchup.find_element(By.CLASS_NAME, "starting-lineups__pitchers")
                pitcher_summaries = pitchers_container.find_elements(By.CLASS_NAME, "starting-lineups__pitcher-summary")
                
                # First pitcher is away, second is home
                away_pitcher = "TBD"
                home_pitcher = "TBD"
                
                if len(pitcher_summaries) > 0:
                    try:
                        away_pitcher_name = pitcher_summaries[0].find_element(By.CLASS_NAME, "starting-lineups__pitcher-name")
                        away_pitcher = away_pitcher_name.text.strip()
                    except:
                        away_pitcher = "TBD"
                
                if len(pitcher_summaries) > 1:
                    try:
                        home_pitcher_name = pitcher_summaries[1].find_element(By.CLASS_NAME, "starting-lineups__pitcher-name")
                        home_pitcher = home_pitcher_name.text.strip()
                    except:
                        home_pitcher = "TBD"
                
                logger.info(f"Pitchers: {away_pitcher} vs {home_pitcher}")
                
                # Get lineups - looks for the first teams container (for full names)
                teams_containers = matchup.find_elements(By.CLASS_NAME, "starting-lineups__teams")
                if teams_containers:
                    teams_container = teams_containers[0]  # Use the first one which has full names
                    
                    # Get away lineup
                    away_lineup = teams_container.find_element(By.CLASS_NAME, "starting-lineups__team--away")
                    away_players = away_lineup.find_elements(By.CLASS_NAME, "starting-lineups__player")
                    
                    logger.info(f"Processing {len(away_players)} away players")
                    
                    for j, player_element in enumerate(away_players):
                        try:
                            # Get player name from the link
                            player_link = player_element.find_element(By.CLASS_NAME, "starting-lineups__player--link")
                            player_name = player_link.text.strip()
                            
                            # Get position/handedness from the span
                            position_element = player_element.find_element(By.CLASS_NAME, "starting-lineups__player--position")
                            position_text = position_element.text.strip()
                            
                            # Parse position text like "(R) DH" or "(L) 2B"
                            if position_text.startswith("(") and ")" in position_text:
                                parts = position_text.split(")")
                                handedness = parts[0].strip("(").strip()
                                position = parts[1].strip()
                            else:
                                handedness = ""
                                position = position_text
                            
                            all_lineups.append({
                                'Game': f"{away_team} @ {home_team}",
                                'GameTime': game_status,
                                'Team': away_team_code,
                                'TeamType': 'Away',
                                'BattingOrder': j + 1,
                                'Player': player_name,
                                'Position': position,
                                'Handedness': handedness,
                                'Opponent': home_team_code,
                                'OpposingPitcher': home_pitcher,
                                'StartingPitcher': away_pitcher if position == 'P' else None
                            })
                            logger.info(f"Added away player: {player_name} ({position})")
                        except Exception as e:
                            logger.error(f"Error parsing away player {j+1}: {str(e)}")
                    
                    # Get home lineup
                    home_lineup = teams_container.find_element(By.CLASS_NAME, "starting-lineups__team--home")
                    home_players = home_lineup.find_elements(By.CLASS_NAME, "starting-lineups__player")
                    
                    logger.info(f"Processing {len(home_players)} home players")
                    
                    for j, player_element in enumerate(home_players):
                        try:
                            # Get player name from the link
                            player_link = player_element.find_element(By.CLASS_NAME, "starting-lineups__player--link")
                            player_name = player_link.text.strip()
                            
                            # Get position/handedness from the span
                            position_element = player_element.find_element(By.CLASS_NAME, "starting-lineups__player--position")
                            position_text = position_element.text.strip()
                            
                            # Parse position text like "(R) DH" or "(L) 2B"
                            if position_text.startswith("(") and ")" in position_text:
                                parts = position_text.split(")")
                                handedness = parts[0].strip("(").strip()
                                position = parts[1].strip()
                            else:
                                handedness = ""
                                position = position_text
                            
                            all_lineups.append({
                                'Game': f"{away_team} @ {home_team}",
                                'GameTime': game_status,
                                'Team': home_team_code,
                                'TeamType': 'Home',
                                'BattingOrder': j + 1,
                                'Player': player_name,
                                'Position': position,
                                'Handedness': handedness,
                                'Opponent': away_team_code,
                                'OpposingPitcher': away_pitcher,
                                'StartingPitcher': home_pitcher if position == 'P' else None
                            })
                            logger.info(f"Added home player: {player_name} ({position})")
                        except Exception as e:
                            logger.error(f"Error parsing home player {j+1}: {str(e)}")
                else:
                    logger.warning(f"Could not find lineups for matchup {i+1}")
                
            except Exception as e:
                logger.error(f"Error parsing matchup {i+1}: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                continue
        
        driver.quit()
        
        logger.info(f"Total lineup entries collected: {len(all_lineups)}")
        
        # Convert to DF
        if all_lineups:
            df = pd.DataFrame(all_lineups)
            
            # Add some additional useful columns
            df['Date'] = date_str
            df['IsHome'] = df['TeamType'] == 'Home'
            
            logger.info(f"Successfully scraped {len(df)} lineup entries")
            return df
        else:
            logger.warning("No lineup data found")
            return pd.DataFrame()
            
    except Exception as e:
        logger.error(f"Error scraping MLB lineups: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return pd.DataFrame()


def update_starting_lineups_sheet(excel_path=EXCEL_FILE_PATH, date_str=None):
    """
    Update or create TodaysStartingLineups sheet with scraped data
    
    Parameters:
    excel_path (str): Path to the Excel file
    date_str (str): Date to scrape lineups for (YYYY-MM-DD format)
    
    Returns:
    bool: True if successful, False otherwise
    """
    logger.info("Updating TodaysStartingLineups sheet")
    
    try:
        # Scrape the lineups
        lineups_df = scrape_mlb_starting_lineups(date_str)
        
        if lineups_df.empty:
            logger.warning("No lineup data to update")
            return False
        
        # Get reference data for name standardization
        reference_df = get_salaries_for_reference()
        
        # Standardize player names
        if reference_df is not None and not reference_df.empty:
            lineups_df = standardize_all_player_names(lineups_df, reference_df, 'Player')
            logger.info("Standardized player names in lineup data")
        
        # Update Excel sheet with formula preservation
        success = update_starting_lineups_with_formulas(excel_path, lineups_df)
        
        if success:
            logger.info(f"Successfully updated TodaysStartingLineups with {len(lineups_df)} entries")
        else:
            logger.warning("Failed to update TodaysStartingLineups sheet")
            
        return success
        
    except Exception as e:
        logger.error(f"Error updating starting lineups sheet: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False
    
def update_starting_lineups_with_formulas(excel_path, lineups_df, sheet_name="TodaysStartingLineups"):
    """
    Update TodaysStartingLineups sheet while preserving any existing formulas (like VLOOKUP)
    
    Parameters:
    excel_path (str): Path to the Excel file
    lineups_df (pandas DataFrame): New lineup data
    sheet_name (str): Name of the sheet to update
    
    Returns:
    bool: True if successful, False otherwise
    """
    logger.info(f"Updating {sheet_name} with formula preservation...")
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        
        # Check if the sheet exists
        if sheet_name not in wb.sheetnames:
            logger.info(f"Creating new sheet '{sheet_name}'")
            sheet = wb.create_sheet(sheet_name)
            
            # Write headers
            for col_idx, col_name in enumerate(lineups_df.columns, 1):
                sheet.cell(row=1, column=col_idx, value=col_name)
            
            # Write data
            for row_idx, row in enumerate(lineups_df.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=value)
        else:
            # Get the existing sheet
            sheet = wb[sheet_name]
            
            # Store all existing data and formulas first
            original_data = {}
            formula_columns = set()
            
            # Check all cells to identify formulas
            for row in range(1, sheet.max_row + 1):
                row_data = {}
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    is_formula = False
                    formula_value = None
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        is_formula = True
                        formula_value = cell.value
                        formula_columns.add(col)
                    
                    row_data[col] = {
                        'value': cell.value,
                        'is_formula': is_formula,
                        'formula_value': formula_value
                    }
                
                original_data[row] = row_data
            
            logger.info(f"Found formulas in columns: {sorted(formula_columns)}")
            
            # Get current headers and map to dataframe columns
            header_map = {}
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                if header:
                    header_map[header] = col
            
            # Identify which columns to update (those without formulas)
            data_columns = []
            for col_name in lineups_df.columns:
                if col_name in header_map:
                    col_idx = header_map[col_name]
                    if col_idx not in formula_columns:
                        data_columns.append((col_name, col_idx))
            
            logger.info(f"Will update data columns: {[col[0] for col in data_columns]}")
            
            # Clear only the data columns we'll update
            for row in range(2, sheet.max_row + 1):
                for col_name, col_idx in data_columns:
                    sheet.cell(row=row, column=col_idx).value = None
            
            # Write new data
            for row_idx, (_, row) in enumerate(lineups_df.iterrows(), start=2):
                for col_name, col_idx in data_columns:
                    if col_name in row.index:
                        sheet.cell(row=row_idx, column=col_idx).value = row[col_name]
            
            # Restore formulas for existing rows
            max_data_row = len(lineups_df) + 1
            for row in range(2, min(max_data_row + 1, sheet.max_row + 1)):
                if row in original_data:
                    for col in formula_columns:
                        if col in original_data[row] and original_data[row][col]['is_formula']:
                            formula = original_data[row][col]['formula_value']
                            sheet.cell(row=row, column=col).value = formula
            
            # For new rows beyond original data, copy formulas from the last template row
            if max_data_row > len(original_data):
                # Find the last row with formulas to use as template
                template_row = None
                for row in sorted(original_data.keys(), reverse=True):
                    if any(original_data[row][col]['is_formula'] for col in formula_columns if col in original_data[row]):
                        template_row = row
                        break
                
                if template_row:
                    # Copy formulas to new rows
                    for new_row in range(len(original_data) + 1, max_data_row + 1):
                        for col in formula_columns:
                            if col in original_data[template_row] and original_data[template_row][col]['is_formula']:
                                formula = original_data[template_row][col]['formula_value']
                                # Adjust formula references
                                adjusted_formula = adjust_formula_row_references(formula, template_row, new_row)
                                sheet.cell(row=new_row, column=col).value = adjusted_formula
            
            # Preserve formulas for rows beyond the data
            for row in range(max_data_row + 1, sheet.max_row + 1):
                if row in original_data:
                    for col in formula_columns:
                        if col in original_data[row] and original_data[row][col]['is_formula']:
                            formula = original_data[row][col]['formula_value']
                            sheet.cell(row=row, column=col).value = formula
        
        # Save the workbook
        wb.save(excel_path)
        logger.info(f"Successfully updated {sheet_name} with formula preservation")
        return True
        
    except Exception as e:
        logger.error(f"Error updating {sheet_name}: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False


# Add this to your main update routine or create a separate function
def run_lineup_update():
    """Run the starting lineup update"""
    logger.info("Starting lineup update process")
    
    try:
        # Update lineups for today
        today = datetime.now().strftime("%Y-%m-%d")
        success = update_starting_lineups_sheet(date_str=today)
        
        if success:
            logger.info("Successfully completed lineup update")
        else:
            logger.warning("Lineup update completed with issues")
            
        return success
        
    except Exception as e:
        logger.error(f"Error in lineup update process: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return False

def import_vba_module(excel_path, bas_file_path, module_name="Module1"):
    """
    Re-imports a VBA .bas file into a .xlsm workbook using xlwings.
    
    Parameters:
    excel_path (str): Path to the .xlsm workbook
    bas_file_path (str): Path to the .bas macro module file
    module_name (str): Name to use for the VBA module inside Excel
    """
    try:
        app = xw.App(visible=False)
        wb = app.books.open(excel_path)

        # Remove old module if it exists
        try:
            vb_module = wb.api.VBProject.VBComponents(module_name)
            wb.api.VBProject.VBComponents.Remove(vb_module)
        except Exception as e:
            pass  # Module might not exist yet

        # Import new module
        wb.api.VBProject.VBComponents.Import(bas_file_path)

        wb.save()
        wb.close()
        app.quit()
        print(f"Successfully imported {bas_file_path} into {excel_path}")

    except Exception as e:
        print(f"Error importing VBA module: {str(e)}")
        import traceback
        print(traceback.format_exc())

def run_update():
    """
    Main function to run the update process
    
    Returns:
    bool: True if successful, False otherwise
    """
    logger.info("Starting MLB projections update")
    global EXCEL_FILE_PATH, BACKUP_FOLDER
    
    # Create a backup first
    if not create_backup():
        logger.error("Update aborted due to backup failure")
        return False
    
    # Close any open Excel instances of this file
    try:
        import subprocess
        subprocess.run(['taskkill', '/f', '/im', 'excel.exe'], 
                      stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=False)
        logger.info("Closed any running Excel instances")
        time.sleep(2)  # Give time for Excel to close
    except Exception as e:
        logger.warning(f"Could not close Excel instances: {str(e)}")

    # Get barrel data from Baseball Savant with overwrite=True
    try:
        logger.info("Downloading barrel data from Baseball Savant (overwriting existing files)...")
        barrel_results = download_barrel_data_from_savant(overwrite=True)
        
        if barrel_results.get('processed', {}).get('success'):
            logger.info("Successfully downloaded and processed barrel data")
            
            # Get the file path
            barrel_file = barrel_results['processed'].get('file_path')
            
            # Update Statcast sheet with the new data
            if barrel_file and os.path.exists(barrel_file):
                statcast_update_result = update_statcast_sheet_in_excel(
                    excel_path=EXCEL_FILE_PATH,
                    barrel_data_path=barrel_file,
                    sheet_name="Statcast"
                )
                
                if statcast_update_result:
                    logger.info("Successfully updated Statcast sheet with barrel data")
                else:
                    logger.warning("Failed to update Statcast sheet")
            else:
                logger.warning("Processed barrel file not found or doesn't exist")
        else:
            # Try using the raw barrel_stats file
            if barrel_results.get('barrel_stats', {}).get('success'):
                logger.info("Using raw Statcast barrel data instead of processed file")
                barrel_file = barrel_results['barrel_stats'].get('file_path')
                
                if barrel_file and os.path.exists(barrel_file):
                    statcast_update_result = update_statcast_sheet_in_excel(
                        excel_path=EXCEL_FILE_PATH,
                        barrel_data_path=barrel_file,
                        sheet_name="Statcast"
                    )
                    
                    if statcast_update_result:
                        logger.info("Successfully updated Statcast sheet with raw barrel data")
                    else:
                        logger.warning("Failed to update Statcast sheet with raw data")
                else:
                    logger.warning("Raw barrel stats file not found or doesn't exist")
            else:
                logger.warning("Failed to download or process barrel data")
    except Exception as e:
        logger.error(f"Error handling barrel data: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())

    # Get park factors data
    try:
        logger.info("Updating park factors data...")
        update_park_factors()
    except Exception as e:
        logger.error(f"Error updating park factors: {str(e)}")
    
    # Get handedness splits data
    try:
        logger.info("Updating handedness splits data...")
        # Get reference data for name standardization
        reference_df = get_salaries_for_reference()
        
        # Get splits data from FanGraphs
        splits_data = get_handedness_splits()
        
        if splits_data:
            # Process the splits data
            processed_splits = process_splits_data(splits_data, reference_df)
            
            # Update Excel sheets
            update_result = update_splits_sheets(EXCEL_FILE_PATH, processed_splits)
            
            if update_result:
                logger.info("Successfully updated handedness splits data")
            else:
                logger.warning("Failed to update handedness splits data")
        else:
            logger.warning("Failed to retrieve handedness splits data")
    except Exception as e:
        logger.error(f"Error updating handedness splits: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
    
    # Get data using pybaseball and custom scrapers
    try:
        # Get salary data to use as reference for player name formats
        # Get reference data from salaries first
        reference_df = get_salaries_for_reference()

        # Process all data sources with the same standardization
        hitters_season = process_data(get_fangraphs_hitters, reference_df, "season")
        pitchers_season = process_data(get_fangraphs_pitchers, reference_df, "season")
        hitters_last3years = process_data(get_fangraphs_hitters, reference_df, "last3years")
        pitchers_last3years = process_data(get_fangraphs_pitchers, reference_df, "last3years")
        
        hitters_last7 = process_ranged_stats(days=7, player_type="hitters", reference_df=reference_df)
        pitchers_last30 = process_ranged_stats(days=30, player_type="pitchers", reference_df=reference_df)
        
        # Get team batting stats for offensive analysis
        try:
            logger.info("Getting team batting stats for offensive analysis...")
            current_year = datetime.now().year
            team_batting_stats = get_team_batting_stats(current_year)
            
            if not team_batting_stats.empty:
                logger.info(f"Successfully retrieved team batting stats for {len(team_batting_stats)} teams")
            else:
                logger.warning("Team batting stats retrieval returned empty DataFrame")
        except Exception as team_e:
            logger.error(f"Error getting team batting stats: {str(team_e)}")
            team_batting_stats = pd.DataFrame()
        
        # Get DraftKings salaries with encoding fixes
        dk_salaries = process_data(get_draftkings_salaries_csv)
        
        # Get probable pitchers with enhanced parsing for today's games
        probable_pitchers = get_fangraphs_probables()
        
        # Get player handedness data from local desktop files
        player_handedness = get_player_handedness()

        try:
            logger.info("Scraping today's starting lineups from MLB.com...")
            today = datetime.now().strftime("%Y-%m-%d")
            
            # Scrape the lineups
            lineups_df = scrape_mlb_starting_lineups(today)
            
            if not lineups_df.empty:
                success = update_starting_lineups_with_formulas(EXCEL_FILE_PATH, lineups_df)
                if success:
                    logger.info("Successfully updated TodaysStartingLineups sheet with lineup data while preserving formulas")
                else:
                    logger.warning("Failed to update TodaysStartingLineups sheet")
                lineups_df = pd.DataFrame()
                
        except Exception as lineup_e:
            logger.error(f"Error scraping starting lineups: {str(lineup_e)}")
            import traceback
            logger.error(traceback.format_exc())
            lineups_df = pd.DataFrame()
        
        if not player_handedness.empty and reference_df is not None:
            player_handedness = standardize_all_player_names(player_handedness, reference_df)
            logger.info("Successfully standardized handedness data with unified approach")
        
        # Add after retrieving the probable pitchers data:
        if not probable_pitchers.empty and reference_df is not None:
            probable_pitchers = standardize_all_player_names(probable_pitchers, reference_df, 'AwayPitcher')
            probable_pitchers = standardize_all_player_names(probable_pitchers, reference_df, 'HomePitcher')
            logger.info("Successfully standardized probable pitchers data with unified approach")
        
        # Calculate DraftKings points for each dataset
        if not hitters_season.empty:
            hitters_season = add_dk_points_to_dataframe(hitters_season, "hitters")
            hitters_season = standardize_all_player_names(hitters_season, reference_df)
            
        if not hitters_last7.empty:
            hitters_last7 = add_dk_points_to_dataframe(hitters_last7, "hitters")
            hitters_last7 = standardize_all_player_names(hitters_last7, reference_df)
            
        if not hitters_last3years.empty:
            hitters_last3years = add_dk_points_to_dataframe(hitters_last3years, "hitters")
            hitters_last3years = standardize_all_player_names(hitters_last3years, reference_df)
            
        if not pitchers_season.empty:
            pitchers_season = add_dk_points_to_dataframe(pitchers_season, "pitchers")
            pitchers_season = standardize_all_player_names(pitchers_season, reference_df)
            
        if not pitchers_last30.empty:
            pitchers_last30 = add_dk_points_to_dataframe(pitchers_last30, "pitchers")
            pitchers_last30 = standardize_all_player_names(pitchers_last30, reference_df)
            
        if not pitchers_last3years.empty:
            pitchers_last3years = add_dk_points_to_dataframe(pitchers_last3years, "pitchers")
            pitchers_last3years = standardize_all_player_names(pitchers_last3years, reference_df)
        
        # Get a fresh copy of the backup to work with
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        working_copy = os.path.join(BACKUP_FOLDER, f"MLBProjections_working_{timestamp}.xlsm")
        
        # Copy the latest backup to the working copy
        latest_backup = get_latest_backup()
        if latest_backup:
            shutil.copy2(latest_backup, working_copy)
            logger.info(f"Created working copy at {working_copy}")
        else:
            logger.error("No backup found to create working copy")
            return False
        
        # Set the working copy as the file to update
        original_file = EXCEL_FILE_PATH
        EXCEL_FILE_PATH = working_copy
        
        # Update Excel sheets with openpyxl
        if not hitters_season.empty:
            update_excel_data_only(hitters_season, "FGHitters")
        
        if not hitters_last7.empty:
            update_excel_data_only(hitters_last7, "FGHittersL7")
            
        if not hitters_last3years.empty:
            update_excel_data_only(hitters_last3years, "FGHittersL3Yrs")
        
        if not pitchers_season.empty:
            update_excel_data_only(pitchers_season, "FGPitchers")
            
        if not pitchers_last30.empty:
            update_excel_data_only(pitchers_last30, "FGPitchersL30")
            
        if not pitchers_last3years.empty:
            # Update with the old sheet name as a fallback
            update_excel_data_only(pitchers_last3years, "FGPitchersL3Yrs", old_sheet_name="FGPitchers2018")
        
        # Update team batting stats
        if not team_batting_stats.empty:
            logger.info("Updating team batting stats sheet...")
            update_team_hitting_sheet(EXCEL_FILE_PATH, team_batting_stats, "FGTmHitting")
            logger.info("Successfully updated team batting stats sheet")
        
        if not dk_salaries.empty:
            update_excel_data_only(dk_salaries, "Salaries")
        
        if not probable_pitchers.empty:
            update_probables_with_formulas(EXCEL_FILE_PATH, probable_pitchers, "probables")
        
        # Update Handedness sheet with player handedness data
        if not player_handedness.empty:
            update_excel_data_only(player_handedness, "Handedness")
            logger.info("Successfully updated Handedness sheet with player handedness data")

        if not lineups_df.empty:
            update_excel_data_only(lineups_df, "TodaysStartingLineups")
            logger.info("Successfully updated TodaysStartingLineups sheet with lineup data")
        
        # Also update the Statcast sheet in the working copy
        if barrel_results and barrel_results.get('processed', {}).get('success'):
            logger.info("Updating Statcast sheet in working copy...")
            barrel_file = barrel_results['processed'].get('file_path')
            if barrel_file and os.path.exists(barrel_file):
                statcast_result = update_statcast_sheet_in_excel(
                    excel_path=EXCEL_FILE_PATH,  # Now points to working copy
                    barrel_data_path=barrel_file,
                    sheet_name="Statcast"
                )
                if statcast_result:
                    logger.info("Successfully added Statcast sheet to working copy")
                else:
                    logger.warning("Failed to add Statcast sheet to working copy")
        
        # Update handedness splits in the working copy as well
        try:
            logger.info("Updating handedness splits in working copy...")
            if 'processed_splits' in locals() and processed_splits:
                update_result = update_splits_sheets(EXCEL_FILE_PATH, processed_splits)
                if update_result:
                    logger.info("Successfully updated handedness splits in working copy")
                else:
                    logger.warning("Failed to update handedness splits in working copy")
        except Exception as splits_e:
            logger.error(f"Error updating handedness splits in working copy: {str(splits_e)}")
    
        # Get park factors data
        try:
            logger.info("Updating park factors data...")
            park_factors_success = update_park_factors()
            if park_factors_success:
                logger.info("Park factors update completed successfully")
            else:
                logger.warning("Park factors update did not complete successfully")
        except Exception as e:
            logger.error(f"Error updating park factors: {str(e)}")

        # Always update player data from the existing Salaries sheet
        # This will preserve existing formulas and structure
        try:
            logger.info("Updating Hitter and Pitcher sheets from existing Salaries data (preserving formulas)")
            update_result = update_player_data()
            if update_result:
                logger.info("Successfully updated Hitter and Pitcher sheets while preserving formulas")
            else:
                logger.warning("Failed to update Hitter and Pitcher sheets from existing Salaries data")
        except Exception as update_e:
            logger.error(f"Error while updating player sheets: {str(update_e)}")
            import traceback
            logger.error(traceback.format_exc())
        
        # Identify starting pitchers for ML training
        try:
            logger.info("Filtering for starting pitchers before pitch weight calculation...")
            
            # Create a function to find starting pitchers across datasets
            def find_starting_pitchers():
                # Dictionary to collect starting pitchers
                starters = {}
                
                # Look for GS column in FanGraphs data
                for df, name in [
                    (pitchers_season, "FGPitchers"),
                    (pitchers_last30, "FGPitchersL30"),
                    (pitchers_last3years, "FGPitchersL3Yrs")
                ]:
                    if df is None or df.empty:
                        continue
                    
                    # Look for GS column
                    gs_col = None
                    for col in df.columns:
                        col_str = str(col).upper()
                        if col_str == 'GS' or 'GS' in col_str:
                            gs_col = col
                            break
                    
                    if gs_col:
                        logger.info(f"Found GS column '{gs_col}' in {name}")
                        # Convert to numeric and handle NaNs
                        df[gs_col] = pd.to_numeric(df[gs_col], errors='coerce').fillna(0)
                        
                        # Find starters (GS > 0)
                        starter_rows = df[df[gs_col] > 0]
                        
                        for _, row in starter_rows.iterrows():
                            if 'Name' in row and pd.notna(row['Name']):
                                starters[row['Name']] = True
                        
                        logger.info(f"Found {len(starter_rows)} starters in {name}")
                
                # Also check probable pitchers
                if not probable_pitchers.empty:
                    for col in ['AwayPitcher', 'HomePitcher']:
                        if col in probable_pitchers.columns:
                            for pitcher in probable_pitchers[col]:
                                if isinstance(pitcher, str) and pitcher != 'TBD':
                                    starters[pitcher] = True
                    
                    logger.info(f"Added probable pitchers as starters")
                
                # Create DataFrame from starters
                starter_names = list(starters.keys())
                if starter_names:
                    starter_df = pd.DataFrame({'Name': starter_names, 'IsStarter': True})
                    logger.info(f"Identified {len(starter_df)} unique starting pitchers")
                    return starter_df
                else:
                    logger.warning("No starting pitchers found")
                    return pd.DataFrame({'Name': [], 'IsStarter': []})
            
            # Get filtered DataFrame of starters
            filtered_pitcher_df = find_starting_pitchers()
            
            if not filtered_pitcher_df.empty:
                try:
                    # Read the Pitcher sheet to add starting flags
                    pitcher_df = pd.read_excel(EXCEL_FILE_PATH, sheet_name="Pitcher")
                    
                    # Create a 'Starting' flag column
                    pitcher_df['Starting'] = False
                    
                    # Convert to set for faster lookups
                    starter_names = set(filtered_pitcher_df['Name'].values)
                    
                    # Mark pitchers as starters
                    starter_count = 0
                    for idx, row in pitcher_df.iterrows():
                        if pd.notna(row.get('Name')) and row['Name'] in starter_names:
                            pitcher_df.at[idx, 'Starting'] = True
                            starter_count += 1
                    
                    logger.info(f"Marked {starter_count} pitchers as starters in the dataset")
                    
                    # Create a starter flag sheet
                    flag_df = pitcher_df[['Name', 'Starting']].copy()
                    update_excel_data_only(flag_df, "PitcherStartingFlag")
                    logger.info("Created PitcherStartingFlag sheet with starting pitcher indicators")
                except Exception as flag_e:
                    logger.error(f"Error creating starter flags: {str(flag_e)}")
                    # Continue without flag sheet
            else:
                logger.warning("No starting pitchers found, will use all pitchers for training")
            
        except Exception as filter_e:
            logger.error(f"Error filtering starting pitchers: {str(filter_e)}")
            import traceback
            logger.error(traceback.format_exc())
            filtered_pitcher_df = pd.DataFrame({'Name': [], 'IsStarter': []})
            
        # Calculate and add advanced pitch weight to Pitcher sheet - now using filtered pitchers
        try:
            logger.info("Calculating advanced pitch weight metric using starting pitchers for training...")
            add_pitch_weight_to_excel(EXCEL_FILE_PATH, sheet_name="Pitcher", filtered_pitcher_df=filtered_pitcher_df)
            logger.info("Successfully added pitch weight to Pitcher sheet")
        except Exception as pw_e:
            logger.error(f"Error calculating pitch weight: {str(pw_e)}")
            import traceback
            logger.error(traceback.format_exc())
        
        # Generate offensive matchup data using team batting stats
        try:
            if not team_batting_stats.empty:
                logger.info("Calculating team offensive metrics for matchup analysis...")
                
                # Ensure we have the necessary columns
                if 'TeamCode' in team_batting_stats.columns and 'OffenseScore' in team_batting_stats.columns:
                    # Convert to a lookup dictionary
                    offense_dict = dict(zip(team_batting_stats['TeamCode'], team_batting_stats['OffenseScore']))
                    
                    # Create DataFrame for Excel
                    offense_df = pd.DataFrame({
                        'TeamCode': list(offense_dict.keys()),
                        'OffenseScore': list(offense_dict.values())
                    })
                    
                    # Update offense sheet
                    update_excel_data_only(offense_df, "TeamOffense")
                    logger.info(f"Created TeamOffense sheet with {len(offense_df)} team offensive metrics")
        except Exception as offense_e:
            logger.error(f"Error calculating offensive metrics: {str(offense_e)}")
            import traceback
            logger.error(traceback.format_exc())

        calculate_adjusted_projections(EXCEL_FILE_PATH, "Hitter")
        
        try:
            logger.info("Synchronizing datasets for cross-sheet lookups...")
            synchronize_datasets_for_lookups()
            logger.info("Datasets synchronized for cross-sheet compatibility")
        except Exception as sync_e:
            logger.error(f"Error synchronizing datasets: {str(sync_e)}")

        try:
            shutil.copy2(working_copy, original_file)
            logger.info(f"Updated file copied to original location: {original_file}")
            EXCEL_FILE_PATH = original_file
        except Exception as e:
            logger.error(f"Error copying updated file to original location: {str(e)}")
            logger.info(f"Updated file is available at: {working_copy}")

        # Re-import macros
        try:
            import_vba_module(EXCEL_FILE_PATH, r"C:\Users\qbend\OneDrive\Desktop\VBMacros.bas", module_name="Module1")
            logger.info("Successfully re-imported VBA macros after update")
        except Exception as vba_e:
            logger.error(f"Error importing VBA macros: {str(vba_e)}")

    except Exception as e:
        logger.error(f"Error in update process: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        
    return True
    
def schedule_daily_update(hour=6, minute=0):
    """Schedule the update to run daily at the specified time"""
    schedule.every().day.at(f"{hour:02d}:{minute:02d}").do(run_update)
    logger.info(f"Scheduled daily update for {hour:02d}:{minute:02d}")
    
    while True:
        schedule.run_pending()
        time.sleep(60)  # Check every minute

def main():
    """Main entry point for the script"""
    import argparse
    
    parser = argparse.ArgumentParser(description='MLB Projections Data Updater')
    parser.add_argument('--run', action='store_true', help='Run the update immediately')
    parser.add_argument('--schedule', action='store_true', help='Schedule daily updates')
    parser.add_argument('--hour', type=int, default=6, help='Hour to run scheduled update (24-hour format)')
    parser.add_argument('--minute', type=int, default=0, help='Minute to run scheduled update')
    parser.add_argument('--barrel-only', action='store_true', help='Run only the barrel data update')
    
    args = parser.parse_args()
    
    if args.barrel_only:
        # Run only barrel data update
        update_barrel_data_only()
        return
    
    if args.run:
        # Run full update immediately
        run_update()
    
    if args.schedule:
        # Schedule daily updates
        schedule_daily_update(args.hour, args.minute)
    
    if not args.run and not args.schedule and not args.barrel_only:
        # If no arguments provided, run once and then schedule
        run_update()
        schedule_daily_update()

if __name__ == "__main__":
    main()